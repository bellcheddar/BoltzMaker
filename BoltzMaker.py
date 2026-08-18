#!/usr/bin/env python3
"""BoltzMaker.py -- generate, preflight, run, and analyze Boltz-2 batch campaigns.

Usage:
    python3 BoltzMaker.py setup                          # create managed venv + install boltz
    python3 BoltzMaker.py setup-plip                      # optional: separate env for cif2plip
    python3 BoltzMaker.py new       [boltz_input.md]      # interactively write a new campaign
    python3 BoltzMaker.py generate  <boltz_input.md> ...
    python3 BoltzMaker.py preflight <boltz_input.md> ...
    python3 BoltzMaker.py run       <boltz_input.md> ...
    python3 BoltzMaker.py analyze   <boltz_input.md> ...
    python3 BoltzMaker.py all       <boltz_input.md> ...   # (also the default if no subcommand given)
"""

import os
import sys
import subprocess
import shutil
from pathlib import Path

SCRIPT_PATH = Path(__file__).resolve()
SCRIPT_DIR = SCRIPT_PATH.parent
VENV_DIR = SCRIPT_DIR / ".venv"

# Separate, optional environment for cif2plip (protein-ligand interaction analysis).
# Kept apart from VENV_DIR because its dependency chain (OpenBabel, PyMOL) needs
# conda-forge builds -- confirmed empirically: plip's own installer forces a from-source
# OpenBabel-binding build unless a working OpenBabel is already importable at build time,
# and the standalone PyPI `pymol-open-source` wheel has a hardcoded broken rpath to its
# original builder's machine. conda-forge's builds have neither problem.
PLIP_VENV_DIR = SCRIPT_DIR / ".plip_env"
CIF2PLIP_COMMIT = "2c3bf8b086ec022d81599b77a91b4713697a5636"

# Vendored (not CDN-linked) so the dashboard's charts render in contexts that don't
# execute a cross-origin <script src>, e.g. htmlpreview.github.io -- confirmed empirically
# that a CDN-loaded plotly.js silently fails to run there even though the exact same file
# works when opened directly in a browser, leaving every chart card blank.
PLOTLY_JS_PATH = SCRIPT_DIR / "vendor" / "plotly-2.35.2.min.js"

# Same rationale as PLOTLY_JS_PATH -- vendored so the interactive binding-site view has
# no external script dependency at all.
THREEDMOL_JS_PATH = SCRIPT_DIR / "vendor" / "3Dmol-2.5.5-min.js"

# cif2plip's own PLIP visualization never calls cmd.label(...) at all (checked its
# source directly) -- the stock PNG/pse have sticks and dashed interaction lines but no
# residue text. This small script (BoltzMaker's own, not vendored from upstream) loads
# the .pse PLIP already produced -- same camera/view/representations -- adds labels for
# the given contacting residues, and re-renders. Runs inside .plip_env (needs `import
# pymol`), invoked as a subprocess from the main venv.
_LABEL_RESIDUES_SCRIPT = '''\
import sys
import pymol
pymol.finish_launching(["pymol", "-qc"])
from pymol import cmd

session_path, output_path = sys.argv[1], sys.argv[2]
residues = [r.split(":") for r in sys.argv[3:]]

cmd.load(session_path)

# Residue (CA) labels: offset away from the atom (screen-space, so it holds
# regardless of viewing angle) so the text doesn't sit on top of the residue
# itself or its stick representation.
cmd.set("label_size", 18)
cmd.set("label_color", "black")
cmd.set("label_outline_color", "white")
cmd.set("label_font_id", 7)  # sans-serif bold -- more legible than PyMOL's default serif
cmd.set("label_position", (2.2, 2.2, 0))

for chain, resnr, restype in residues:
    sel = f"chain {chain} and resi {resnr} and name CA"
    if cmd.count_atoms(sel) > 0:
        cmd.label(sel, f'"{restype}{resnr}"')

# Interaction-line distance labels: PLIP draws each interaction via
# cmd.distance() (which computes and can show the actual measured distance)
# but hides the label by default. Show them, styled distinctly (smaller,
# grey) from the residue labels so the two don't visually compete.
cmd.set("label_size", 14, "Interactions")
cmd.set("label_color", "gray30", "Interactions")
cmd.show("labels", "Interactions")

cmd.set("ray_opaque_background", 0)
cmd.png(output_path, width=1200, height=900, dpi=150, ray=1)
'''


# ==========================================================================
# CLI styling -- stdlib-only (no pyfiglet/rich dependency) so the banner and
# colored prompts work even in `setup`/`setup-plip`, which run on whatever
# bare python3 the user has, before the managed venv (and rich) exists.
# Respects NO_COLOR (https://no-color.org) and auto-disables on a non-tty
# (piped output, log redirection, CI) so scripted use stays clean.
# ==========================================================================

_NO_COLOR = bool(os.environ.get("NO_COLOR")) or not sys.stdout.isatty()


def _c(code: str) -> str:
    return "" if _NO_COLOR else code


_RESET = _c("\x1b[0m")
_BOLD = _c("\x1b[1m")
_DIM = _c("\x1b[2m")
# marcdeller.com brand palette (matches the HTML dashboard's :root custom properties).
_BLUE = _c("\x1b[38;2;30;115;190m")    # #1e73be -- primary
_CYAN = _c("\x1b[38;2;74;159;212m")    # #4a9fd4 -- primary-light
_AMBER = _c("\x1b[38;2;252;185;0m")    # #fcb900 -- accent
_GREEN = _c("\x1b[38;2;0;208;132m")    # #00d084 -- accent
_RED = _c("\x1b[38;2;214;39;40m")

_BANNER_ART = r"""   ___       ____      __  ___     __
  / _ )___  / / /____ /  |/  /__ _/ /_____ ____
 / _  / _ \/ / __/_ // /|_/ / _ `/  '_/ -_) __/
/____/\___/_/\__//__/_/  /_/\_,_/_/\_\\__/_/
"""


def _print_banner() -> None:
    if _NO_COLOR:
        print("BoltzMaker")
        return
    print()
    for line in _BANNER_ART.splitlines():
        print(f"{_BOLD}{_BLUE}{line}{_RESET}")
    print(f"  {_DIM}Boltz-2 batch campaigns -- structure + affinity{_RESET}")
    print()


def _ok(msg: str) -> None:
    print(f"{_GREEN}✓{_RESET} {msg}")


def _info(msg: str) -> None:
    print(f"{_BLUE}ℹ{_RESET} {msg}")


def _warn(msg: str) -> None:
    print(f"{_AMBER}⚠{_RESET} {msg}")


def _err(msg: str) -> None:
    print(f"{_RED}✗{_RESET} {msg}")


def _step(msg: str) -> None:
    print(f"{_BLUE}→{_RESET} {msg}")


def _venv_bin(name: str) -> Path:
    return VENV_DIR / ("Scripts" if os.name == "nt" else "bin") / name


def _boltz_bin() -> Path:
    # Under pixi, boltz's console-script entry point lands on PATH inside the pixi
    # environment directly (installed via [pypi-dependencies] in pixi.toml) --
    # resolved via shutil.which rather than the .venv-specific path below.
    if _in_pixi_env():
        found = shutil.which("boltz")
        return Path(found) if found else _venv_bin("boltz")
    return _venv_bin("boltz")


def _plip_venv_bin(name: str) -> Path:
    return PLIP_VENV_DIR / "env" / ("Scripts" if os.name == "nt" else "bin") / name


def _plip_python() -> Path:
    # Under pixi, plip/openbabel/pymol/gemmi all live in the one unified environment
    # BoltzMaker.py is already running under (installed via pixi.toml + the
    # `postinstall` task) -- no separate .plip_env/env/bin/python to point at.
    if _in_pixi_env():
        return Path(sys.executable)
    return _plip_venv_bin("python")


def _plip_script() -> Path:
    return PLIP_VENV_DIR / "cif2plip" / "cif2plip.py"


def _plip_label_script() -> Path:
    return PLIP_VENV_DIR / "label_residues.py"


def _ensure_plip_assets_vendored() -> None:
    """Fetch cif2plip.py and (re)write label_residues.py, idempotently.

    Shared by cmd_setup_plip()'s full micromamba-env build and, under pixi, by
    _plip_available()'s lazy self-heal -- `pixi run postinstall` only installs the
    plip/pdb-tools Python packages (see pixi.toml), it doesn't vendor these two
    BoltzMaker-side asset files, so the first real check for plip availability under
    pixi is what actually fetches them.
    """
    cif2plip_dir = PLIP_VENV_DIR / "cif2plip"
    cif2plip_dir.mkdir(parents=True, exist_ok=True)
    script_path = cif2plip_dir / "cif2plip.py"
    if not script_path.exists():
        url = f"https://raw.githubusercontent.com/bellcheddar/cif2plip/{CIF2PLIP_COMMIT}/cif2plip.py"
        _step(f"vendoring cif2plip.py (pinned commit {CIF2PLIP_COMMIT[:10]})")
        _curl_download(url, script_path)
    # Always (re)written, even when reusing an existing env, so an env built before this
    # script existed picks up residue labeling without needing a full rebuild.
    _plip_label_script().write_text(_LABEL_RESIDUES_SCRIPT)


def _plip_available() -> bool:
    if _in_pixi_env():
        try:
            import importlib
            importlib.import_module("plip")
        except ImportError:
            return False
        if not _plip_script().exists():
            _ensure_plip_assets_vendored()
        return _plip_script().exists()
    return _plip_python().exists() and _plip_script().exists()


# ==========================================================================
# `setup` -- stdlib-only, must work under whatever bare python3 the user has
# ==========================================================================

def _find_boltz_python() -> Path:
    # boltz pins numpy<2.0, and numpy 1.26.x has no prebuilt wheel for Python 3.13+
    # (only cp312 and earlier) -- building it from source fails against recent Xcode
    # Clang, so the managed venv must be built on 3.12, not whatever is newest.
    candidates = ["/opt/homebrew/bin/python3.12", "/usr/local/bin/python3.12", "python3.12"]
    for c in candidates:
        path = c if os.path.isabs(c) else shutil.which(c)
        if not path or not os.path.exists(path):
            continue
        try:
            out = subprocess.run([path, "--version"], capture_output=True, text=True)
            if "3.12" in (out.stdout or out.stderr):
                return Path(path)
        except Exception:
            continue
    _err("could not find a python3.12 interpreter (checked /opt/homebrew/bin, /usr/local/bin, PATH).")
    _info("Install one (e.g. `brew install python@3.12`) and re-run `setup`.")
    sys.exit(1)


# boltz's triangular attention computes the row-wise QK^T score matrix for the whole
# complex in one unchunked matmul (batch = token count, so the tensor scales as
# tokens * heads * tokens * tokens). Apple's MPS backend has a hard ceiling on single-
# tensor size; past roughly 1250 residues that matmul crashes the whole process with a
# SIGSEGV inside PyTorch's internal tiled-bmm fallback (confirmed via crash symbolication
# and direct tracing against boltz 2.2.1 on Apple Silicon). Each row's attention is
# independent, so chunking along that axis is exact, not an approximation -- this patches
# the installed boltz package directly (there is no upstream release with this fix yet).
_BOLTZ_MPS_ATTN_ORIGINAL = '''def _attention(
    query: torch.Tensor,
    key: torch.Tensor,
    value: torch.Tensor,
    biases: List[torch.Tensor],
) -> torch.Tensor:
    # [*, H, C_hidden, K]
    key = permute_final_dims(key, (1, 0))

    # [*, H, Q, K]
    a = torch.matmul(query, key)

    for b in biases:
        a += b

    a = softmax_no_cast(a, -1)

    # [*, H, Q, C_hidden]
    a = torch.matmul(a, value)

    return a'''

_BOLTZ_MPS_ATTN_PATCHED = '''# Patched by BoltzMaker setup -- see cmd_setup()'s _patch_boltz_mps_attention().
_MPS_ATTN_MAX_SCORE_ELEMENTS = 400_000_000  # ~1.6GB fp32, well under MPS's tensor-size ceiling


def _attention(
    query: torch.Tensor,
    key: torch.Tensor,
    value: torch.Tensor,
    biases: List[torch.Tensor],
) -> torch.Tensor:
    # [*, H, C_hidden, K]
    key = permute_final_dims(key, (1, 0))

    n_rows = query.shape[1] if query.dim() >= 5 else None
    score_elements = 1
    for d in query.shape[:-1]:
        score_elements *= d
    score_elements *= key.shape[-1]

    if (
        query.device.type == "mps"
        and n_rows is not None
        and n_rows > 1
        and score_elements > _MPS_ATTN_MAX_SCORE_ELEMENTS
    ):
        n_chunks = math.ceil(score_elements / _MPS_ATTN_MAX_SCORE_ELEMENTS)
        chunk_size = math.ceil(n_rows / n_chunks)
        outputs = []
        for start in range(0, n_rows, chunk_size):
            end = min(start + chunk_size, n_rows)
            a_c = torch.matmul(query[:, start:end], key[:, start:end])
            for b in biases:
                b_c = b[:, start:end] if b.shape[1] == n_rows else b
                a_c = a_c + b_c
            a_c = softmax_no_cast(a_c, -1)
            outputs.append(torch.matmul(a_c, value[:, start:end]))
        return torch.cat(outputs, dim=1)

    # [*, H, Q, K]
    a = torch.matmul(query, key)

    for b in biases:
        a += b

    a = softmax_no_cast(a, -1)

    # [*, H, Q, C_hidden]
    a = torch.matmul(a, value)

    return a'''


def _patch_boltz_mps_attention(venv_dir: Path) -> None:
    """Patch boltz's triangular attention to chunk on MPS for large complexes.

    Idempotent and non-fatal: does nothing if already patched, and warns (without
    aborting setup) if boltz's source no longer matches what this patch expects, since
    that means an upstream boltz release changed this function and the patch needs
    re-checking against the new code before it can be safely reapplied.
    """
    import glob

    matches = glob.glob(str(venv_dir / "lib" / "python3.*" / "site-packages" / "boltz" /
                             "model" / "layers" / "triangular_attention" / "primitives.py"))
    if not matches:
        _warn("could not find boltz's triangular_attention/primitives.py "
              "to patch (MPS large-complex fix not applied).")
        return
    target = Path(matches[0])
    text = target.read_text()

    if "_MPS_ATTN_MAX_SCORE_ELEMENTS" in text:
        _ok("MPS triangular-attention chunking patch already applied.")
        return
    if _BOLTZ_MPS_ATTN_ORIGINAL not in text:
        _warn("boltz's _attention() source doesn't match the expected "
              "vanilla text (boltz version changed?) -- skipping MPS large-complex patch. "
              "Large multi-chain complexes (>~1250 residues) may crash on Apple Silicon MPS.")
        return

    target.write_text(text.replace(_BOLTZ_MPS_ATTN_ORIGINAL, _BOLTZ_MPS_ATTN_PATCHED))
    _ok("patched boltz's triangular attention to chunk on MPS for large complexes.")


def cmd_setup(argv: list) -> None:
    force = "--force" in argv
    yes = "--yes" in argv or "-y" in argv

    if VENV_DIR.exists() and force:
        shutil.rmtree(VENV_DIR)
    if not VENV_DIR.exists():
        interpreter = _find_boltz_python()
        _step(f"creating venv at {VENV_DIR} using {interpreter}")
        subprocess.run([str(interpreter), "-m", "venv", str(VENV_DIR)], check=True)
    else:
        _info(f"reusing existing venv at {VENV_DIR} (pass --force to recreate)")

    pip = _venv_bin("pip")
    subprocess.run([str(pip), "install", "--upgrade", "pip"], check=True)

    _step("about to install boltz + dependencies into the managed venv.")
    print(f"  {_DIM}This pulls PyTorch (~2-3 GB) and, on the first `boltz predict` run,{_RESET}")
    print(f"  {_DIM}Boltz will download several GB of model weights over the network.{_RESET}")
    if not yes:
        resp = input(f"{_BLUE}?{_RESET} Continue? {_DIM}(y/N){_RESET} ").strip().lower()
        if resp != "y":
            _err("Aborted.")
            sys.exit(1)

    subprocess.run(
        [str(pip), "install", "boltz", "rich", "pandas", "openpyxl", "pyyaml", "rdkit", "matplotlib", "psutil",
         "scipy", "gemmi", "biopython", "plotly", "reportlab"],
        check=True,
    )
    freeze = subprocess.run([str(pip), "freeze"], capture_output=True, text=True, check=True)
    (VENV_DIR / "requirements.lock.txt").write_text(freeze.stdout)

    boltz_check = subprocess.run([str(_venv_bin("boltz")), "--help"], capture_output=True, text=True)
    (_ok if boltz_check.returncode == 0 else _warn)(f"boltz CLI check exit={boltz_check.returncode}")

    _patch_boltz_mps_attention(VENV_DIR)

    py = _venv_bin("python3")
    torch_check = subprocess.run(
        [str(py), "-c", "import torch; print('mps:', torch.backends.mps.is_available()); print('cuda:', torch.cuda.is_available())"],
        capture_output=True, text=True,
    )
    if torch_check.stdout.strip():
        _info(torch_check.stdout.strip())
    else:
        _warn(f"torch import check failed: {torch_check.stderr[-500:]}")
    _ok("setup complete. Run `python3 BoltzMaker.py all <boltz_input.md>` next.")


# ==========================================================================
# `setup-plip` -- separate, optional environment for cif2plip. stdlib-only, same as
# `setup` above; also runs before the bootstrap-relaunch so it works on a fresh
# checkout regardless of whether `setup` has ever been run.
# ==========================================================================

def _curl_download(url: str, dest: Path) -> None:
    # Shell out to curl (system trust store) rather than urllib.request -- some Python
    # installations (e.g. python.org framework builds) ship without properly configured
    # CA certificates, which breaks urllib's HTTPS entirely; curl doesn't have this problem.
    subprocess.run(["curl", "-fLsS", "-o", str(dest), url], check=True)


def _download_micromamba(dest: Path) -> None:
    import platform
    import tarfile
    import tempfile

    arch = platform.machine()
    system = platform.system()
    if system == "Darwin":
        plat = {"arm64": "osx-arm64", "x86_64": "osx-64"}.get(arch)
    elif system == "Linux":
        plat = {"x86_64": "linux-64", "aarch64": "linux-aarch64"}.get(arch)
    else:
        plat = None
    if plat is None:
        _err(f"no known micromamba build for platform {system}/{arch!r}.")
        _info(f"Install micromamba yourself (https://micro.mamba.pm) and place the binary at {dest}")
        sys.exit(1)
    url = f"https://micro.mamba.pm/api/micromamba/{plat}/latest"
    _step(f"downloading micromamba from {url}")
    dest.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory() as tmp:
        tar_path = Path(tmp) / "micromamba.tar.bz2"
        _curl_download(url, tar_path)
        with tarfile.open(tar_path, "r:bz2") as tf:
            tf.extractall(tmp)
        shutil.copy2(Path(tmp) / "bin" / "micromamba", dest)
    dest.chmod(dest.stat().st_mode | 0o111)


def cmd_setup_plip(argv: list) -> None:
    force = "--force" in argv
    yes = "--yes" in argv or "-y" in argv

    if PLIP_VENV_DIR.exists() and force:
        shutil.rmtree(PLIP_VENV_DIR)

    micromamba = PLIP_VENV_DIR / "bin" / "micromamba"
    if not micromamba.exists():
        _download_micromamba(micromamba)

    env_dir = PLIP_VENV_DIR / "env"
    if not env_dir.exists():
        _step("about to build a separate environment for cif2plip (protein-ligand")
        print(f"  {_DIM}interaction analysis) via conda-forge (python + gemmi + openbabel + pymol-open-source).{_RESET}")
        print(f"  {_DIM}This is roughly 1-1.5GB (mostly Qt/PyMOL's own dependencies), separate from{_RESET}")
        print(f"  {_DIM}BoltzMaker's own pip-only venv, and entirely optional -- BoltzMaker works fully{_RESET}")
        print(f"  {_DIM}without it.{_RESET}")
        if not yes:
            resp = input(f"{_BLUE}?{_RESET} Continue? {_DIM}(y/N){_RESET} ").strip().lower()
            if resp != "y":
                _err("Aborted.")
                sys.exit(1)
        _step(f"creating {env_dir} via micromamba")
        subprocess.run(
            [str(micromamba), "create", "-y", "-p", str(env_dir), "-c", "conda-forge",
             "python=3.11", "gemmi", "openbabel", "pymol-open-source"],
            check=True,
        )
    else:
        _info(f"reusing existing plip env at {env_dir} (pass --force to recreate)")

    pip = env_dir / "bin" / "pip"
    subprocess.run([str(pip), "install", "pdb-tools"], check=True)
    # --no-build-isolation: plip's own installer tries to build OpenBabel's Python
    # bindings from source unless `import openbabel` already succeeds where it runs --
    # isolated build sandboxes can't see this env's already-installed openbabel, so
    # isolation must be off for that check to find it and skip the (broken) rebuild.
    subprocess.run([str(pip), "install", "--no-build-isolation", "plip"], check=True)

    (PLIP_VENV_DIR / "cif2plip" / "cif2plip.py").unlink(missing_ok=True)  # force a re-fetch below
    _ensure_plip_assets_vendored()

    _step("smoke-testing the plip environment...")
    smoke = subprocess.run(
        [str(env_dir / "bin" / "python"), "-c", "import pymol, openbabel, gemmi, plip; print('SMOKE_OK')"],
        capture_output=True, text=True,
    )
    if smoke.returncode != 0 or "SMOKE_OK" not in smoke.stdout:
        _err("plip environment smoke test failed:")
        print(smoke.stderr[-2000:])
        sys.exit(1)
    _ok("setup-plip complete.")
    print(f"  {_DIM}`analyze` will now run interaction analysis automatically, and `new` can suggest{_RESET}")
    print(f"  {_DIM}pocket residues from a reference structure.{_RESET}")


# ==========================================================================
# Bootstrap shim -- relaunch under the managed venv's interpreter so every
# command below this point can assume rich/pandas/yaml/rdkit are importable.
# ==========================================================================

def _in_pixi_env() -> bool:
    # True both for `pixi run`/`pixi shell` (this checkout's pixi.toml, Tier A) and for
    # a Tier B offline install.sh/pixi-pack executable's `source .../activate.sh` --
    # the latter never runs `pixi` at all, so PIXI_PROJECT_ROOT (which only the pixi
    # CLI itself sets) can't be the signal. CONDA_PREFIX is exported by both: pixi
    # itself sets it for `run`/`shell`, and pixi-pack's generated activate.sh sets it
    # unconditionally (confirmed by reading a real extracted activate.sh). Checking it
    # matches sys.prefix (not just "is set to something") confirms this process is
    # actually running from that environment, not just inheriting a stale var from an
    # unrelated parent shell.
    conda_prefix = os.environ.get("CONDA_PREFIX")
    return bool(conda_prefix) and Path(conda_prefix).resolve() == Path(sys.prefix).resolve()


def _bootstrap_or_relaunch(argv: list) -> None:
    subcommand = argv[1] if len(argv) > 1 else None
    # `format` and `preflight --json` are both meant for scripting/tooling consumption
    # (this web app's own subprocess wrapper included) -- the banner text on stdout
    # would corrupt a caller trying to parse JSON straight off stdout, so both stay quiet.
    quiet = subcommand == "format" or (subcommand == "preflight" and "--json" in argv)
    if subcommand in ("setup", "setup-plip"):
        if _in_pixi_env():
            _print_banner()
            _err(f"'{subcommand}' builds BoltzMaker's own .venv/.plip_env -- redundant "
                 "and not needed inside a pixi environment.")
            _info("Dependencies here are already managed by pixi.toml. Run `pixi install`, "
                  "then `pixi run postinstall` once for plip/pdb-tools.")
            sys.exit(1)
        _print_banner()
        (cmd_setup if subcommand == "setup" else cmd_setup_plip)(argv[2:])
        sys.exit(0)

    if _in_pixi_env():
        # Already running under pixi's own interpreter with every dependency this
        # script needs (rich/pandas/yaml/rdkit/...) installed by pixi.toml -- skip the
        # .venv existence check/execv relaunch below entirely, straight through to the
        # rest of the script.
        if not quiet:
            _print_banner()
        return

    venv_python = _venv_bin("python3")
    if not venv_python.exists():
        _err("no managed environment found.")
        _info(f"Run: python3 {SCRIPT_PATH} setup")
        sys.exit(1)
    if Path(sys.executable).resolve() != venv_python.resolve():
        os.execv(str(venv_python), [str(venv_python), str(SCRIPT_PATH)] + argv[1:])
    # else: already running under venv python -- fall through to the rest of the script.
    # Printed here (not at the top of this function) so it appears exactly once per
    # invocation, after the execv relaunch above, not once in each process image.
    if not quiet:
        _print_banner()


_bootstrap_or_relaunch(sys.argv)

# --------------------------------------------------------------------------
# Everything below only ever executes inside the managed venv.
# --------------------------------------------------------------------------

import argparse
import ast
import base64
import contextlib
import html
import io
import json
import re
import threading
import time
from dataclasses import dataclass, field, asdict

import yaml
import pandas as pd
import psutil
from rich import box as _rich_box
from rich.console import Console
from rich.table import Table
from rich.progress import Progress, BarColumn, TextColumn
from rich.table import Column

# marcdeller.com brand palette, for rich components (Table/Progress) that render
# post-bootstrap -- matches the plain-ANSI _BLUE/_AMBER/_GREEN used pre-bootstrap above.
_RICH_BLUE = "#1e73be"
_RICH_AMBER = "#fcb900"
_RICH_GREEN = "#00d084"
# The brand's alert colour, matching --md-accent-red in the dashboard CSS. The
# preflight table used a bare "red", which is the terminal's own red and does not
# match anything else BoltzMaker draws.
_RICH_RED = "#d81b8c"
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

import plotly.graph_objects as go
import plotly.io as pio


# ==========================================================================
# Data model
# ==========================================================================

@dataclass
class Settings:
    output_dir: str = "./boltz_yamls"
    predict_affinity: bool = False
    # How far a ligand may sit from the named pocket contacts. Boltz's own default
    # is 6.0; only emitted when a family actually has pocket contacts, so a campaign
    # without them is unaffected.
    pocket_distance: float = 6.0
    # How many targets one `boltz predict` process may run before BoltzMaker starts a
    # fresh one. Apple's MPS allocator never returns everything: measured on a live
    # campaign, driver-held memory floored at ~20GB after the first target and rose
    # ~1.9GB per target after that, while a single target in a fresh process peaked at
    # 47.6GB of a 55.7GB ceiling. Left to run, the two meet and every later target
    # fails for memory -- 0 out-of-memory skips in the first four targets of one run,
    # 3 in the last four. Only process exit frees it, so the process is recycled.
    # 0 disables the recycling and restores the old single-invocation behaviour.
    targets_per_invocation: int = 4


@dataclass
class Partner:
    id: str
    sequence: str
    type: str = "protein"  # protein / dna / rna
    modifications: object = None
    cyclic: bool = False
    msa: object = None


@dataclass
class ProteinFamily:
    id: str
    sequence: str
    partners: list = field(default_factory=list)
    pocket_contacts: object = None
    # {code: [contact tokens]}. A protein may define several pockets; every ligand is
    # then run against each, plus once unconstrained. Distinct from pocket_contacts,
    # which is the older unnamed form and still means exactly one constrained target.
    pockets: object = None
    ligands: object = None
    modifications: object = None
    cyclic: bool = False
    msa: object = None
    bond_constraints: object = None
    contact_constraints: object = None
    templates: object = None
    apo_structure: object = None   # raw path string to a reference apo structure, or None
    apo_chain: object = None        # explicit apo chain id, or None (triggers auto-detect)
    family_type: str = "auto"        # "gpcr" | "kinase" | "auto" -- selects the compare-sse MotifAnnotator
    group: object = None             # optional display/report grouping name shared across
                                       # multiple Protein: blocks of the same underlying
                                       # receptor (e.g. with/without a partner, or an apo
                                       # variant) -- defaults to this family's own id if unset


@dataclass
class Ligand:
    id: str
    smiles: object = None
    ccd: object = None
    role: object = None  # optional "agonist" / "antagonist" -- purely for reporting
                           # (dashboard charts, compare-sse), never affects generate/run
    # Pocket contacts scoped to THIS ligand, overriding the protein's. A pocket is
    # not purely a property of the receptor: measured on GLP1R/GIPR, the site where
    # orforglipron binds GLP1R (7E14) and the site where LSN1 binds GIPR (7RBT) share
    # 3 residues out of ~60 once projected onto each other. One pocket per protein
    # would force one of those chemotypes into the wrong site.
    pocket_contacts: object = None


@dataclass
class Campaign:
    settings: Settings
    partners: dict
    families: list
    ligands: list
    source_path: object = None


@dataclass
class Target:
    stem: str
    family_id: str
    ligand_id: object  # str, or None for a ligand-free (apo) target
    pocket_contacts_used: object = None
    needs_affinity: bool = True  # False for apo targets even when the campaign predicts affinity


MANIFEST_FILENAME = ".boltzmaker_manifest.json"
RUN_HISTORY_FILENAME = ".boltzmaker_run_history.jsonl"


class MDParseError(Exception):
    pass


# ==========================================================================
# MDParser -- boltz_input.md is plain labelled text: blank-line-separated
# blocks (`Settings:` / `Protein: <name>` / `Partner: <name>` / `Ligand:
# <name>`, each followed by `Field: value` lines), plus standalone
# constraint sentences ("Covalent bond: X residue N atom A to Y residue M
# atom B") recognized anywhere in the file. No markdown, no YAML, no
# brackets or quoting -- one rule (`Label: value`, blank line between
# records) so the file reads like a filled-in form, not code.
# ==========================================================================

_RECORD_START_RE = re.compile(r"^(Settings|Protein|Partner|Ligand)\s*:\s*(.*)$", re.IGNORECASE)
_FIELD_RE = re.compile(r"^([A-Za-z][A-Za-z ]*?)\s*:\s*(.*)$")

_RECORD_ALLOWED_FIELDS = {
    "settings": {"output folder", "predict affinity", "pocket distance",
                 "targets per invocation"},
    "protein": {"sequence", "partners", "ligands", "modifications", "cyclic", "msa", "templates",
                "apo structure", "apo chain", "family type", "group"},
    "partner": {"sequence", "type", "copies", "modifications", "cyclic", "msa"},
    "ligand": {"smiles", "ccd", "role"},
}

# A statement's owner is always its first-mentioned chain, which must be a
# Protein -- this is how a standalone constraint sentence attaches to the
# family it belongs to.
_ENDPOINT = r"(\w+)\s+residue\s+(\d+)(?:\s+atom\s+(\w+))?"
_COVALENT_RE = re.compile(
    rf"^covalent bond:\s*(\w+)\s+residue\s+(\d+)\s+atom\s+(\w+)\s+to\s+(\w+)\s+residue\s+(\d+)\s+atom\s+(\w+)\s*$",
    re.IGNORECASE)
# The optional "for <ligand>" scopes a pocket to one ligand instead of to every
# ligand of the protein. Needed because a pocket is not purely a property of the
# receptor: measured on GLP1R/GIPR, orforglipron's site on GLP1R and LSN1's site
# on GIPR share 3 residues out of ~60 once projected onto each other.
# "as CODE" names the pocket, so one protein can carry several and every ligand is
# run against each of them plus an unconstrained baseline -- the matrix that answers
# "where does this compound actually want to sit". "for LIGAND" instead scopes a
# single pocket to one ligand.
_POCKET_RE = re.compile(
    rf"^pocket contact:\s*{_ENDPOINT}"
    rf"(?:\s+for\s+(?P<lig>\S+))?(?:\s+as\s+(?P<code>[A-Za-z0-9_-]+))?\s*$",
    re.IGNORECASE)
_DISTANCE_RE = re.compile(
    rf"^distance constraint:\s*{_ENDPOINT}\s+to\s+{_ENDPOINT}(?:\s+within\s+([\d.]+)(?:\s+\w+)?)?\s*$",
    re.IGNORECASE)


def _find_comment_start(line: str):
    # '#' starts a comment only at col 0 or after whitespace, so a SMILES
    # triple bond like 'C#N' is never mistaken for one (no quoting exists in
    # this format, so that's the only rule needed).
    for i, ch in enumerate(line):
        if ch == "#" and (i == 0 or line[i - 1].isspace()):
            return i
    return None


def _strip_comment(line: str) -> str:
    idx = _find_comment_start(line)
    return line if idx is None else line[:idx]


def _strip_quotes(s: str) -> str:
    if len(s) >= 2 and s[0] == s[-1] and s[0] in ("'", '"'):
        return s[1:-1]
    return s


def _parse_yesno(value: str, default: bool) -> bool:
    v = value.strip().lower()
    if v in ("yes", "y", "true", "1"):
        return True
    if v in ("no", "n", "false", "0"):
        return False
    return default


def _parse_csv(value: str) -> list:
    return [_strip_quotes(v.strip()) for v in value.split(",") if v.strip()]


def _parse_modification_token(s: str) -> list:
    parts = [p.strip() for p in s.split(":")]
    if len(parts) != 2:
        raise MDParseError(f"invalid modification (expected CCD:position): {s!r}")
    return [parts[0], int(parts[1])]


def _match_statement(stripped: str, lineno: int):
    m = _COVALENT_RE.match(stripped)
    if m:
        c1, r1, a1, c2, r2, a2 = m.groups()
        return {"type": "bond", "owner": c1, "atom1": [c1, int(r1), a1], "atom2": [c2, int(r2), a2], "line": lineno}
    m = _POCKET_RE.match(stripped)
    if m:
        chain, res, atom = m.group(1), m.group(2), m.group(3)
        token = [chain, int(res), atom] if atom else [chain, int(res)]
        return {"type": "pocket", "owner": chain, "token": token,
                "ligand": m.group("lig"), "code": m.group("code"), "line": lineno}
    m = _DISTANCE_RE.match(stripped)
    if m:
        c1, r1, a1, c2, r2, a2, dist = m.groups()
        t1 = [c1, int(r1), a1] if a1 else [c1, int(r1)]
        t2 = [c2, int(r2), a2] if a2 else [c2, int(r2)]
        entry = [t1, t2, float(dist)] if dist else [t1, t2]
        return {"type": "distance", "owner": c1, "entry": entry, "line": lineno}
    return None


def _extract_statements(raw_lines: list):
    statements, remaining = [], []
    for lineno, raw in raw_lines:
        stripped = _strip_comment(raw).strip()
        if stripped:
            stmt = _match_statement(stripped, lineno)
            if stmt is not None:
                statements.append(stmt)
                continue
        remaining.append((lineno, raw))
    return statements, remaining


def _split_records(lines: list) -> list:
    records = []
    current = None  # [record_type, name, fields, lineno]
    for lineno, raw in lines:
        if raw.strip() == "":  # a genuinely blank line ends the record
            if current is not None:
                records.append(current)
                current = None
            continue
        stripped = _strip_comment(raw).strip()
        if not stripped:
            continue  # a comment-only line -- doesn't end the record, doesn't set anything
        m = _RECORD_START_RE.match(stripped)
        if m:
            if current is not None:
                records.append(current)
            record_type, name = m.group(1).lower(), _strip_quotes(m.group(2).strip())
            if record_type == "settings" and name:
                raise MDParseError(f"'Settings:' takes no value (line {lineno}): {stripped!r}")
            if record_type != "settings" and not name:
                raise MDParseError(f"'{m.group(1)}:' needs a name (line {lineno}): {stripped!r}")
            current = [record_type, name, {}, lineno]
            continue
        if current is None:
            continue  # stray content before the first record -- ignore
        fm = _FIELD_RE.match(stripped)
        if not fm:
            continue
        field_name, field_value = fm.group(1).strip().lower(), _strip_quotes(fm.group(2).strip())
        allowed = _RECORD_ALLOWED_FIELDS.get(current[0], set())
        if field_name not in allowed:
            _warn(f"unrecognized field '{fm.group(1).strip()}:' in "
                  f"{current[0].capitalize()} '{current[1]}' (line {lineno}) -- ignored, typo?")
            continue
        current[2][field_name] = field_value
    if current is not None:
        records.append(current)
    return records


def _build_partner_record(name: str, fields: dict, lineno: int) -> Partner:
    if "sequence" not in fields:
        raise MDParseError(f"partner '{name}' missing Sequence (line {lineno})")
    copies = _parse_csv(fields["copies"]) if "copies" in fields else None
    modifications = [_parse_modification_token(t) for t in _parse_csv(fields["modifications"])] if "modifications" in fields else None
    return Partner(
        id=copies if copies else name, sequence=fields["sequence"], type=fields.get("type", "protein").lower(),
        modifications=modifications, cyclic=_parse_yesno(fields.get("cyclic", ""), False), msa=fields.get("msa"),
    )


def _build_family_record(name: str, fields: dict, partners: dict, statements: list, lineno: int) -> ProteinFamily:
    if "sequence" not in fields:
        raise MDParseError(f"protein '{name}' missing Sequence (line {lineno})")
    partner_ids = _parse_csv(fields["partners"]) if "partners" in fields else []
    for pid in partner_ids:
        if pid not in partners:
            raise MDParseError(f"protein '{name}' references unknown partner '{pid}' (line {lineno})")
    modifications = [_parse_modification_token(t) for t in _parse_csv(fields["modifications"])] if "modifications" in fields else None
    pocket_contacts = [s["token"] for s in statements
                       if s["type"] == "pocket" and not s.get("code")] or None
    pockets: dict = {}
    for stmt in statements:
        if stmt["type"] == "pocket" and stmt.get("code"):
            pockets.setdefault(stmt["code"], []).append(stmt["token"])
    bond_constraints = [(s["atom1"], s["atom2"]) for s in statements if s["type"] == "bond"] or None
    contact_constraints = [s["entry"] for s in statements if s["type"] == "distance"] or None
    family_type = fields.get("family type", "auto").lower()
    if family_type not in ("gpcr", "kinase", "auto"):
        raise MDParseError(f"protein '{name}' has invalid Family type '{family_type}' "
                            f"(expected gpcr/kinase/auto, line {lineno})")
    # "Ligands: none" is a distinct sentinel from omitting the field entirely: omitted
    # means "default to every campaign ligand" (fan-out, existing behavior); "none" means
    # this family gets exactly one ligand-free (apo) target -- no ligand entity in the
    # YAML, no affinity property, stem is just the family id.
    if "ligands" in fields and fields["ligands"].strip().lower() == "none":
        ligands = []
    elif "ligands" in fields:
        ligands = _parse_csv(fields["ligands"])
    else:
        ligands = None
    return ProteinFamily(
        id=name, sequence=fields["sequence"], partners=partner_ids,
        pocket_contacts=pocket_contacts, pockets=(pockets or None), ligands=ligands,
        modifications=modifications, cyclic=_parse_yesno(fields.get("cyclic", ""), False), msa=fields.get("msa"),
        bond_constraints=bond_constraints, contact_constraints=contact_constraints,
        templates=_parse_csv(fields["templates"]) if "templates" in fields else None,
        apo_structure=fields.get("apo structure"), apo_chain=fields.get("apo chain"), family_type=family_type,
        group=fields.get("group"),
    )


def _canonicalize_smiles(smiles: str) -> str:
    # Silent normalization only -- an invalid SMILES is left as-is here and reported by
    # preflight's check_smiles (parsing shouldn't fail or change error timing over a
    # chemistry problem). A consistent canonical form flowing through the YAML, the
    # summary table, and cif2plip's own ligand-matching (see _analyze_target_interactions)
    # is the actual payoff, not just cosmetics.
    try:
        from rdkit import Chem
        mol = Chem.MolFromSmiles(smiles)
        return Chem.MolToSmiles(mol) if mol is not None else smiles
    except Exception:
        return smiles


def _smiles_to_inchikey(smiles: str):
    try:
        from rdkit import Chem
        mol = Chem.MolFromSmiles(smiles)
        return Chem.MolToInchiKey(mol) if mol is not None else None
    except Exception:
        return None


def _build_ligand_record(name: str, fields: dict, statements: list, lineno: int) -> Ligand:
    has_smiles, has_ccd = "smiles" in fields, "ccd" in fields
    if has_smiles == has_ccd:  # both True (ambiguous) or both False (missing)
        raise MDParseError(f"ligand '{name}' must specify exactly one of SMILES/CCD (line {lineno})")
    smiles = _canonicalize_smiles(fields["smiles"]) if has_smiles else None
    role = fields.get("role", "").strip().lower() or None
    if role and role not in ("agonist", "antagonist"):
        raise MDParseError(f"ligand '{name}' has invalid Role '{role}' "
                            f"(expected agonist/antagonist, line {lineno})")
    pocket_contacts = [s["token"] for s in statements if s["type"] == "pocket"] or None
    return Ligand(id=name, smiles=smiles, ccd=fields.get("ccd"), role=role,
                  pocket_contacts=pocket_contacts)


# ==========================================================================
# Formatter -- purely cosmetic (comment-column alignment, blank-line spacing
# around record boundaries). Never changes parsed meaning; `cmd_format`
# validates via parse_md() before writing anything back.
# ==========================================================================

def _normalize_blank_lines(text: str) -> str:
    lines = text.splitlines()
    out = []
    for raw in lines:
        stripped = _strip_comment(raw).strip()
        if stripped and _RECORD_START_RE.match(stripped) and out and out[-1].strip() != "":
            out.append("")
        out.append(raw)
    result = "\n".join(out)
    return result + "\n" if text.endswith("\n") else result


def _format_block(block: list) -> list:
    parsed = []
    for line in block:
        idx = _find_comment_start(line)
        parsed.append((line.rstrip(), None) if idx is None else (line[:idx].rstrip(), line[idx:].rstrip()))
    candidates = [len(code) for code, comment in parsed if comment is not None and code.strip() != ""]
    if not candidates:
        return list(block)  # nothing to align (e.g. a pure comment block) -- keep as-is
    target_col = max(candidates) + 2
    out = []
    for code, comment in parsed:
        if comment is None:
            out.append(code)
        elif code.strip() == "":
            out.append(" " * target_col + comment)
        else:
            out.append(code.ljust(target_col) + comment)
    return out


def format_md_text(text: str) -> str:
    text = _normalize_blank_lines(text)
    lines = text.splitlines()
    out, i = [], 0
    while i < len(lines):
        if lines[i].strip() == "":
            out.append(lines[i])
            i += 1
            continue
        start = i
        while i < len(lines) and lines[i].strip() != "":
            i += 1
        out.extend(_format_block(lines[start:i]))
    result = "\n".join(out)
    return result + "\n" if text.endswith("\n") else result


def cmd_format(md_path: Path, check: bool = False) -> None:
    original = md_path.read_text()
    parse_md(md_path)  # validate first: surfaces MDParseError / unknown-field warnings
    formatted = format_md_text(original)
    if formatted == original:
        _ok(f"{md_path} already formatted.")
        return
    if check:
        _warn(f"{md_path} would be reformatted (comment alignment / blank-line spacing).")
        sys.exit(1)
    md_path.write_text(formatted)
    _ok(f"reformatted {md_path}.")


# ==========================================================================
# Wizard -- `BoltzMaker.py new` interviews a non-specialist user in plain
# language and writes a boltz_input.md in the format above. Covers the
# common case only (proteins, partners, ligands, the three constraint
# sentence-types); rarer fields (modifications, cyclic, MSA override,
# templates, homo-oligomer copies) are left for hand-editing the file
# afterward.
# ==========================================================================

def _wiz_prompt(msg: str, default: str = None) -> str:
    suffix = f" {_DIM}[{default}]{_RESET}" if default is not None else ""
    while True:
        val = input(f"{_BLUE}?{_RESET} {msg}{suffix}{_DIM}:{_RESET} ").strip()
        if val:
            return val
        if default is not None:
            return default


def _wiz_yesno(msg: str, default: bool = False) -> bool:
    hint = f"{_GREEN}Y{_RESET}{_DIM}/n{_RESET}" if default else f"{_DIM}y/{_RESET}{_GREEN}N{_RESET}"
    while True:
        val = input(f"{_BLUE}?{_RESET} {msg} {_DIM}({_RESET}{hint}{_DIM}){_RESET} ").strip().lower()
        if not val:
            return default
        if val in ("y", "yes"):
            return True
        if val in ("n", "no"):
            return False
        _warn("please answer y or n")


def _wiz_name(msg: str, taken: set) -> str:
    while True:
        val = input(f"{_BLUE}?{_RESET} {msg}{_DIM}:{_RESET} ").strip()
        if not val:
            _warn("a name is required")
        elif len(val) > 5:
            _warn(f"'{val}' is {len(val)} characters -- Boltz needs chain ids MAX 5 CHARACTERS, try again")
        elif val in taken:
            _warn(f"'{val}' is already used, pick a different name")
        else:
            return val


def _align_positions(ref_seq: str, target_seq: str) -> dict:
    # Maps 0-indexed ref_seq positions -> 0-indexed target_seq positions for aligned
    # (non-gap) regions only. BLOSUM62 + affine gaps -- standard protein alignment
    # defaults, not hand-rolled, since residue-index remapping here feeds directly into
    # constraint statements a user may accept without double-checking.
    from Bio import Align
    from Bio.Align import substitution_matrices

    aligner = Align.PairwiseAligner()
    aligner.substitution_matrix = substitution_matrices.load("BLOSUM62")
    aligner.open_gap_score = -10
    aligner.extend_gap_score = -0.5
    aligner.mode = "global"
    alignment = aligner.align(ref_seq, target_seq)[0]
    mapping = {}
    for (r_start, r_end), (t_start, t_end) in zip(*alignment.aligned):
        for offset in range(r_end - r_start):
            mapping[r_start + offset] = t_start + offset
    return mapping


def _wiz_reference_structure_suggestions(name: str, sequence: str) -> list:
    """Optionally analyze a reference structure with a bound ligand and suggest pocket
    residues for `name`, remapped onto `sequence`'s own numbering via pairwise sequence
    alignment. Returns a list of target-numbered residue ints (possibly empty)."""
    if not _wiz_yesno(f"Do you have a reference structure with a ligand already bound for {name} "
                       f"(co-crystal or homology model)", default=False):
        return []
    ref_path = Path(_wiz_prompt("  Path to the reference structure (.cif/.pdb/.mmcif)")).expanduser()
    if not ref_path.exists():
        _warn(f"{ref_path} not found -- skipping")
        return []

    import tempfile
    with tempfile.TemporaryDirectory() as tmp:
        work_dir = Path(tmp)
        _step("running cif2plip on the reference structure...")
        try:
            proc = _run_cif2plip(ref_path, work_dir)
        except subprocess.TimeoutExpired:
            _warn("cif2plip timed out on that structure -- skipping")
            return []
        inter_csv = work_dir / f"{ref_path.stem}_interactions.csv"
        summ_csv = work_dir / f"{ref_path.stem}_ligand_summary.csv"
        pdb_path = work_dir / f"{ref_path.stem}.pdb"
        if proc.returncode != 0 or not inter_csv.exists() or not summ_csv.exists():
            _warn("cif2plip couldn't process that structure -- skipping")
            return []

        inter_df = pd.read_csv(inter_csv)
        summ_df = pd.read_csv(summ_csv).reset_index(drop=True)
        if summ_df.empty:
            _warn("no ligands detected in that structure -- skipping")
            return []

        if len(summ_df) == 1:
            chosen = summ_df.iloc[0]["ligand"]
        else:
            _info("Multiple ligands detected in the reference structure:")
            for i, r in summ_df.iterrows():
                print(f"    {_DIM}[{i + 1}]{_RESET} {r['ligand']} (SMILES: {r.get('smiles', '?')}, "
                      f"{r.get('total_interactions', '?')} interactions)")
            choice = input(f"{_BLUE}?{_RESET}   Which one is the relevant bound ligand? "
                           f"{_DIM}[1-{len(summ_df)}]:{_RESET} ").strip()
            try:
                chosen = summ_df.iloc[int(choice) - 1]["ligand"]
            except (ValueError, IndexError):
                _warn("not a valid choice -- skipping")
                return []

        sub = inter_df[inter_df["ligand"] == chosen]
        if sub.empty:
            _warn("that ligand has no recorded interactions -- skipping")
            return []
        ref_chain_id = sub.iloc[0]["prot_chain"]
        contact_resnrs = sorted(int(x) for x in sub["prot_resnr"].unique())

        try:
            import gemmi
            st = gemmi.read_structure(str(pdb_path))
            st.setup_entities()
            polymer = st[0][ref_chain_id].get_polymer()
            ref_seq = polymer.make_one_letter_sequence()
            resnr_to_pos = {res.seqid.num: i for i, res in enumerate(polymer)}
        except Exception as exc:
            _warn(f"couldn't extract the reference chain's sequence -- skipping ({exc})")
            return []

        mapping = _align_positions(ref_seq, sequence)
        suggested, skipped = [], 0
        for resnr in contact_resnrs:
            ref_pos = resnr_to_pos.get(resnr)
            target_pos = mapping.get(ref_pos) if ref_pos is not None else None
            if target_pos is None:
                skipped += 1
                continue
            suggested.append(target_pos + 1)  # 1-indexed residue number, matching Boltz convention

        if not suggested:
            _warn(f"none of the reference structure's contact residues could be mapped onto "
                  f"{name}'s sequence -- skipping")
            return []
        if skipped:
            _info(f"({skipped} reference residue(s) had no equivalent position in {name}'s sequence, skipped)")
        _ok(f"Found {len(suggested)} candidate pocket residue(s) from the reference structure: "
            f"{', '.join(str(r) for r in suggested)}")
        if _wiz_yesno("  Add these as pocket-contact constraints", default=True):
            return suggested
        return []


def _wiz_section(title: str) -> None:
    print(f"\n{_BOLD}{_AMBER}{title}{_RESET}")


def cmd_new(md_path: Path) -> None:
    if md_path.exists() and not _wiz_yesno(f"{md_path} already exists -- overwrite", default=False):
        _err("aborted, nothing written.")
        return

    _info("let's set up a new campaign. Press Ctrl-C any time to cancel.")
    try:
        predict_affinity = _wiz_yesno("Predict binding affinity too (slower, adds Kd/pIC50 estimates)", default=False)
        out = ["Settings:", "Output folder: ./boltz_yamls",
               f"Predict affinity: {'yes' if predict_affinity else 'no'}"]

        used_names, known_partners = set(), set()
        partner_blocks, protein_blocks, statement_lines = [], [], []

        _wiz_section("Protein(s) -- at least one is required")
        first = True
        while first or _wiz_yesno("Add another protein", default=False):
            first = False
            name = _wiz_name("Protein short name (max 5 letters)", used_names)
            used_names.add(name)
            sequence = _wiz_prompt(f"Paste the amino acid sequence for {name}")
            block = [f"Protein: {name}", f"Sequence: {sequence}"]

            if _plip_available():
                for r in _wiz_reference_structure_suggestions(name, sequence):
                    statement_lines.append(f"Pocket contact: {name} residue {r}")

            if _wiz_yesno(f"Does {name} co-fold with any partner chains", default=False):
                partner_ids = []
                add_more = True
                while add_more:
                    pname = _wiz_name("  Partner short name (max 5 letters)", used_names)
                    if pname not in known_partners:
                        used_names.add(pname)
                        known_partners.add(pname)
                        psequence = _wiz_prompt(f"  Paste the sequence for partner {pname}")
                        partner_blocks.append([f"Partner: {pname}", f"Sequence: {psequence}"])
                    partner_ids.append(pname)
                    add_more = _wiz_yesno("  Add another partner", default=False)
                block.append(f"Partners: {', '.join(partner_ids)}")

            while _wiz_yesno(f"Add a constraint on {name}", default=False):
                choice = input(f"{_BLUE}?{_RESET}   {_DIM}[1]{_RESET} Covalent bond  {_DIM}[2]{_RESET} Pocket "
                               f"contact  {_DIM}[3]{_RESET} Distance constraint{_DIM}:{_RESET} ").strip()
                if choice == "1":
                    r1 = _wiz_prompt(f"  {name} residue number")
                    a1 = _wiz_prompt(f"  {name} atom name (e.g. SG for a cysteine sulfur)")
                    other = _wiz_prompt("  Name of the ligand/protein it bonds to")
                    r2 = _wiz_prompt(f"  {other} residue number")
                    a2 = _wiz_prompt(f"  {other} atom name")
                    statement_lines.append(f"Covalent bond: {name} residue {r1} atom {a1} to {other} residue {r2} atom {a2}")
                elif choice == "2":
                    r1 = _wiz_prompt(f"  {name} residue number")
                    statement_lines.append(f"Pocket contact: {name} residue {r1}")
                elif choice == "3":
                    r1 = _wiz_prompt(f"  {name} residue number")
                    other = _wiz_prompt("  Name of the other protein")
                    r2 = _wiz_prompt(f"  {other} residue number")
                    dist = _wiz_prompt("  Maximum distance in Angstrom", default="6.0")
                    statement_lines.append(f"Distance constraint: {name} residue {r1} to {other} residue {r2} within {dist} Angstrom")
                else:
                    _warn("not a recognized choice, skipping")
            protein_blocks.append(block)

        _wiz_section("Ligand(s) -- at least one is required")
        ligand_blocks = []
        first = True
        while first or _wiz_yesno("Add another ligand", default=False):
            first = False
            name = _wiz_name("Ligand short name (max 5 letters)", used_names)
            used_names.add(name)
            kind = input(f"{_BLUE}?{_RESET}   SMILES or CCD code? {_DIM}[1]{_RESET} SMILES  "
                         f"{_DIM}[2]{_RESET} CCD{_DIM}:{_RESET} ").strip()
            if kind == "2":
                code = _wiz_prompt(f"  CCD code for {name}")
                ligand_blocks.append([f"Ligand: {name}", f"CCD: {code}"])
            else:
                smiles = _wiz_prompt(f"  SMILES for {name}")
                try:
                    from rdkit import Chem
                    if Chem.MolFromSmiles(smiles) is None:
                        _warn("rdkit couldn't parse that SMILES -- saved anyway, double-check it")
                except Exception:
                    pass
                ligand_blocks.append([f"Ligand: {name}", f"SMILES: {smiles}"])
    except (KeyboardInterrupt, EOFError):
        _err("cancelled, nothing written.")
        return

    for block in protein_blocks + partner_blocks + ligand_blocks:
        out.append("")
        out.extend(block)
    if statement_lines:
        out.append("")
        out.extend(statement_lines)

    md_path.write_text("\n".join(out) + "\n")
    print()
    _ok(f"wrote {md_path}")
    _info(f"Next: python3 BoltzMaker.py preflight {md_path}")


def parse_md(path: Path) -> Campaign:
    text = path.read_text()
    raw_lines = list(enumerate(text.splitlines(), start=1))
    statements, remaining = _extract_statements(raw_lines)
    records = _split_records(remaining)

    settings = Settings()
    partners: dict = {}
    protein_records, ligand_records = [], []
    for record_type, name, fields, lineno in records:
        if record_type == "settings":
            settings.output_dir = fields.get("output folder", settings.output_dir)
            raw_chunk = fields.get("targets per invocation", "").strip()
            if raw_chunk:
                try:
                    settings.targets_per_invocation = int(raw_chunk)
                except ValueError:
                    raise MDParseError(
                        f"Settings: 'Targets per invocation: {raw_chunk}' is not a whole number.")
                if settings.targets_per_invocation < 0:
                    raise MDParseError(
                        "Settings: 'Targets per invocation' cannot be negative (0 disables recycling).")
            raw_distance = fields.get("pocket distance", "").strip()
            if raw_distance:
                try:
                    settings.pocket_distance = float(raw_distance)
                except ValueError:
                    raise MDParseError(
                        f"Settings: 'Pocket distance: {raw_distance}' is not a number.")
                if not 1.0 <= settings.pocket_distance <= 50.0:
                    raise MDParseError(
                        f"Settings: 'Pocket distance: {raw_distance}' is outside 1-50 A.")
            settings.predict_affinity = _parse_yesno(fields.get("predict affinity", ""), settings.predict_affinity)
        elif record_type == "partner":
            partners[name] = _build_partner_record(name, fields, lineno)
        elif record_type == "protein":
            protein_records.append((name, fields, lineno))
        elif record_type == "ligand":
            ligand_records.append((name, fields, lineno))

    # A pocket written "... for LIG" belongs to that ligand, not to every ligand of
    # the protein it names. Routed here rather than by which block it appeared in,
    # because statements are extracted from the whole file before it is split into
    # blocks, so there is no block to attribute them to.
    statements_by_owner: dict = {}
    statements_by_ligand: dict = {}
    for stmt in statements:
        if stmt.get("ligand"):
            statements_by_ligand.setdefault(stmt["ligand"], []).append(stmt)
        else:
            statements_by_owner.setdefault(stmt["owner"], []).append(stmt)

    families, seen_fam = [], set()
    for name, fields, lineno in protein_records:
        if name in seen_fam:
            raise MDParseError(f"duplicate protein '{name}' (line {lineno})")
        seen_fam.add(name)
        families.append(_build_family_record(name, fields, partners, statements_by_owner.pop(name, []), lineno))

    if statements_by_owner:
        owner, stmts = next(iter(statements_by_owner.items()))
        raise MDParseError(f"a constraint (line {stmts[0]['line']}) names '{owner}' as the owning protein, "
                            f"but no 'Protein: {owner}' block exists")

    ligands, seen_lig = [], set()
    for name, fields, lineno in ligand_records:
        if name in seen_lig:
            raise MDParseError(f"duplicate ligand '{name}' (line {lineno})")
        seen_lig.add(name)
        ligands.append(_build_ligand_record(name, fields, statements_by_ligand.pop(name, []), lineno))

    if statements_by_ligand:
        lig, stmts = next(iter(statements_by_ligand.items()))
        raise MDParseError(f"a pocket contact (line {stmts[0]['line']}) is written "
                            f"'for {lig}', but no 'Ligand: {lig}' block exists")
    if not families:
        raise MDParseError("no 'Protein:' blocks found")
    if not ligands:
        raise MDParseError("no 'Ligand:' blocks found")
    return Campaign(settings=settings, partners=partners, families=families, ligands=ligands, source_path=path)


# ==========================================================================
# YamlGenerator
# ==========================================================================

def _expand_targets(campaign: Campaign):
    ligand_by_id = {l.id: l for l in campaign.ligands}
    targets = []
    for fam in campaign.families:
        if fam.ligands == []:  # explicit "Ligands: none" -- one ligand-free (apo) target
            targets.append((fam, None, None))
            continue
        ligand_ids = fam.ligands if fam.ligands else [l.id for l in campaign.ligands]
        for lig_id in ligand_ids:
            if lig_id not in ligand_by_id:
                raise MDParseError(f"protein '{fam.id}' references unknown ligand '{lig_id}'")
            lig = ligand_by_id[lig_id]
            if not fam.pockets:
                # No named pockets: one target, exactly as before this existed.
                targets.append((fam, lig, None))
                continue
            # Named pockets fan out: one target per pocket, plus an unconstrained
            # baseline, so what the constraint changed is visible inside one campaign
            # rather than by cross-referencing an older run with different settings.
            targets.append((fam, lig, None))
            for code in sorted(fam.pockets):
                targets.append((fam, lig, code))
    return targets


def _target_stem(fam, lig, code) -> str:
    """protein_ligand, plus the pocket's ligand code when one is in use."""
    stem = fam.id if lig is None else f"{fam.id}_{lig.id}"
    return f"{stem}_{code}" if code else stem


# Homo-oligomer copies: `id: [A, B]` on a *partner* shares one sequence across
# multiple chain ids -- real YAML already parses this straight into a Python
# list, so `Partner.id` is just used as-is (str or list). Only partners support
# this: the primary family id also names the output file
# (`{family_id}_{ligand_id}.yaml`), so it must stay a plain single token.


def _chain_entry(chain_id, sequence: str, entity_type: str = "protein",
                  modifications=None, cyclic: bool = False, msa=None) -> dict:
    entry = {"id": chain_id, "sequence": sequence}
    if modifications:
        entry["modifications"] = [{"ccd": m[0], "position": m[1]} for m in modifications]
    if cyclic:
        entry["cyclic"] = True
    if msa:
        entry["msa"] = msa
    return {entity_type: entry}


def _ligand_entry(lig: Ligand) -> dict:
    key, value = ("smiles", lig.smiles) if lig.smiles is not None else ("ccd", lig.ccd)
    return {"ligand": {key: value, "id": lig.id}}


def _pocket_for(fam, lig, code):
    """Which contacts this target actually uses.

    A named pocket wins (it is the thing being varied across the matrix), then a
    ligand's own pocket, then the protein's unnamed one. None means unconstrained,
    which is a real choice here rather than an absence: every ligand also gets a
    baseline target so the effect of a constraint is visible within one campaign.
    """
    if code:
        return (fam.pockets or {}).get(code)
    if lig is not None and lig.pocket_contacts:
        return lig.pocket_contacts
    if fam.pockets:
        return None          # the baseline target of a matrix campaign
    return fam.pocket_contacts


def _build_yaml_doc(fam: ProteinFamily, lig: object, campaign: Campaign,
                     pocket_code: object = None) -> dict:
    # lig is None for a ligand-free (apo) target ("Ligands: none") -- no ligand entity,
    # no pocket/affinity binder (both are meaningless without a ligand to bind).
    sequences = [_chain_entry(fam.id, fam.sequence, "protein", fam.modifications, fam.cyclic, fam.msa)]
    for pid in fam.partners:
        p = campaign.partners[pid]
        sequences.append(_chain_entry(p.id, p.sequence, p.type, p.modifications, p.cyclic, p.msa))
    if lig is not None:
        sequences.append(_ligand_entry(lig))
    doc = {"sequences": sequences}

    constraints = []
    # A ligand's own pocket wins over the protein's: the ligand is what has an
    # experimentally observed site, and two chemotypes on the same receptor can
    # occupy different ones.
    pocket_contacts = _pocket_for(fam, lig, pocket_code)
    if pocket_contacts and lig is not None:
        # Boltz's pocket constraint requires every contact entry to be an explicit
        # [chain, residue_or_atom] pair (verified against the installed boltz 2.2.1
        # schema parser) -- there is no whole-chain-only shorthand, so a family with
        # no pocket_contacts gets no pocket constraint at all (unconstrained folding).
        pocket = {"binder": lig.id, "contacts": pocket_contacts}
        # Only written when it differs from Boltz's own default, so an existing
        # campaign's YAML is byte-identical to what it was before this setting existed.
        if campaign.settings.pocket_distance != 6.0:
            pocket["max_distance"] = campaign.settings.pocket_distance
        constraints.append({"pocket": pocket})
    for atom1, atom2 in (fam.bond_constraints or []):
        constraints.append({"bond": {"atom1": atom1, "atom2": atom2}})
    for entry in (fam.contact_constraints or []):
        token1, token2 = entry[0], entry[1]
        contact = {"token1": token1, "token2": token2}
        if len(entry) > 2:
            contact["max_distance"] = entry[2]
        constraints.append({"contact": contact})
    if constraints:
        doc["constraints"] = constraints

    if fam.templates:
        doc["templates"] = [
            {("pdb" if str(path).lower().endswith(".pdb") else "cif"): path} for path in fam.templates
        ]

    if campaign.settings.predict_affinity and lig is not None:
        doc["properties"] = [{"affinity": {"binder": lig.id}}]
    return doc


def generate_yamls(campaign: Campaign, output_dir: Path) -> list:
    output_dir.mkdir(parents=True, exist_ok=True)
    manifest, seen = [], set()
    for fam, lig, pocket_code in _expand_targets(campaign):
        stem = _target_stem(fam, lig, pocket_code)
        if stem in seen:
            raise MDParseError(f"duplicate target filename '{stem}.yaml' -- check for duplicate family/ligand ids")
        seen.add(stem)
        doc = _build_yaml_doc(fam, lig, campaign, pocket_code)
        with (output_dir / f"{stem}.yaml").open("w") as f:
            yaml.safe_dump(doc, f, sort_keys=False, default_flow_style=False)
        needs_affinity = campaign.settings.predict_affinity and lig is not None
        manifest.append(Target(stem=stem, family_id=fam.id, ligand_id=(lig.id if lig is not None else None),
                                # same precedence as _build_yaml_doc: the ligand's own
                                # pocket wins over the protein's.
                                pocket_contacts_used=(
                                    _pocket_for(fam, lig, pocket_code)
                                    if lig is not None else None),
                                needs_affinity=needs_affinity))
    with (output_dir / MANIFEST_FILENAME).open("w") as f:
        json.dump([asdict(t) for t in manifest], f, indent=2)
    return manifest


def load_manifest(output_dir: Path) -> list:
    manifest_path = output_dir / MANIFEST_FILENAME
    if not manifest_path.exists():
        _err(f"no manifest at {manifest_path} -- run `generate` first.")
        sys.exit(1)
    with manifest_path.open() as f:
        return [Target(**d) for d in json.load(f)]


# ==========================================================================
# Preflight
# ==========================================================================

@dataclass
class CheckResult:
    name: str
    status: str  # PASS / WARN / FAIL
    message: str


def check_boltz_cli() -> CheckResult:
    boltz_path = _boltz_bin()
    if not boltz_path.exists():
        if not _in_pixi_env():
            fix = "run `setup`"
        elif shutil.which("pixi"):
            fix = "run `pixi install`"
        else:
            fix = "the environment looks incomplete -- re-run the offline installer"
        return CheckResult("boltz_cli", "FAIL", f"{boltz_path} not found -- {fix}")
    try:
        out = subprocess.run([str(boltz_path), "--help"], capture_output=True, text=True,
                             timeout=BOLTZ_CLI_HELP_TIMEOUT)
        ok = out.returncode == 0
        return CheckResult("boltz_cli", "PASS" if ok else "WARN",
                            f"{boltz_path} {'reachable' if ok else f'exited {out.returncode} on --help'}")
    except subprocess.TimeoutExpired:
        # Deliberately WARN, not FAIL. The binary exists -- that was checked above --
        # so a slow `--help` is evidence about how long an import took, not about
        # whether boltz works. `boltz --help` imports the entire torch stack, and the
        # very first invocation in a freshly-solved environment also byte-compiles it
        # with a cold page cache; on an iCloud-backed directory that alone can run to
        # minutes. Failing here aborted whole campaigns before any GPU work on nothing
        # more than a cold start, which is the wrong trade: if boltz really is broken,
        # `run` says so immediately and loudly. --strict still promotes this to FAIL
        # for anyone who wants the cautious behaviour.
        return CheckResult(
            "boltz_cli", "WARN",
            f"{boltz_path} exists but did not answer --help within {BOLTZ_CLI_HELP_TIMEOUT}s "
            "-- usually just a cold first import (torch is large); the run itself will "
            "confirm. Re-run preflight once it has warmed up to clear this.",
        )
    except Exception as e:
        return CheckResult("boltz_cli", "FAIL", f"error invoking {boltz_path}: {e}")


def check_sse_comparison() -> CheckResult:
    """The compare-sse package must sit beside the script.

    `analyze` imports sse_comparison late, only when it reaches compare-sse, so a
    missing package is not noticed until the prediction has already finished. A
    real bundled run spent 45 minutes predicting and then died on
    "ModuleNotFoundError: No module named 'sse_comparison'". Checking it in
    preflight costs nothing and moves that failure to before the GPU work.
    """
    package = SCRIPT_DIR / "sse_comparison"
    if not (package / "__init__.py").is_file():
        return CheckResult(
            "sse_comparison", "FAIL",
            f"{package} is missing -- `analyze` imports it after prediction finishes, so "
            "the run would complete and then fail. Re-download the bundle, or re-clone.",
        )
    missing = [name for name in ("cli.py", "report.py", "alignment.py", "metrics.py")
               if not (package / name).is_file()]
    if missing:
        return CheckResult("sse_comparison", "FAIL",
                           f"sse_comparison is incomplete -- missing {', '.join(missing)}")
    return CheckResult("sse_comparison", "PASS", "compare-sse package present")


def check_vendor_assets() -> CheckResult:
    """The dashboard embeds these; without them it silently reaches for a CDN."""
    missing = [path.name for path in (PLOTLY_JS_PATH, THREEDMOL_JS_PATH) if not path.exists()]
    if missing:
        return CheckResult(
            "vendor_assets", "WARN",
            f"missing {', '.join(missing)} in vendor/ -- the dashboard will fall back to a "
            "CDN, so it will need the network to render and will not work offline",
        )
    return CheckResult("vendor_assets", "PASS", "Plotly and 3Dmol bundled for an offline dashboard")


def check_result_packer() -> CheckResult:
    """Only meaningful inside a bundle: a plain checkout has no packer and wants none.

    A bundle is identified by the runner script it was built with, not by the
    packer itself -- otherwise a bundle whose packer went missing would look like
    a plain checkout and the check would pass by disappearing.
    """
    packer = SCRIPT_DIR / "pack_results.py"
    if not (SCRIPT_DIR / "run_campaign.sh").is_file():
        return CheckResult("result_packer", "PASS", "not a bundle -- no results file to pack")
    if not packer.is_file():
        return CheckResult(
            "result_packer", "FAIL",
            "pack_results.py is missing from this bundle -- the campaign would run and "
            "then have no way to write its .bmz. Re-download the bundle.",
        )
    try:
        ast.parse(packer.read_text())
    except (SyntaxError, OSError) as exc:
        return CheckResult("result_packer", "FAIL", f"pack_results.py will not parse: {exc}")
    return CheckResult("result_packer", "PASS", "results packer present and valid")


def check_gpu() -> CheckResult:
    try:
        import torch
        if torch.backends.mps.is_available():
            return CheckResult("gpu", "PASS", "Apple MPS backend available")
        if torch.cuda.is_available():
            return CheckResult("gpu", "PASS", "CUDA backend available")
        return CheckResult("gpu", "WARN", "no GPU/MPS backend detected -- boltz will run on CPU (slow)")
    except ImportError:
        return CheckResult("gpu", "FAIL", "torch not importable -- run `setup`")


def check_disk_space(out_dir: Path, n_targets: int) -> CheckResult:
    out_dir.mkdir(parents=True, exist_ok=True)
    usage = shutil.disk_usage(out_dir)
    threshold = max(10 * 1024**3, n_targets * 200 * 1024**2)
    free_gb, need_gb = usage.free / 1e9, threshold / 1e9
    if usage.free < threshold:
        return CheckResult("disk_space", "FAIL", f"only {free_gb:.1f} GB free at {out_dir}, need ~{need_gb:.1f} GB")
    return CheckResult("disk_space", "PASS", f"{free_gb:.1f} GB free at {out_dir}")


def _is_dataless(path: Path) -> bool:
    st = path.stat()
    blocks = getattr(st, "st_blocks", None)
    return blocks is not None and st.st_size > 0 and blocks == 0


def ensure_materialized(path: Path, timeout: float = 30.0) -> CheckResult:
    # macOS "Optimize Mac Storage" can evict iCloud-synced files to dataless
    # placeholders (non-zero logical size, zero blocks on disk).
    if not path.exists():
        return CheckResult(f"materialize:{path.name}", "FAIL", f"{path} does not exist")
    if not _is_dataless(path):
        return CheckResult(f"materialize:{path.name}", "PASS", f"{path.name} is local")
    subprocess.run(["brctl", "download", str(path)], capture_output=True)
    start = time.time()
    while time.time() - start < timeout:
        if not _is_dataless(path):
            return CheckResult(f"materialize:{path.name}", "PASS", f"{path.name} materialized via brctl")
        time.sleep(1)
    return CheckResult(f"materialize:{path.name}", "FAIL",
                        f"{path.name} still dataless/evicted after {timeout:.0f}s (iCloud eviction?)")


def check_all_materialized(paths: list) -> CheckResult:
    failed, recovered = [], 0
    for p in paths:
        r = ensure_materialized(p)
        if r.status == "FAIL":
            failed.append(p.name)
        elif "materialized via brctl" in r.message:
            recovered += 1
    if failed:
        shown = failed[:5]
        return CheckResult("icloud_materialize", "FAIL",
                            f"{len(failed)} file(s) still dataless/evicted: {shown}{'...' if len(failed) > 5 else ''}")
    msg = f"{len(paths)} file(s) local"
    if recovered:
        msg += f" ({recovered} recovered from iCloud eviction via brctl)"
    return CheckResult("icloud_materialize", "PASS", msg)


def check_yaml_validity(manifest: list, output_dir: Path, campaign: Campaign) -> CheckResult:
    bad = []
    for t in manifest:
        path = output_dir / f"{t.stem}.yaml"
        try:
            with path.open() as f:
                doc = yaml.safe_load(f)
        except Exception as e:
            bad.append(f"{path.name}: parse error ({e})")
            continue
        if not doc.get("sequences"):
            bad.append(f"{path.name}: missing/empty 'sequences'")
            continue
        binder_ids = {e["ligand"]["id"] for e in doc["sequences"] if "ligand" in e}
        for c in doc.get("constraints", []):
            if "pocket" not in c:
                continue  # bond/contact constraints reference chains/atoms directly, not a ligand "binder"
            binder = c["pocket"].get("binder")
            if binder not in binder_ids:
                bad.append(f"{path.name}: pocket constraint binder '{binder}' not among ligand ids")
        if campaign.settings.predict_affinity and binder_ids:
            # binder_ids empty means a ligand-free (apo) target -- predict_affinity is
            # meaningless without a ligand, so it correctly has no properties block.
            aff_binders = {p["affinity"]["binder"] for p in doc.get("properties", []) if "affinity" in p}
            if not aff_binders & binder_ids:
                bad.append(f"{path.name}: predict_affinity is on but no matching affinity property found")
    if bad:
        return CheckResult("yaml_validity", "FAIL", f"{len(bad)} problem(s): {bad[:5]}")
    return CheckResult("yaml_validity", "PASS", f"{len(manifest)} yaml file(s) valid")


def check_smiles(campaign: Campaign) -> CheckResult:
    # ccd-code ligands aren't SMILES and are validated by Boltz itself at runtime.
    from rdkit import Chem
    smiles_ligands = [lig for lig in campaign.ligands if lig.smiles is not None]
    bad, oversized, warn = [], [], []
    for lig in smiles_ligands:
        mol = Chem.MolFromSmiles(lig.smiles)
        if mol is None:
            bad.append(lig.id)
            continue
        if campaign.settings.predict_affinity:
            n_atoms = mol.GetNumAtoms()
            if n_atoms > 128:
                oversized.append(f"{lig.id} ({n_atoms} atoms)")
            elif n_atoms > 56:
                warn.append(f"{lig.id} ({n_atoms} atoms)")
    if bad:
        return CheckResult("smiles_validity", "FAIL", f"invalid SMILES for: {bad}")
    if oversized:
        return CheckResult("smiles_validity", "FAIL",
                            f"exceeds Boltz's 128-heavy-atom affinity limit: {oversized}")
    if warn:
        return CheckResult("smiles_validity", "WARN",
                            f"{len(smiles_ligands)} SMILES parse OK; {warn} exceed Boltz's 56-atom "
                            "affinity 'trained size' threshold (still runs, may be less accurate)")
    return CheckResult("smiles_validity", "PASS", f"{len(smiles_ligands)} SMILES parse OK")


_IONIZABLE_SMARTS = {
    "carboxylic acid": "[CX3](=O)[OX2H1]",
    "primary/secondary amine": "[NX3;H2,H1;!$(NC=O);!$(N=*)]",
    "phenol": "[OX2H][cX3]",
    "sulfonic acid": "[SX4](=O)(=O)[OX2H1]",
}


def _ligand_chemistry_notes(campaign: Campaign) -> dict:
    # Bad input chemistry (an undefined stereocentre, a salt/counterion left attached, an
    # ionizable group whose intended protonation state is ambiguous) doesn't error --
    # Boltz folds whatever it's given and the pose/affinity is silently wrong. Shared by
    # preflight's check_ligand_preparation, the dashboard's "Ligand preparation" card, and
    # the ligand-grid panel's per-atom highlighting, so none of the three can drift out of
    # sync. This only advises; it never blocks -- these are legitimate modelling choices
    # the user may have already made deliberately.
    #
    # Returns {ligand_id: {"notes": [str, ...], "stereo_atoms": [int, ...],
    #                      "ionizable_atoms": {group_name: [int, ...]}, "has_fragments": bool}}
    # -- only for ligands with at least one finding.
    from rdkit import Chem

    patterns = {name: Chem.MolFromSmarts(smarts) for name, smarts in _IONIZABLE_SMARTS.items()}

    notes_by_ligand = {}
    for lig in campaign.ligands:
        if lig.smiles is None:
            continue  # CCD ligands are pre-defined dictionary entries, not raw SMILES
        mol = Chem.MolFromSmiles(lig.smiles)
        if mol is None:
            continue  # already reported as a FAIL by check_smiles

        notes = []
        has_fragments = len(Chem.GetMolFrags(mol)) > 1
        if has_fragments:
            notes.append("multiple disconnected fragments (salt/counterion?)")

        centres = Chem.FindMolChiralCenters(mol, includeUnassigned=True, useLegacyImplementation=False)
        stereo_atoms = [idx for idx, chirality in centres if chirality == "?"]
        if stereo_atoms:
            notes.append(f"undefined stereocentre(s) at atom index {', '.join(str(i) for i in stereo_atoms)}")

        ionizable_atoms = {}
        for name, patt in patterns.items():
            if patt is None:
                continue
            matches = mol.GetSubstructMatches(patt)
            if matches:
                ionizable_atoms[name] = sorted({idx for match in matches for idx in match})
        if ionizable_atoms:
            notes.append(f"ionizable group(s) present ({', '.join(ionizable_atoms)}) -- verify the SMILES "
                          "reflects your intended protonation state")

        if notes:
            notes_by_ligand[lig.id] = {"notes": notes, "stereo_atoms": stereo_atoms,
                                        "ionizable_atoms": ionizable_atoms, "has_fragments": has_fragments}
    return notes_by_ligand


def check_ligand_preparation(campaign: Campaign) -> CheckResult:
    notes_by_ligand = _ligand_chemistry_notes(campaign)
    if notes_by_ligand:
        issues = [f"{lig_id}: {'; '.join(info['notes'])}" for lig_id, info in notes_by_ligand.items()]
        shown = issues[:5]
        more = f" (+{len(issues) - 5} more)" if len(issues) > 5 else ""
        return CheckResult("ligand_preparation", "WARN",
                            f"{len(issues)} ligand(s) may need chemistry review: {shown}{more}")
    return CheckResult("ligand_preparation", "PASS", "no stereo/protonation/fragment concerns detected")


def check_hidden_files(input_dir: Path, dry_run: bool = False) -> CheckResult:
    # exclude BoltzMaker's own manifest -- it's bookkeeping, not Finder/OS cruft.
    hidden = [p for p in input_dir.glob(".*") if p.is_file() and p.name != MANIFEST_FILENAME]
    if not hidden:
        return CheckResult("hidden_files", "PASS", "no hidden files in input dir")
    if dry_run:
        return CheckResult("hidden_files", "WARN", f"{len(hidden)} hidden file(s) would be removed: {[p.name for p in hidden]}")
    for p in hidden:
        p.unlink()
    return CheckResult("hidden_files", "PASS", f"removed {len(hidden)} hidden file(s)")


def check_chain_id_length(campaign: Campaign) -> CheckResult:
    # Boltz stores chain names in a fixed 5-character numpy field (`Chain` dtype in its
    # own source, data/types.py) -- longer ids are silently truncated on write, then
    # crash later with a confusing KeyError deep inside Boltz's schema parser (this was
    # discovered via an actual failed run, not from reading docs).
    offenders = set()
    for fam in campaign.families:
        if len(fam.id) > 5:
            offenders.add(f"protein '{fam.id}' ({len(fam.id)} chars)")
        for pid in fam.partners:
            raw_id = campaign.partners[pid].id
            for cid in (raw_id if isinstance(raw_id, list) else [raw_id]):
                if len(str(cid)) > 5:
                    offenders.add(f"partner '{pid}' chain id '{cid}' ({len(str(cid))} chars)")
    for lig in campaign.ligands:
        if len(lig.id) > 5:
            offenders.add(f"ligand '{lig.id}' ({len(lig.id)} chars)")
    if offenders:
        return CheckResult("chain_id_length", "FAIL",
                            f"chain id(s) exceed Boltz's 5-character limit: {sorted(offenders)}")
    return CheckResult("chain_id_length", "PASS", "all chain ids <= 5 characters")


def check_duplicate_targets(manifest: list) -> CheckResult:
    stems = [t.stem for t in manifest]
    dupes = {s for s in stems if stems.count(s) > 1}
    if dupes:
        return CheckResult("duplicate_targets", "FAIL", f"duplicate target stem(s): {dupes}")
    return CheckResult("duplicate_targets", "PASS", f"{len(stems)} unique target(s)")


def check_memory_heuristic(campaign: Campaign, manifest: list, threshold: int) -> CheckResult:
    # Rough empirical heuristic, not a precise memory model: a ~1250-token 4-chain
    # complex (GPCR + 3 G-protein subunits) used ~65GB RAM on a 64GB M1 Max in testing
    # and swap-thrashed for 20+ minutes with zero progress before being killed.
    from rdkit import Chem
    fam_by_id = {f.id: f for f in campaign.families}
    lig_by_id = {l.id: l for l in campaign.ligands}
    atom_cache = {}

    def ligand_atoms(lig: Ligand) -> int:
        if lig.id not in atom_cache:
            mol = Chem.MolFromSmiles(lig.smiles) if lig.smiles is not None else None
            atom_cache[lig.id] = mol.GetNumAtoms() if mol else 0
        return atom_cache[lig.id]

    offenders = []
    for t in manifest:
        fam = fam_by_id.get(t.family_id)
        if fam is None:
            continue
        total = len(fam.sequence) + sum(len(campaign.partners[pid].sequence) for pid in fam.partners)
        if t.ligand_id is not None:  # apo (ligand-free) targets contribute no ligand atoms
            lig = lig_by_id.get(t.ligand_id)
            if lig is not None:
                total += ligand_atoms(lig)
        if total > threshold:
            offenders.append(f"{t.stem} (~{total} tokens)")
    if offenders:
        shown = offenders[:5]
        return CheckResult(
            "memory_heuristic", "WARN",
            f"{len(offenders)} target(s) exceed ~{threshold} tokens: {shown}{'...' if len(offenders) > 5 else ''} "
            "-- consider --workers 1 --max-parallel-samples 1 and/or a lower --mps-watermark",
        )
    return CheckResult("memory_heuristic", "PASS", f"all targets under ~{threshold} tokens (empirical heuristic)")


def check_plip_env() -> CheckResult:
    # Purely informational -- cif2plip interaction analysis is optional and additive,
    # so this must never WARN/FAIL (which --strict would otherwise promote to blocking
    # an ordinary run over a feature that isn't required).
    if _plip_available():
        return CheckResult("plip_env", "PASS", "cif2plip environment found -- interaction analysis will run")
    if not _in_pixi_env():
        fix = "run `setup-plip`"
    elif shutil.which("pixi"):
        fix = "run `pixi run postinstall`"
    else:
        # A Tier B offline pack's extracted environment: no `pixi` CLI is present at
        # all (that's the point), so the fix is the same underlying pip command
        # `postinstall` itself runs, spelled out directly.
        fix = "run `python3 -m pip install --no-build-isolation plip pdb-tools`"
    return CheckResult("plip_env", "PASS", "cif2plip environment not found -- interaction analysis will be "
                        f"skipped (optional; {fix} to enable)")


# check_boltz_patches() applies as well as checks, so it must not run twice in one
# process: `all` calls it from preflight and again from run_boltz (for the entry points
# that skip preflight), and the second call would re-scan every patched file for nothing.
_PATCH_STATE: dict = {"result": None}


def _remember_patch_state(result: "CheckResult") -> "CheckResult":
    _PATCH_STATE["result"] = result
    return result


def check_boltz_patches() -> CheckResult:
    """Apply patches/apply_boltz_patches.py to the installed boltz, then verify.

    Applies rather than merely reporting, because the patches are what make an
    unattended run survive: numerical-failure containment, the MPS flush between
    targets, the NaN guards in steering, and the precision fixes. They used to be
    applied only by run_campaign.sh, so a campaign started directly with
    `BoltzMaker.py all` -- and any campaign whose environment had been rebuilt,
    since installing boltz reverts every patch -- ran unpatched with nothing but a
    WARN nobody was awake to read.

    Applying is idempotent and a no-op when they are already in place, so the cost
    on the normal path is one subprocess. Still WARN and never FAIL if it does not
    take: an unpatched boltz predicts, it just loses the containment, and refusing
    to start would turn a degraded overnight run into no run at all.

    Self-healing here follows _plip_available()'s precedent rather than adding a
    separate step, so there is one place that knows the environment must be fixed.
    """
    if _PATCH_STATE.get("result") is not None:
        return _PATCH_STATE["result"]
    script = SCRIPT_DIR / "patches" / "apply_boltz_patches.py"
    if not script.is_file():
        return _remember_patch_state(CheckResult("boltz_patches", "WARN", f"patch script not found at {script}"))
    try:
        applied = subprocess.run([sys.executable, str(script)],
                                 capture_output=True, text=True, timeout=300)
        healed = sum(1 for l in applied.stdout.splitlines() if l.strip().startswith("applied "))
        proc = subprocess.run([sys.executable, str(script), "--check"],
                              capture_output=True, text=True, timeout=60)
    except (OSError, subprocess.SubprocessError) as e:
        return _remember_patch_state(CheckResult("boltz_patches", "WARN", f"could not check boltz patches: {e}"))
    if proc.returncode == 0:
        return _remember_patch_state(CheckResult("boltz_patches", "PASS",
                           f"boltz numerical-failure patches applied (repaired {healed} this run)"
                           if healed else "boltz numerical-failure patches applied"))
    outstanding = [l.strip() for l in proc.stdout.splitlines() if "NOT APPLIED" in l or "CANNOT" in l]
    return _remember_patch_state(CheckResult("boltz_patches", "WARN",
                       f"{len(outstanding)} boltz patch(es) not applied -- one target's numerical "
                       f"failure can abort the whole batch; run `python3 {script}`"))


def run_preflight(manifest: list, output_dir: Path, campaign: Campaign, md_path: Path, strict: bool = False,
                   memory_warn_tokens: int = 1000, json_output: bool = False) -> bool:
    results = [
        check_boltz_cli(),
        check_sse_comparison(),
        check_vendor_assets(),
        check_result_packer(),
        check_gpu(),
        check_disk_space(output_dir, len(manifest)),
        check_all_materialized([md_path] + [output_dir / f"{t.stem}.yaml" for t in manifest]),
        check_yaml_validity(manifest, output_dir, campaign),
        check_smiles(campaign),
        check_ligand_preparation(campaign),
        check_chain_id_length(campaign),
        check_memory_heuristic(campaign, manifest, memory_warn_tokens),
        check_hidden_files(output_dir),
        check_duplicate_targets(manifest),
        check_plip_env(),
        check_boltz_patches(),
    ]
    worst = "PASS"
    for r in results:
        if r.status == "FAIL" or (strict and r.status == "WARN"):
            worst = "FAIL"
        elif r.status == "WARN" and worst != "FAIL":
            worst = "WARN"

    if json_output:
        print(json.dumps([asdict(r) for r in results]))
        return worst != "FAIL"

    table = Table(title=f"[bold {_RICH_BLUE}]BoltzMaker preflight[/bold {_RICH_BLUE}]",
                  box=_rich_box.ROUNDED, header_style=f"bold {_RICH_BLUE}", border_style="dim")
    table.add_column("Check")
    table.add_column("Status")
    table.add_column("Detail")
    style = {"PASS": _RICH_GREEN, "WARN": _RICH_AMBER, "FAIL": "red"}
    icon = {"PASS": "✓", "WARN": "⚠", "FAIL": "✗"}
    for r in results:
        table.add_row(r.name, f"[{style[r.status]}]{icon[r.status]} {r.status}[/{style[r.status]}]", r.message)
    Console().print(table)
    return worst != "FAIL"


# ==========================================================================
# Runner -- resume via staged symlinks, progress via filesystem polling
# ==========================================================================

def _predictions_dir_for(out_dir: Path, staging_name: str) -> Path:
    return out_dir / f"boltz_results_{staging_name}" / "predictions"


def _target_complete(pred_dir: Path, stem: str, need_affinity: bool) -> bool:
    d = pred_dir / stem
    if not d.is_dir():
        return False
    if not any(d.glob(f"{stem}_model_0.cif")):
        return False
    if not any(d.glob(f"confidence_{stem}_model_0.json")):
        return False
    if need_affinity and not any(d.glob(f"affinity_{stem}.json")):
        return False
    return True


def _partition_targets(manifest: list, pred_dir: Path):
    complete, pending = [], []
    for t in manifest:
        (complete if pred_dir and _target_complete(pred_dir, t.stem, t.needs_affinity) else pending).append(t)
    return complete, pending


def _stage_targets(yaml_dir: Path, targets: list, stage_dir: Path) -> None:
    if stage_dir.exists():
        shutil.rmtree(stage_dir)
    stage_dir.mkdir(parents=True)
    for t in targets:
        src = (yaml_dir / f"{t.stem}.yaml").resolve()
        (stage_dir / f"{t.stem}.yaml").symlink_to(src)


def _unpark_boltz_records(results_dir: Path) -> int:
    """Restore any records parked by an earlier run that was killed mid-batch.

    `results_dir` is Boltz's *internal* output root -- <out_dir>/boltz_results_<stage>,
    the parent of `predictions` -- not the --out_dir we hand the CLI. Pointing this at
    --out_dir finds no records/ directory and silently parks nothing, which looks
    exactly like success.
    """
    records_dir = results_dir / "processed" / "records"
    parked_dir = results_dir / "processed" / "records_parked"
    if not parked_dir.is_dir():
        return 0
    records_dir.mkdir(parents=True, exist_ok=True)
    restored = 0
    for rec in parked_dir.glob("*.json"):
        dest = records_dir / rec.name
        if dest.exists():
            rec.unlink()          # the live copy wins; the parked one is a stale duplicate
        else:
            rec.rename(dest)
        restored += 1
    with contextlib.suppress(OSError):
        parked_dir.rmdir()
    return restored


@contextlib.contextmanager
def _boltz_records_restricted_to(results_dir: Path, pending: list):
    """Make Boltz's manifest match exactly the targets we staged.

    Staging one YAML is not enough to make Boltz run one target. `boltz predict`
    ignores the input dir once a target has been processed before: check_inputs()
    filters the staged input out as "All inputs are already processed", then
    rebuilds the manifest from *every* record in <out_dir>/processed/records/, and
    predict iterates that manifest. So a supposedly isolated single-target retry
    silently re-runs the whole campaign in one process -- which is how a run here
    spent 14.5 hours and 20 invocations producing nothing: the first target in
    manifest order raised, and the nine queued behind it were never attempted.
    It is also what turns one missing pre_affinity_*.npz into a crash of the shared
    affinity phase, because that phase iterates the same over-broad manifest.

    Parking the other records for the duration makes the rebuilt manifest contain
    only `pending`, so isolation is real and the affinity phase only ever asks for
    files that this batch produced. Records are small JSON pointers -- the MSAs,
    processed structures and predictions they refer to are untouched, so parking
    costs nothing and is fully reversible.
    """
    records_dir = results_dir / "processed" / "records"
    parked_dir = results_dir / "processed" / "records_parked"
    _unpark_boltz_records(results_dir)  # recover from an earlier hard kill before parking again
    if not records_dir.is_dir():
        # Genuine on a first run (Boltz has processed nothing yet). If it happens on a
        # resume, the path is wrong and scoping is silently inert -- hence the _info.
        _info(f"no processed records yet at {records_dir} -- Boltz batch scoping not needed")
        yield
        return
    keep = {t.stem for t in pending}
    moved = []
    try:
        for rec in sorted(records_dir.glob("*.json")):
            if rec.stem not in keep:
                parked_dir.mkdir(parents=True, exist_ok=True)
                dest = parked_dir / rec.name
                rec.rename(dest)
                moved.append((dest, rec))
        _info(f"Boltz batch scoped to {len(keep)} target(s); {len(moved)} other record(s) "
              f"parked for the duration.")
        yield
    finally:
        for dest, orig in moved:
            if dest.exists():
                dest.rename(orig)
        with contextlib.suppress(OSError):
            parked_dir.rmdir()


def resolve_accelerator(choice: str) -> str:
    if choice != "auto":
        return choice
    try:
        import torch
        if torch.backends.mps.is_available() or torch.cuda.is_available():
            return "gpu"
    except ImportError:
        pass
    return "cpu"


# Phase-transition markers Boltz actually prints (verified against installed boltz 2.2.1
# source -- there is no diffusion/recycling step-level signal anywhere in its output, so
# this is the finest *honest* granularity available: which phase, and that phase's own
# per-target count/rate from Lightning's own reporting).
_PHASE_PATTERNS = [
    (re.compile(r"Calling MSA server for target (\S+)"), "MSA"),
    (re.compile(r"Running structure prediction for (\d+) input"), "structure prediction"),
    (re.compile(r"Running affinity prediction for (\d+) input"), "affinity prediction"),
]
_DATALOADER_RE = re.compile(r"Predicting DataLoader 0:\s*\d+%\|.*?\|\s*(\d+)/(\d+)\s*\[([^\]]*)\]")

def _historical_seconds_per_target(campaign_dir: Path, accelerator: str) -> tuple:
    """Mean seconds per completed target from this campaign's own past runs.

    Returns (seconds_per_target, number_of_runs_used), or (None, 0).

    Boltz gives no usable signal inside a single target -- its progress bar counts
    dataloader items, and one target is one item -- so there is nothing to
    extrapolate from within a run until a target finishes. Past runs of the same
    campaign on the same machine are the next best thing, and the run-history file
    has been recording exactly what is needed (duration, targets completed,
    accelerator) since long before anything read it back.

    Only runs that completed at least one target are used: a run that failed in the
    first minute would otherwise drag the mean toward zero. The accelerator is
    matched because a CPU run is a different machine for these purposes.
    """
    history = campaign_dir / RUN_HISTORY_FILENAME
    if not history.is_file():
        return None, 0
    per_target = []
    try:
        for line in history.read_text().splitlines():
            if not line.strip():
                continue
            try:
                record = json.loads(line)
            except json.JSONDecodeError:
                continue
            completed = record.get("targets_completed") or 0
            # working_seconds excludes time the run spent suspended; older records
            # predate it and only have wall clock, which is the best they can offer.
            duration = record.get("working_seconds") or record.get("duration_seconds") or 0
            if completed > 0 and duration > 0 and record.get("accelerator") == accelerator:
                per_target.append(duration / completed)
    except OSError:
        return None, 0
    if not per_target:
        return None, 0
    # Median, not mean: one swap-thrashing run that took ten times as long as the
    # rest should not dominate the estimate for every run after it.
    per_target.sort()
    middle = len(per_target) // 2
    median = (per_target[middle] if len(per_target) % 2
              else (per_target[middle - 1] + per_target[middle]) / 2)
    return median, len(per_target)


# Wide enough for "structure prediction" abbreviated, narrow enough to leave the
# bar room on an 80-column terminal.
_LABEL_WIDTH = 10

# One mark, always in the same place, replacing the second spinner: two spinners
# turning in step for one process said nothing the first had not.
_STATE_GLYPHS = {
    "running": f"[{_RICH_GREEN}]\u25b6[/{_RICH_GREEN}]",
    "paused": f"[{_RICH_AMBER}]\u23f8[/{_RICH_AMBER}]",
    "stopping": f"[{_RICH_RED}]\u25a0[/{_RICH_RED}]",
}

# The phase names Boltz reports are too long for a 10-character column, and
# truncation would lose the word that distinguishes them.
_PHASE_SHORT = {
    "structure prediction": "structure",
    "affinity prediction": "affinity",
    "MSA": "MSA",
    "starting": "starting",
}


def _compact_duration(seconds: float) -> str:
    """A duration for a fixed-width column: no spaces, no more precision than the
    scale deserves. _format_duration renders "1h 13m 31s", which is eleven
    characters of which the last three are noise on an hour-long estimate."""
    seconds = int(max(0, seconds))
    h, rem = divmod(seconds, 3600)
    m, s = divmod(rem, 60)
    if h:
        return f"{h}h{m:02d}m"
    if m:
        return f"{m}m{s:02d}s"
    return f"{s}s"


def _memory_gauge(rss_gb: float, total_gb: float, width: int = 8) -> tuple:
    """A small bar for memory, and the colour to draw it in.

    Filled against the point where this machine starts to hurt, not against total
    RAM. A 4-chain GPCR complex took ~65GB on a 64GB Mac and swap-thrashed for 20
    minutes with no progress, which is the failure this gauge exists to make
    visible -- and measured against 69GB of installed memory that run would have
    shown a gauge that was merely "quite full" until the moment it died. The
    threshold is the same one the thrash warning already uses, so the gauge turns
    red at the point the log starts complaining rather than at some other number.
    """
    if total_gb <= 0:
        return "", _RICH_GREEN
    ceiling = MEMORY_THRASH_FRACTION * total_gb
    fraction = max(0.0, min(rss_gb / ceiling, 1.0)) if ceiling else 0.0
    filled = int(round(width * fraction))
    if fraction >= 0.85:
        colour = _RICH_RED
    elif fraction >= 0.60:
        colour = _RICH_AMBER
    else:
        colour = _RICH_GREEN
    return ("\u2593" * filled) + ("\u2591" * (width - filled)), colour


TARGET_MEMORY_FILE = ".boltzmaker_target_memory.jsonl"


def _record_target_memory(campaign_dir: Path, stem: str, peak_gb: float, tokens: int = 0) -> None:
    """Append one measured peak-RSS observation for a completed target.

    Preflight's size check is a hand-set token threshold, and on a real campaign it
    could not separate success from failure at all: every target sat between 1307 and
    1333 tokens, and both the ones that completed and the ones that OOM'd were inside
    that 26-token band. A measured peak, on this machine, for a target of this size is
    the thing that would actually have predicted trouble -- so record it as we go and
    let the check be derived from it rather than guessed.
    """
    try:
        record = {"target": stem, "peak_rss_gb": round(peak_gb, 2), "tokens": tokens,
                  "recorded": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
                  "total_ram_gb": round(psutil.virtual_memory().total / 1e9, 1)}
        with (campaign_dir / TARGET_MEMORY_FILE).open("a", encoding="utf-8") as handle:
            handle.write(json.dumps(record) + "\n")
    except OSError:
        pass          # a measurement is never worth failing a run over


MEMORY_THRASH_FRACTION = 0.90  # fraction of total system RAM considered "at risk of thrashing"
MEMORY_THRASH_SECONDS = 60  # how long sustained high memory must persist before warning


class _RunControls:
    """Single-keypress QUIT and PAUSE/RESUME while `boltz predict` runs.

    Pause is a real SIGSTOP of the whole Boltz process tree, not a soft flag.
    Nothing is discarded and nothing is recomputed: the process is frozen exactly
    where it stood and SIGCONT resumes it mid-diffusion. The alternative -- kill
    now, rely on `run` being idempotent, restart later -- only resumes at target
    granularity, so pausing an hour into a target would throw that hour away.

    The cost of a real stop is that a paused run keeps everything it holds: RAM,
    and the GPU allocations with it. That is the right trade for "pause while I
    need the machine for something else", and the wrong one for "pause for the
    weekend" -- so the message says which it is.

    Children are stopped before the parent and resumed in the opposite order. A
    dataloader worker that keeps running against a stopped parent can fill a pipe
    and wedge; resuming the parent first gives it somewhere to write.

    Controls are only offered on a real terminal. Under nohup, a CI log or a pipe
    there is no keyboard to read, and putting a non-tty into cbreak mode fails.
    """

    def __init__(self, proc):
        self.proc = proc
        self.paused = False
        self.quit_requested = False
        self.paused_seconds = 0.0
        self._paused_at = None
        self._stop = threading.Event()
        self._thread = None
        self._fd = None
        self._saved_termios = None

    @property
    def available(self) -> bool:
        try:
            return sys.stdin.isatty()
        except (ValueError, AttributeError):
            return False

    def start(self):
        if not self.available:
            return self
        import termios, tty
        self._fd = sys.stdin.fileno()
        try:
            self._saved_termios = termios.tcgetattr(self._fd)
            tty.setcbreak(self._fd)
        except termios.error:
            self._saved_termios = None
            return self
        self._thread = threading.Thread(target=self._read_keys, daemon=True)
        self._thread.start()
        return self

    def stop(self):
        self._stop.set()
        if self._saved_termios is not None:
            import termios
            try:
                termios.tcsetattr(self._fd, termios.TCSADRAIN, self._saved_termios)
            except termios.error:
                pass
            self._saved_termios = None

    def _read_keys(self):
        import select
        while not self._stop.is_set():
            try:
                ready, _, _ = select.select([sys.stdin], [], [], 0.25)
            except (ValueError, OSError):
                return
            if not ready:
                continue
            try:
                key = sys.stdin.read(1)
            except (ValueError, OSError):
                return
            if not key:
                continue
            key = key.lower()
            if key == "q":
                self.quit_requested = True
                return
            if key == "p":
                self.toggle_pause()

    def _tree(self):
        """The Boltz process and its children, parents last."""
        try:
            parent = psutil.Process(self.proc.pid)
        except psutil.NoSuchProcess:
            return []
        try:
            children = parent.children(recursive=True)
        except psutil.NoSuchProcess:
            children = []
        return children + [parent]

    def toggle_pause(self):
        if self.paused:
            self.resume()
        else:
            self.pause()

    def pause(self):
        if self.paused or self.proc.poll() is not None:
            return
        for process in self._tree():          # children first, parent last
            try:
                process.suspend()
            except (psutil.NoSuchProcess, psutil.AccessDenied):
                pass
        self.paused = True
        self._paused_at = time.time()

    def resume(self):
        if not self.paused:
            return
        for process in reversed(self._tree()):  # parent first, then children
            try:
                process.resume()
            except (psutil.NoSuchProcess, psutil.AccessDenied):
                pass
        self.paused = False
        if self._paused_at is not None:
            # Paused time is real time but not work time. Counting it would make the
            # measured seconds-per-target -- and so every future ETA -- wrong for
            # every run after the one that happened to be paused.
            self.paused_seconds += time.time() - self._paused_at
            self._paused_at = None

    def paused_for(self) -> float:
        return time.time() - self._paused_at if self._paused_at else 0.0

    def terminate_tree(self, timeout: float = 15.0) -> list:
        """Stop Boltz and everything it started. Returns any survivors.

        proc.terminate() signals only the process Popen started. Boltz runs
        dataloader workers as children, and with --workers 2 a plain terminate
        left them alive holding their share of RAM and the GPU -- verified
        directly: a two-process tree with the parent terminated left one child
        running. Nothing tore them down afterwards, so they leaked for the life
        of the shell.

        The child list is taken BEFORE the parent is signalled. Once the parent
        dies its children are reparented to init and are no longer reachable
        from its pid, so collecting them afterwards finds nothing.
        """
        tree = self._tree()                     # children first, parent last
        if self.paused:
            # SIGTERM is queued but not acted on by a stopped process; resume so
            # each one can actually run its handler and exit.
            self.resume()
        for process in tree:
            try:
                process.terminate()
            except (psutil.NoSuchProcess, psutil.AccessDenied):
                pass
        _gone, alive = psutil.wait_procs(tree, timeout=timeout)
        for process in alive:
            try:
                process.kill()
            except (psutil.NoSuchProcess, psutil.AccessDenied):
                pass
        _gone2, still_alive = psutil.wait_procs(alive, timeout=5)
        return still_alive


def run_boltz(yaml_dir: Path, out_dir: Path, manifest: list, workers: int, accelerator: str,
              campaign_dir: Path, limit: int = None,
              mps_watermark: float = 1.0, max_parallel_samples: int = 1,
              recycling_steps: int = None, sampling_steps: int = None, diffusion_samples: int = 1,
              diffusion_samples_affinity: int = None, sampling_steps_affinity: int = None,
              max_msa_seqs: int = None, max_retries: int = 2,
              use_potentials: bool = True, targets_per_invocation: int = 4) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)
    # Covers `run`, which skips preflight. Memoised, so `all` does not repeat the work
    # it already did. An unpatched boltz is the difference between one target's NaN
    # being contained and it aborting every target queued behind it.
    patches = check_boltz_patches()
    if patches.status != "PASS":
        _warn(f"boltz patches: {patches.message}")
    stage_dir = yaml_dir / "_stage_run"
    pred_dir = _predictions_dir_for(out_dir, stage_dir.name)

    complete, pending = _partition_targets(manifest, pred_dir)
    if limit is not None:
        pending = pending[:limit]
    if not pending:
        _ok(f"{len(complete)}/{len(manifest)} target(s) already complete, nothing to run.")
        return

    # Boltz's own "Running affinity prediction for N inputs" phase iterates over every
    # staged/processed input, not just the ones whose YAML declared a properties/affinity
    # block -- mixing apo (Ligands: none) targets into the same batch as affinity targets
    # crashes it with a FileNotFoundError loading a pre_affinity_*.npz that only ever gets
    # generated for the affinity ones (found via a real crash on examples/5ht2_gq_panel,
    # not documented Boltz behavior). Splitting into two invocations avoids it; both reuse
    # the same stage_dir/pred_dir so `find_any_predictions_dir` keeps seeing one
    # consistent output tree across both.
    groups = [g for g in (
        [t for t in pending if t.needs_affinity],
        [t for t in pending if not t.needs_affinity],
    ) if g]

    for batch in groups:
        _run_boltz_batch_with_retry(batch, len(pending), len(manifest), len(complete), yaml_dir, stage_dir,
                          pred_dir, out_dir, workers, accelerator, campaign_dir, mps_watermark,
                          max_parallel_samples, recycling_steps, sampling_steps, diffusion_samples,
                          diffusion_samples_affinity, sampling_steps_affinity, max_msa_seqs, max_retries,
                          use_potentials, targets_per_invocation)


# How long boltz may say nothing before it is treated as wedged rather than busy. The
# quietest legitimate stretch is model load plus MSA setup at the start of an invocation,
# a few minutes; a target that is genuinely sampling refreshes its progress bar
# continuously. 60 minutes is far above the former and far below a wasted night -- the
# external watchdog this replaces used 75.
_STALL_TIMEOUT_SECONDS = 3600

_RETRY_SETTLE_SECONDS = 15  # pause between attempts so the OS fully reclaims a crashed subprocess's memory


def _run_boltz_batch_with_retry(batch: list, total_pending: int, manifest_len: int, complete_len: int,
                                 yaml_dir: Path, stage_dir: Path, pred_dir: Path, out_dir: Path, workers: int,
                                 accelerator: str, campaign_dir: Path, mps_watermark: float,
                                 max_parallel_samples: int, recycling_steps: int, sampling_steps: int,
                                 diffusion_samples: int, diffusion_samples_affinity: int,
                                 sampling_steps_affinity: int, max_msa_seqs: int, max_retries: int,
                                 use_potentials: bool = True,
                                 targets_per_invocation: int = 4) -> None:
    """Runs a batch, then automatically retries any target that didn't complete.

    A real 4-target cascade on `5ht2_gq_panel` showed why this matters: an OOM during
    structure prediction for 2 of 6 large targets run together (boltz's own clean-fail
    path, not a crash) left a `pre_affinity_*.npz` missing, which then crashed the
    shared affinity phase and took 2 more already-succeeded targets down with it. The
    first attempt runs exactly as submitted (grouped, unchanged from before); each
    retry after that isolates every still-incomplete target into its own single-target
    `boltz predict` invocation, since running fewer targets per process is the mitigation
    that actually recovered from this in practice -- a crashed subprocess's MPS/Metal
    memory is only released once the process fully exits, so a short pause before each
    retry gives the OS a moment to reclaim it before the next attempt.
    """
    remaining = batch
    attempt = 0
    while remaining:
        if attempt == 0:
            # Recycle the process every `targets_per_invocation` targets. Apple's MPS
            # allocator only returns everything on process exit: measured on a live
            # campaign, driver-held memory floored at ~20GB after one target and grew
            # ~1.9GB per target thereafter, while a single target in a fresh process
            # peaked at 47.6GB of a 55.7GB ceiling. Left in one process the two meet --
            # that run had 0 out-of-memory skips in its first four targets and 3 in its
            # last four. Each fresh process starts the allocator at zero, and Boltz
            # skips targets that already have a structure, so a chunk boundary costs one
            # model load (~4 min against ~45 min of sampling) and nothing is recomputed.
            size = targets_per_invocation if targets_per_invocation > 0 else len(remaining)
            chunks = [remaining[i:i + size] for i in range(0, len(remaining), size)]
            if len(chunks) > 1:
                _info(f"running {len(remaining)} target(s) in {len(chunks)} invocation(s) of up to "
                      f"{size}, so MPS memory is released between them "
                      f"(Settings: 'Targets per invocation: 0' to disable)")
            for chunk in chunks:
                _run_boltz_batch(chunk, total_pending, manifest_len, complete_len, yaml_dir, stage_dir,
                                  pred_dir, out_dir, workers, accelerator, campaign_dir, mps_watermark,
                                  max_parallel_samples, recycling_steps, sampling_steps, diffusion_samples,
                                  diffusion_samples_affinity, sampling_steps_affinity, max_msa_seqs,
                                  use_potentials)
        else:
            _step(f"retrying {len(remaining)} incomplete target(s) in isolation "
                  f"(attempt {attempt}/{max_retries}, one at a time, {_RETRY_SETTLE_SECONDS}s pause "
                  f"before each so memory settles): {[t.stem for t in remaining]}")
            for t in remaining:
                time.sleep(_RETRY_SETTLE_SECONDS)
                _run_boltz_batch([t], total_pending, manifest_len, complete_len, yaml_dir, stage_dir,
                                  pred_dir, out_dir, workers, accelerator, campaign_dir, mps_watermark,
                                  max_parallel_samples, recycling_steps, sampling_steps, diffusion_samples,
                                  diffusion_samples_affinity, sampling_steps_affinity, max_msa_seqs,
                                  use_potentials)

        before = len(remaining)
        remaining = [t for t in remaining if not _target_complete(pred_dir, t.stem, t.needs_affinity)]
        if not remaining:
            return
        # A retry pass that completes nothing has proved the failure is deterministic, so
        # running the identical pass again just burns hours: 20 invocations over 14.5h once
        # died on the same linalg.svd error in the same target and produced not one model.
        # Stop and say so, rather than spending the remaining budget re-proving it.
        if attempt >= 1 and len(remaining) == before:
            _err(f"attempt {attempt} completed no targets at all -- all {len(remaining)} failed "
                 f"again in isolation: {[t.stem for t in remaining]}. The failure is reproducible, "
                 f"so further identical retries are not attempted. Read the newest per-target log "
                 f"under {campaign_dir} for the real error before re-running.")
            return
        attempt += 1
        if attempt > max_retries:
            _err(f"{len(remaining)} target(s) still incomplete after {max_retries} "
                 f"automatic retr{'y' if max_retries == 1 else 'ies'} -- giving up: "
                 f"{[t.stem for t in remaining]}. Re-run `run`/`all` later to try again, or investigate "
                 f"the per-target log under {campaign_dir}.")
            return


def _run_boltz_batch(pending: list, total_pending: int, manifest_len: int, complete_len: int,
                      yaml_dir: Path, stage_dir: Path, pred_dir: Path, out_dir: Path, workers: int,
                      accelerator: str, campaign_dir: Path, mps_watermark: float, max_parallel_samples: int,
                      recycling_steps: int, sampling_steps: int, diffusion_samples: int,
                      diffusion_samples_affinity: int, sampling_steps_affinity: int,
                      max_msa_seqs: int, use_potentials: bool = True) -> None:
    # Staging alone does not scope the run -- see _boltz_records_restricted_to. The
    # records live under Boltz's internal results root (pred_dir's parent), not --out_dir.
    with _boltz_records_restricted_to(pred_dir.parent, pending):
        _run_boltz_batch_body(pending, total_pending, manifest_len, complete_len, yaml_dir, stage_dir,
                               pred_dir, out_dir, workers, accelerator, campaign_dir, mps_watermark,
                               max_parallel_samples, recycling_steps, sampling_steps, diffusion_samples,
                               diffusion_samples_affinity, sampling_steps_affinity, max_msa_seqs,
                               use_potentials)


def _run_boltz_batch_body(pending: list, total_pending: int, manifest_len: int, complete_len: int,
                      yaml_dir: Path, stage_dir: Path, pred_dir: Path, out_dir: Path, workers: int,
                      accelerator: str, campaign_dir: Path, mps_watermark: float, max_parallel_samples: int,
                      recycling_steps: int, sampling_steps: int, diffusion_samples: int,
                      diffusion_samples_affinity: int, sampling_steps_affinity: int,
                      max_msa_seqs: int, use_potentials: bool = True) -> None:
    _stage_targets(yaml_dir, pending, stage_dir)
    check_hidden_files(stage_dir)

    boltz_bin = _boltz_bin()
    cmd = [
        str(boltz_bin), "predict", str(stage_dir),
        "--diffusion_samples", str(diffusion_samples), "--use_msa_server",
        "--num_workers", str(workers), "--accelerator", accelerator,
        "--out_dir", str(out_dir),
    ]
    if use_potentials:
        # FK steering plus the physical-guidance coordinate update. Boltz recommends it and
        # it is on by default, but it is also a plausible source of a diffusion trajectory
        # diverging to NaN, so it has to be switchable per campaign.
        cmd.append("--use_potentials")
    optional_flags = {
        "--max_parallel_samples": max_parallel_samples,
        "--recycling_steps": recycling_steps,
        "--sampling_steps": sampling_steps,
        "--diffusion_samples_affinity": diffusion_samples_affinity,
        "--sampling_steps_affinity": sampling_steps_affinity,
        "--max_msa_seqs": max_msa_seqs,
    }
    for flag, val in optional_flags.items():
        if val is not None:
            cmd += [flag, str(val)]

    caffeinate_bin = shutil.which("caffeinate")

    env = dict(os.environ, PYTORCH_ENABLE_MPS_FALLBACK="1")
    if mps_watermark is not None:
        # Caps MPS allocation at `mps_watermark` x the device's recommended max working
        # set -- an oversized complex then raises a clean MPS OOM error instead of
        # silently spilling into swap (this is what actually happened in testing: a
        # ~1250-token complex used ~65GB RAM on a 64GB Mac and thrashed for 20+ minutes).
        # torch's MPS allocator requires low_watermark <= high_watermark and otherwise
        # raises immediately (its own unset default low watermark is ~1.4, which is
        # *above* our default high of 1.0 -- discovered by an actual failed run, not
        # anticipated from docs alone) -- so the low watermark must always be pinned
        # to something <= mps_watermark whenever we override the high one.
        env["PYTORCH_MPS_HIGH_WATERMARK_RATIO"] = str(mps_watermark)
        env["PYTORCH_MPS_LOW_WATERMARK_RATIO"] = "0.0"

    ts = time.strftime("%Y%m%d_%H%M%S")
    log_path = campaign_dir / f"boltz_run_{ts}.log"
    log_f = log_path.open("w")
    log_f.write(f"=== Started at {time.strftime('%a %b %d %H:%M:%S %Y')} ===\n")
    log_f.write(f"=== Command: {' '.join(cmd)} ===\n")
    log_f.write(f"=== PYTORCH_MPS_HIGH_WATERMARK_RATIO={env.get('PYTORCH_MPS_HIGH_WATERMARK_RATIO', 'unset')} "
                f"PYTORCH_MPS_LOW_WATERMARK_RATIO={env.get('PYTORCH_MPS_LOW_WATERMARK_RATIO', 'unset')} ===\n")
    log_f.flush()

    _info(f"{complete_len}/{manifest_len} campaign target(s) already complete; "
          f"running {len(pending)}/{total_pending} target(s) in this batch.")
    _info(f"log -> {log_path}")

    run_start = time.time()
    # stdin=DEVNULL so the terminal belongs to the keypress reader below. Boltz does
    # not read stdin, but an inherited terminal in cbreak mode is shared state, and
    # a child that decides to read from it would swallow the control keys.
    proc = subprocess.Popen(cmd, cwd=str(yaml_dir), env=env, stdout=subprocess.PIPE,
                             stderr=subprocess.STDOUT, stdin=subprocess.DEVNULL,
                             text=True, bufsize=1)

    # macOS idle/system sleep can interrupt an in-flight Metal/MPS GPU kernel mid-
    # execution and wedge the process in an uninterruptible kernel wait, unkillable by
    # SIGTERM -- confirmed directly via `pmset -g log` showing the kernel's
    # IOPMrootDomain idle-sleep-preventer assertion toggling on/off every few seconds
    # during a real multi-hour campaign with no held assertion, correlating exactly with
    # observed stalls (a small target that had already succeeded once stalled on a
    # retry, ruling out complex size as the cause). `-w <pid>` (rather than wrapping the
    # boltz predict command directly) keeps `proc` as boltz predict's own direct
    # subprocess -- `.terminate()`/`.kill()`/PID-based memory monitoring below are
    # unaffected -- and caffeinate exits on its own once that pid exits, however it exits.
    if caffeinate_bin:
        subprocess.Popen([caffeinate_bin, "-i", "-s", "-m", "-w", str(proc.pid)])

    latest_line = {"text": ""}
    # `since` and `updated` are ours, not Boltz's. Boltz's own tqdm line is the only
    # fine-grained signal available, and for a single target its DataLoader has
    # exactly one item -- so it renders 0/1 at the start, 1/1 at the end, and
    # nothing in between. Displaying that captured string unchanged meant a bar
    # that sat at 0/1 with a frozen "[00:00<?, ?it/s]" for the entire run, which
    # reads as a hung program. These two let the display distinguish "no progress"
    # from "no news", and keep a clock that is genuinely ours ticking either way.
    phase = {"name": "starting", "done": 0, "total": 0, "rate": "",
             "since": time.time(), "updated": 0.0, "heard": time.time()}

    def _reader():
        for line in proc.stdout:
            log_f.write(line)
            log_f.flush()
            latest_line["text"] = line.rstrip()
            phase["heard"] = time.time()
            for pattern, name in _PHASE_PATTERNS:
                if pattern.search(line):
                    phase.update(name=name, done=0, total=0, rate="",
                                 since=time.time(), updated=0.0)
            m = _DATALOADER_RE.search(line)
            if m:
                phase["done"], phase["total"], phase["rate"] = int(m.group(1)), int(m.group(2)), m.group(3)
                phase["updated"] = time.time()

    reader_thread = threading.Thread(target=_reader, daemon=True)
    reader_thread.start()

    # Live memory monitor: sums RSS across the whole boltz process tree (--num_workers
    # forks matter here) so the progress bar can show real usage, and so a repeat of the
    # thrashing incident gets a loud warning instead of Marc staring at a silent hang.
    total_ram_gb = psutil.virtual_memory().total / 1e9
    # `segment_peak` is the running peak since the last target completed, so a grouped
    # batch still yields one figure per target rather than a single number for the whole
    # invocation. `seen` starts from what already exists so a resumed campaign does not
    # re-attribute old targets to this run's memory profile.
    mem_state = {"rss_gb": 0.0, "high_since": None, "peak_gb": 0.0, "segment_peak": 0.0}
    seen_complete = {t.stem for t in pending
                     if (pred_dir / t.stem / f"{t.stem}_model_0.cif").is_file()}
    stop_monitor = threading.Event()

    def _memory_monitor():
        while not stop_monitor.is_set():
            try:
                p = psutil.Process(proc.pid)
                rss = p.memory_info().rss
                for child in p.children(recursive=True):
                    try:
                        rss += child.memory_info().rss
                    except psutil.NoSuchProcess:
                        pass
                mem_state["rss_gb"] = rss / 1e9
                mem_state["peak_gb"] = max(mem_state["peak_gb"], mem_state["rss_gb"])
                mem_state["segment_peak"] = max(mem_state["segment_peak"], mem_state["rss_gb"])
                for t in pending:
                    if t.stem in seen_complete:
                        continue
                    if (pred_dir / t.stem / f"{t.stem}_model_0.cif").is_file():
                        seen_complete.add(t.stem)
                        _record_target_memory(campaign_dir, t.stem, mem_state["segment_peak"])
                        mem_state["segment_peak"] = mem_state["rss_gb"]
                if rss > MEMORY_THRASH_FRACTION * psutil.virtual_memory().total:
                    if mem_state["high_since"] is None:
                        mem_state["high_since"] = time.time()
                    elif time.time() - mem_state["high_since"] > MEMORY_THRASH_SECONDS:
                        log_f.write(
                            "=== WARNING: memory usage has been above "
                            f"{MEMORY_THRASH_FRACTION:.0%} of system RAM for "
                            f"{MEMORY_THRASH_SECONDS}+s -- likely swap-thrashing, not "
                            "genuine progress. Consider Ctrl-C and re-running with a "
                            "lower --max-parallel-samples/--workers or --mps-watermark. ===\n"
                        )
                        log_f.flush()
                        mem_state["high_since"] = time.time()  # re-arm, don't spam every tick
                else:
                    mem_state["high_since"] = None
            except psutil.NoSuchProcess:
                pass
            stop_monitor.wait(2)

    mem_thread = threading.Thread(target=_memory_monitor, daemon=True)
    mem_thread.start()

    total = len(pending)
    # Seeded before the first target finishes, when there is nothing measured yet.
    # TimeRemainingColumn used to sit here and showed "-:--:--" for the whole run,
    # because a target count only moves when a target completes -- on a single-target
    # campaign it never moves at all until the end.
    history_per_target, history_runs = _historical_seconds_per_target(campaign_dir, accelerator)
    # Short enough for a fixed-width column; where it came from is stated once here,
    # rather than re-rendered on every frame as a sentence that changes width.
    if history_per_target:
        eta_initial = f"~{_format_duration(history_per_target * total)}"
        _info(f"estimate: about {_format_duration(history_per_target * total)} for {total} "
              f"target(s), from the median of {history_runs} past "
              f"run{'s' if history_runs != 1 else ''} of this campaign on {accelerator}.")
    else:
        eta_initial = "ETA --"
        _info("no past runs of this campaign to estimate from -- the ETA appears once the "
              "first target finishes.")
    controls = _RunControls(proc).start()
    if controls.available:
        # _info is a plain print, not a rich Console, so rich markup would appear
        # literally as "[bold]q[/bold]". ANSI directly, matching the other messages.
        _info(f"controls:  {_BOLD}q{_RESET} quit (stops boltz and exits cleanly)   "
              f"{_BOLD}p{_RESET} pause / resume")
    else:
        _info("run controls need a terminal -- not available when piped or under nohup.")
    try:
        # Every measurable value is right-aligned in a fixed-width column, and only
        # the bar is elastic. Previously each field simply followed the text before
        # it, so a phase name changing length dragged the whole row sideways and no
        # two numbers ever sat under each other. Two columns are also gone: the
        # elapsed clock was rendered on both rows (identical, because both tasks
        # start together) and the memory column printed "mem:" on the phase row,
        # whose value is deliberately empty.
        with Progress(
            TextColumn("{task.fields[state]}", table_column=Column(width=1)),
            TextColumn(f"[bold {_RICH_BLUE}]{{task.description}}[/bold {_RICH_BLUE}]",
                        table_column=Column(width=_LABEL_WIDTH, no_wrap=True)),
            BarColumn(complete_style=_RICH_GREEN, finished_style=_RICH_GREEN,
                       pulse_style=_RICH_AMBER),
            TextColumn("{task.fields[count]}", table_column=Column(width=6, justify="right")),
            TextColumn(f"[{_RICH_GREEN}]{{task.fields[clock]}}[/{_RICH_GREEN}]",
                        table_column=Column(width=8, justify="right")),
            TextColumn("{task.fields[gauge]}", table_column=Column(width=9, justify="right")),
            TextColumn("{task.fields[right]}", table_column=Column(width=13, justify="right")),
        ) as progress:
            blank = {"state": "", "count": "", "clock": "", "gauge": "", "right": ""}
            outer = progress.add_task("targets", total=total,
                                       **{**blank, "state": _STATE_GLYPHS["running"],
                                          "count": f"0/{total}", "right": eta_initial})
            inner = progress.add_task("starting", total=None, **blank)
            first_done_at = None
            done_prev = 0
            stalled = {"fired": False}
            while proc.poll() is None:
                now = time.time()
                if controls.quit_requested:
                    # Same shutdown as Ctrl-C, deliberately: one path to get right.
                    raise KeyboardInterrupt

                # A wedged boltz does not exit and does not error -- it goes quiet with
                # its worker in an uninterruptible wait, holding its GPU memory, and
                # waits forever. Observed on a live campaign: 24 minutes of silence at
                # 39.8GB held with no GPU work, which would have run to morning. This
                # was the external watchdog's job; doing it here means an unattended run
                # needs nothing watching it. Killing the tree drops the targets back to
                # the retry ladder, which reruns them one at a time in fresh processes,
                # and boltz keeps whatever it had already written.
                if not stalled["fired"] and now - phase["heard"] > _STALL_TIMEOUT_SECONDS:
                    stalled["fired"] = True
                    _warn(f"boltz has produced no output for "
                          f"{_compact_duration(now - phase['heard'])} "
                          f"(limit {_compact_duration(_STALL_TIMEOUT_SECONDS)}) -- treating it as "
                          f"wedged and stopping it so the remaining target(s) can be retried")
                    controls.terminate_tree()
                done = sum(1 for t in pending if _target_complete(pred_dir, t.stem, t.needs_affinity))
                if done > done_prev:
                    # Timestamp of the first completion, so the in-run rate below
                    # measures only whole targets and not the leading MSA/setup work.
                    if first_done_at is None:
                        first_done_at = now
                    done_prev = done

                # Prefer what this run has actually measured over anything historical:
                # once a target has finished here, that is this machine, this campaign,
                # this configuration, today.
                # The estimate is short and fixed-width: its provenance was a
                # 38-character sentence in a column, which is why nothing lined up.
                # Where it came from is said once, before the bars start.
                if done > 0:
                    working = now - run_start - controls.paused_seconds - controls.paused_for()
                    measured = max(working, 1.0) / done
                    eta_str = f"~{_compact_duration(measured * (total - done))}"
                elif history_per_target:
                    eta_str = f"~{_compact_duration(history_per_target * total)}"
                else:
                    eta_str = "ETA --"

                if controls.paused:
                    state = _STATE_GLYPHS["paused"]
                    # Labelled: a bare duration in the estimate's column reads as an
                    # estimate, and this is the opposite -- time the run is not working.
                    right = (f"[{_RICH_AMBER}]held {_compact_duration(controls.paused_for())}"
                             f"[/{_RICH_AMBER}]")
                else:
                    state = _STATE_GLYPHS["running"]
                    right = eta_str

                progress.update(outer, completed=done, state=state,
                                description="paused" if controls.paused else "targets",
                                count=f"{done}/{total}",
                                clock=_compact_duration(now - run_start), right=right)

                # Boltz's own rate string is only meaningful while it is still being
                # refreshed. For a single target it is written once and never again, so
                # showing it indefinitely presents a stale snapshot as live data.
                fresh_rate = phase["rate"] and (now - phase["updated"]) < 60
                # total=None renders a pulsing bar. A determinate bar is only honest
                # when Boltz is actually reporting items, which for one target it is not.
                inner_total = phase["total"] if phase["total"] > 1 else None
                gauge, gauge_colour = _memory_gauge(mem_state["rss_gb"], total_ram_gb)
                progress.update(
                    inner, completed=phase["done"], total=inner_total,
                    description=_PHASE_SHORT.get(phase["name"], phase["name"])[:_LABEL_WIDTH],
                    count="" if inner_total is None else f"{phase['done']}/{inner_total}",
                    clock=phase["rate"].split("<")[0].strip() if fresh_rate
                          else _compact_duration(now - phase["since"]),
                    gauge=f"[{gauge_colour}]{gauge}[/{gauge_colour}]",
                    # One decimal: a 3.7GB working set rounds to 4 and a 0.4GB one to
                    # zero, which reads as "no memory in use" rather than "barely any".
                    right=f"[{gauge_colour}]{mem_state['rss_gb']:.1f}/{total_ram_gb:.0f}G[/{gauge_colour}]",
                )
                time.sleep(1)
            reader_thread.join(timeout=5)
            done = sum(1 for t in pending if _target_complete(pred_dir, t.stem, t.needs_affinity))
            progress.update(outer, completed=done, count=f"{done}/{total}",
                            clock=_compact_duration(time.time() - run_start))
    except KeyboardInterrupt:
        print()
        if controls.paused:
            _info("resuming the paused run so it can shut down cleanly...")
        _warn("stopping boltz predict and its worker processes...")
        survivors = controls.terminate_tree()
        try:
            proc.wait(timeout=5)
        except subprocess.TimeoutExpired:
            pass
        if survivors:
            _warn(f"{len(survivors)} boltz process(es) could not be stopped: "
                  f"{[p.pid for p in survivors]}")
        else:
            _ok("boltz predict stopped; no worker processes left behind.")
        raise
    finally:
        controls.stop()
        stop_monitor.set()
        run_end = time.time()
        elapsed = time.strftime('%a %b %d %H:%M:%S %Y')
        log_f.write(f"=== Finished at {elapsed} ===\n")
        log_f.close()

        # Recorded even on interrupt (partial progress is still worth knowing) -- read
        # back by write_html() to show run parameters/runtime in the summary table.
        completed_count = sum(1 for t in pending if _target_complete(pred_dir, t.stem, t.needs_affinity))
        record = {
            "started_at": time.strftime("%Y-%m-%dT%H:%M:%S", time.localtime(run_start)),
            "finished_at": time.strftime("%Y-%m-%dT%H:%M:%S", time.localtime(run_end)),
            "duration_seconds": round(run_end - run_start, 1),
            # Wall clock minus any time spent suspended. duration_seconds stays
            # honest about how long the run occupied the machine; this is the one
            # the ETA model reads, because a run that was paused for lunch is not
            # evidence that targets take an extra hour.
            "working_seconds": round(run_end - run_start - controls.paused_seconds, 1),
            "paused_seconds": round(controls.paused_seconds, 1),
            "workers": workers, "accelerator": accelerator, "mps_watermark": mps_watermark,
            "max_parallel_samples": max_parallel_samples, "recycling_steps": recycling_steps,
            "sampling_steps": sampling_steps, "diffusion_samples_affinity": diffusion_samples_affinity,
            "sampling_steps_affinity": sampling_steps_affinity, "max_msa_seqs": max_msa_seqs,
            "targets_submitted": len(pending), "targets_completed": completed_count,
            "exit_code": proc.returncode,
        }
        with (campaign_dir / RUN_HISTORY_FILENAME).open("a") as hf:
            hf.write(json.dumps(record) + "\n")

    still_missing = [t.stem for t in pending if not _target_complete(pred_dir, t.stem, t.needs_affinity)]
    _info(f"boltz predict exited with code {proc.returncode}")
    if still_missing:
        _warn(f"{len(still_missing)} target(s) did not complete: {still_missing}")
    else:
        _ok(f"all {total} submitted target(s) completed successfully.")


# ==========================================================================
# Analyzer -- generic JSON flattening so the schema can evolve between
# Boltz versions without breaking the tool.
# ==========================================================================

def _flatten_json(prefix: str, obj, out: dict) -> None:
    if isinstance(obj, dict):
        for k, v in obj.items():
            _flatten_json(f"{prefix}_{k}" if prefix else k, v, out)
    elif isinstance(obj, list):
        if not obj:
            return
        if all(isinstance(x, (int, float)) for x in obj):
            out[f"{prefix}_mean"] = sum(obj) / len(obj)
            out[f"{prefix}_min"] = min(obj)
            out[f"{prefix}_max"] = max(obj)
        elif all(isinstance(x, list) for x in obj):
            out[f"{prefix}_skipped"] = f"<{len(obj)}x{len(obj[0]) if obj[0] else 0} matrix omitted>"
        else:
            out[prefix] = json.dumps(obj)
    else:
        out[prefix] = obj


def _load_json_flat(path: Path) -> dict:
    with path.open() as f:
        data = json.load(f)
    out = {}
    _flatten_json("", data, out)
    return out


_AFFINITY_KEY_RE = re.compile(r"^affinity_pred_value(\d*)$")


def _compute_pic50_columns(flat: dict) -> dict:
    out = {}
    for k, v in list(flat.items()):
        m = _AFFINITY_KEY_RE.match(k)
        if m and isinstance(v, (int, float)):
            suffix = m.group(1)
            out[f"pIC50{'_' + suffix if suffix else ''}"] = (6 - float(v)) * 1.364
    members = [v for k, v in out.items() if re.match(r"^pIC50_\d+$", k)]
    if len(members) >= 2:
        mean = sum(members) / len(members)
        out["pIC50_ensemble_mean"] = mean
        out["pIC50_ensemble_std"] = (sum((x - mean) ** 2 for x in members) / len(members)) ** 0.5
    return out


def find_any_predictions_dir(out_dir: Path):
    if not out_dir.exists():
        return None
    for root, dirs, _files in os.walk(out_dir):
        if "predictions" in dirs:
            return Path(root) / "predictions"
    return None


_FLAG_TEMPLATES = {
    "MISSING_OUTPUTS": "prediction did not complete -- re-run this target.",
    "LOW_CONFIDENCE": "low structural confidence.",
    "HIGH_CONFIDENCE_POOR_AFFINITY": "high structural confidence but weak predicted affinity -- verify pocket/binding mode.",
    "LOW_CONFIDENCE_STRONG_AFFINITY": "strong predicted affinity but low structural confidence -- verify pose before trusting.",
    "LOW_POCKET_PLDDT": "low pLDDT near the specified pocket (approximate, complex-level proxy).",
}

# `boltz --help` imports the whole torch stack. Warm, that is well under a second;
# cold, in a just-installed environment that still has to byte-compile it, it has
# been measured at over 20s -- which is what the old timeout was, so a first run
# failed its own preflight. Generous on purpose: the check only costs this long in
# the one case where it was previously wrong.
BOLTZ_CLI_HELP_TIMEOUT = 120

LOW_CONFIDENCE_THRESHOLD = 0.5
POCKET_PLDDT_THRESHOLD = 0.7


def _flags_to_note(flags_str: str) -> str:
    if not flags_str:
        return ""
    return " ".join(_FLAG_TEMPLATES.get(p, p) for p in flags_str.split(";") if p)


def _set_flag(df: pd.DataFrame, pos: int, flag: str) -> None:
    col = df.columns.get_loc("flags")
    existing = df.iat[pos, col]
    flags = set(existing.split(";")) if existing else set()
    flags.add(flag)
    df.iat[pos, col] = ";".join(sorted(f for f in flags if f))


def apply_confidence_flags(df: pd.DataFrame) -> pd.DataFrame:
    conf_col = "confidence_score" if "confidence_score" in df.columns else ("ptm" if "ptm" in df.columns else None)
    pic50_col = "pIC50" if "pIC50" in df.columns else None

    if conf_col:
        for pos in df.index[df[conf_col] < LOW_CONFIDENCE_THRESHOLD]:
            _set_flag(df, df.index.get_loc(pos), "LOW_CONFIDENCE")

    if conf_col and pic50_col:
        valid = df[[conf_col, pic50_col]].dropna()
        if len(valid) >= 3:
            try:
                conf_tercile = pd.qcut(df[conf_col].rank(method="first"), 3, labels=["low", "mid", "high"])
                pic50_tercile = pd.qcut(df[pic50_col].rank(method="first"), 3, labels=["low", "mid", "high"])
                for pos in range(len(df)):
                    c, p = conf_tercile.iloc[pos], pic50_tercile.iloc[pos]
                    if pd.isna(c) or pd.isna(p):
                        continue
                    if c == "high" and p == "low":
                        _set_flag(df, pos, "HIGH_CONFIDENCE_POOR_AFFINITY")
                    elif c == "low" and p == "high":
                        _set_flag(df, pos, "LOW_CONFIDENCE_STRONG_AFFINITY")
            except ValueError:
                pass  # too few distinct values to tercile-split -- skip the mismatch flags

    df["notes"] = df["flags"].apply(_flags_to_note)
    return df


def apply_pocket_plddt_flag(df: pd.DataFrame, manifest: list) -> pd.DataFrame:
    pocket_targets = {t.stem for t in manifest if t.pocket_contacts_used}
    if not pocket_targets:
        return df
    proxy_col = next((c for c in ("complex_iplddt", "complex_plddt") if c in df.columns), None)
    if proxy_col is None:
        return df
    for pos in range(len(df)):
        row = df.iloc[pos]
        if row["target_id"] in pocket_targets and pd.notna(row.get(proxy_col)) and row[proxy_col] < POCKET_PLDDT_THRESHOLD:
            _set_flag(df, pos, "LOW_POCKET_PLDDT")
    df["notes"] = df["flags"].apply(_flags_to_note)
    return df


_PLIP_STATUS_FILE = "_plip_status.json"


def _plip_dir(campaign_dir: Path) -> Path:
    return campaign_dir / "boltz_plip"


def _run_cif2plip(cif_path: Path, work_dir: Path) -> subprocess.CompletedProcess:
    # cif2plip.py shells out to the bare `pdb_tidy` command by name -- it must resolve
    # via PATH, so the plip env's own bin/ has to be on PATH, not just used by absolute
    # path for the interpreter itself.
    env = os.environ.copy()
    env["PATH"] = str(_plip_python().parent) + os.pathsep + env.get("PATH", "")
    return subprocess.run(
        [str(_plip_python()), str(_plip_script()), str(cif_path), "-o", str(work_dir)],
        capture_output=True, text=True, env=env, timeout=600,
    )


def _label_plip_image(pse_path: Path, out_png: Path, residues: pd.DataFrame) -> bool:
    if not _plip_label_script().exists():
        return False
    args = [f"{r.prot_chain}:{int(r.prot_resnr)}:{r.prot_restype}"
            for r in residues.drop_duplicates().itertuples()]
    if not args:
        return False
    env = os.environ.copy()
    # A user's own ~/.pymolrc.py (e.g. a plugin unrelated to BoltzMaker) can reference
    # modules that don't exist in .plip_env and error out during startup -- override HOME
    # so PyMOL can't find one, regardless of what's on any given machine.
    env["HOME"] = str(PLIP_VENV_DIR)
    proc = subprocess.run(
        [str(_plip_python()), str(_plip_label_script()), str(pse_path), str(out_png)] + args,
        capture_output=True, text=True, env=env, timeout=120,
    )
    return proc.returncode == 0 and out_png.exists()


def _analyze_target_interactions(t: Target, cif_path: Path, campaign: Campaign, campaign_dir: Path, i: int, n: int) -> dict:
    """Runs (or reuses a cached run of) cif2plip for one target. Returns a dict with
    plip_status, per-interaction-type counts, ligand InChIKey, long-format contact rows
    (target_id already attached), and relative paths to the PNG/pse if produced."""
    empty = {"plip_status": "failed", "counts": {}, "inchikey": None, "contacts": [], "png": None, "pse": None}
    final_dir = _plip_dir(campaign_dir) / t.stem
    status_file = final_dir / _PLIP_STATUS_FILE

    if final_dir.exists() and status_file.exists():
        status = json.loads(status_file.read_text())
        contacts = []
        inter_csv = final_dir / f"{cif_path.stem}_interactions.csv"
        chosen = status.get("chosen_ligand_key")
        if inter_csv.exists() and chosen:
            idf = pd.read_csv(inter_csv)
            contacts = idf[idf["ligand"] == chosen].to_dict("records")
            for c in contacts:
                c["target_id"] = t.stem
        return {"plip_status": status["plip_status"], "counts": status.get("counts", {}),
                "inchikey": status.get("inchikey"), "contacts": contacts,
                "png": status.get("png"), "pse": status.get("pse")}

    _step(f"interaction profiling {i}/{n} ({t.stem})...")
    stage_dir = campaign_dir / f".boltz_plip_staging_{t.stem}"
    if stage_dir.exists():
        shutil.rmtree(stage_dir)
    stage_dir.mkdir(parents=True)
    try:
        proc = _run_cif2plip(cif_path, stage_dir)
    except subprocess.TimeoutExpired:
        shutil.rmtree(stage_dir, ignore_errors=True)
        return empty

    inter_csv = stage_dir / f"{cif_path.stem}_interactions.csv"
    summ_csv = stage_dir / f"{cif_path.stem}_ligand_summary.csv"
    if proc.returncode != 0 or not inter_csv.exists() or not summ_csv.exists():
        shutil.rmtree(stage_dir, ignore_errors=True)
        return empty

    inter_df = pd.read_csv(inter_csv)
    summ_df = pd.read_csv(summ_csv)

    lig = next((l for l in campaign.ligands if l.id == t.ligand_id), None)
    chosen = None
    if lig is not None and lig.smiles:
        # InChIKey first: cif2plip's own SMILES is re-derived from the 3D structure via
        # PLIP/OpenBabel, whose canonicalization scheme differs from RDKit's, so an exact
        # string match against our (RDKit-canonical) SMILES can miss even the correct
        # ligand. InChIKey is algorithm-independent -- both tools implement the same IUPAC
        # standard -- so it survives that mismatch; the SMILES check below is just a
        # fallback in case InChIKey generation fails on either side.
        lig_inchikey = _smiles_to_inchikey(lig.smiles)
        if lig_inchikey and "inchikey" in summ_df.columns:
            matches = summ_df[summ_df["inchikey"] == lig_inchikey]
            if len(matches) == 1:
                chosen = matches.iloc[0]["ligand"]
        if chosen is None and "smiles" in summ_df.columns:
            matches = summ_df[summ_df["smiles"] == lig.smiles]
            if len(matches) == 1:
                chosen = matches.iloc[0]["ligand"]
    if chosen is None and len(summ_df) == 1:
        chosen = summ_df.iloc[0]["ligand"]

    if chosen is None:
        status_str, counts, inchikey, contacts = "ambiguous_ligand", {}, None, []
        _warn(f"{t.stem} -- couldn't unambiguously match the campaign "
              f"ligand against cif2plip's detected ligands, skipping interaction analysis")
    else:
        sub = inter_df[inter_df["ligand"] == chosen]
        contacts = sub.to_dict("records")
        for c in contacts:
            c["target_id"] = t.stem
        counts = sub["interaction_type"].value_counts().to_dict()
        srow = summ_df[summ_df["ligand"] == chosen]
        inchikey = srow.iloc[0]["inchikey"] if not srow.empty and "inchikey" in srow.columns else None
        status_str = "ok" if len(sub) > 0 else "no_interactions"

    plip_subdir = stage_dir / f"{cif_path.stem}_plip"
    pngs = sorted(plip_subdir.glob("*.png")) if plip_subdir.is_dir() else []
    pses = sorted(plip_subdir.glob("*.pse")) if plip_subdir.is_dir() else []

    labeled_name = None
    if pses and status_str == "ok":
        labeled_path = plip_subdir / f"{pses[0].stem}_labeled.png"
        if _label_plip_image(pses[0], labeled_path, sub[["prot_chain", "prot_resnr", "prot_restype"]]):
            labeled_name = labeled_path.name

    png_rel = f"boltz_plip/{t.stem}/{cif_path.stem}_plip/{labeled_name or (pngs[0].name if pngs else '')}" \
        if (labeled_name or pngs) else None
    pse_rel = f"boltz_plip/{t.stem}/{cif_path.stem}_plip/{pses[0].name}" if pses else None

    final_dir.parent.mkdir(parents=True, exist_ok=True)
    if final_dir.exists():
        shutil.rmtree(final_dir)
    shutil.move(str(stage_dir), str(final_dir))
    (final_dir / _PLIP_STATUS_FILE).write_text(json.dumps({
        "plip_status": status_str, "counts": counts, "inchikey": inchikey,
        "chosen_ligand_key": chosen, "png": png_rel, "pse": pse_rel,
    }))
    return {"plip_status": status_str, "counts": counts, "inchikey": inchikey,
            "contacts": contacts, "png": png_rel, "pse": pse_rel}


def analyze(yaml_dir: Path, out_dir: Path, campaign_dir: Path,
            campaign: Campaign, skip_interactions: bool = False) -> pd.DataFrame:
    manifest = load_manifest(yaml_dir)
    pred_dir = find_any_predictions_dir(out_dir)
    cif_dst = campaign_dir / "boltz_cif"
    cif_dst.mkdir(exist_ok=True)

    run_plip = _plip_available() and not skip_interactions
    all_contacts = []
    plip_targets_done = 0
    plip_targets_total = sum(1 for t in manifest if t.ligand_id is not None and pred_dir
                              and (pred_dir / t.stem).is_dir()) if run_plip else 0

    ligand_by_id = {l.id: l for l in campaign.ligands}
    family_by_id = {f.id: f for f in campaign.families}

    rows = []
    for t in manifest:
        lig = ligand_by_id.get(t.ligand_id)
        ligand_smiles = (lig.smiles or lig.ccd) if lig else None
        fam = family_by_id.get(t.family_id)
        family_group = (fam.group if fam and fam.group else None) or t.family_id
        partner_ids = ", ".join(_partner_display_id(p) for p in fam.partners) if fam and fam.partners else ""
        display_name = _target_display_name(fam, t.ligand_id)
        row = {"target_id": t.stem, "family_id": t.family_id, "family_group": family_group,
               "partner_ids": partner_ids, "display_name": display_name, "ligand_id": t.ligand_id,
               "ligand_smiles": ligand_smiles, "ligand_role": lig.role if lig else None, "flags": ""}
        d = pred_dir / t.stem if pred_dir else None
        if not d or not d.is_dir():
            row["flags"] = "MISSING_OUTPUTS"
            rows.append(row)
            continue

        conf_files = sorted(d.glob(f"confidence_{t.stem}_model_0.json"))
        aff_files = sorted(d.glob(f"affinity_{t.stem}.json"))
        cif_files = sorted(d.glob(f"{t.stem}_model_0.cif"))

        if conf_files:
            row.update(_load_json_flat(conf_files[0]))
        if aff_files:
            aff_flat = _load_json_flat(aff_files[0])
            row.update(aff_flat)
            row.update(_compute_pic50_columns(aff_flat))
        if cif_files:
            shutil.copy2(cif_files[0], cif_dst / cif_files[0].name)
            row["cif_file"] = cif_files[0].name

            if run_plip and t.ligand_id is not None:
                plip_targets_done += 1
                result = _analyze_target_interactions(t, cif_dst / cif_files[0].name, campaign, campaign_dir,
                                                       plip_targets_done, plip_targets_total)
                row["plip_status"] = result["plip_status"]
                row["plip_png_path"] = result["png"]
                row["plip_pse_path"] = result["pse"]
                for itype, n in result["counts"].items():
                    row[f"plip_{itype.replace(' ', '_')}_count"] = n
                all_contacts.extend(result["contacts"])
            elif run_plip:
                row["plip_status"] = "not_applicable_apo"
        elif not skip_interactions and t.ligand_id is not None:
            row["plip_status"] = "skipped_no_env"

        if not conf_files or not cif_files or (t.needs_affinity and not aff_files):
            row["flags"] = "MISSING_OUTPUTS"
        rows.append(row)

    df = pd.DataFrame(rows)
    if any(c.startswith("plip_") and c.endswith("_count") for c in df.columns):
        count_cols = [c for c in df.columns if c.startswith("plip_") and c.endswith("_count")]
        df[count_cols] = df[count_cols].fillna(0).astype(int)
    df = apply_confidence_flags(df)
    df = apply_pocket_plddt_flag(df, manifest)

    if all_contacts:
        pd.DataFrame(all_contacts).to_csv(campaign_dir / "boltz_interactions.csv", index=False)

    return df


# ==========================================================================
# Report writers: CSV / XLSX / self-contained HTML dashboard
# ==========================================================================

def write_csv(df: pd.DataFrame, path: Path) -> None:
    df.to_csv(path, index=False)


def write_xlsx(df: pd.DataFrame, path: Path) -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="targets", index=False)
        has_pivot = "pIC50" in df.columns and df["family_id"].nunique() > 1
        pivot = df.pivot_table(index="ligand_id", columns="family_id", values="pIC50", aggfunc="mean") if has_pivot else None
        if pivot is not None:
            fam_label = _family_label_map(df)
            pivot = pivot.rename(columns=fam_label)
            pivot.to_excel(writer, sheet_name="selectivity")

        wb = writer.book
        ws = wb["targets"]
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True)
        if len(df) > 0:
            ws.auto_filter.ref = ws.dimensions
            for col_name in (c for c in ("confidence_score", "ptm", "iptm", "pIC50") if c in df.columns):
                col_letter = get_column_letter(df.columns.get_loc(col_name) + 1)
                ws.conditional_formatting.add(
                    f"{col_letter}2:{col_letter}{len(df) + 1}",
                    ColorScaleRule(start_type="min", start_color="F8696B",
                                   mid_type="percentile", mid_value=50, mid_color="FFEB84",
                                   end_type="max", end_color="63BE7B"),
                )
        if pivot is not None and pivot.shape[0] and pivot.shape[1]:
            ws2 = wb["selectivity"]
            rng = f"B2:{get_column_letter(1 + pivot.shape[1])}{1 + pivot.shape[0]}"
            ws2.conditional_formatting.add(
                rng, ColorScaleRule(start_type="min", start_color="F8696B",
                                     mid_type="percentile", mid_value=50, mid_color="FFEB84",
                                     end_type="max", end_color="63BE7B"),
            )


def _fig_to_base64(fig) -> str:
    buf = io.BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight", dpi=110)
    plt.close(fig)
    return base64.b64encode(buf.getvalue()).decode("ascii")


_BAR_WIDTH = 0.6
_BAR_MIN_SLOTS = 5  # reserve room for at least this many category slots, so 1-2
                     # bars don't visually dominate the whole plot area

_AXIS_LABEL_FONTSIZE = 12
_TICK_FONTSIZE = 10
_LEGEND_FONTSIZE = 10
_ANNOTATION_FONTSIZE = 9


_CHART_HEIGHT_PX = 260  # matches the .md-chart-grid img sizing these replaced


def _plotly_font() -> dict:
    return dict(family="Inter, -apple-system, BlinkMacSystemFont, sans-serif", size=_TICK_FONTSIZE)


def _plotly_to_div(fig, div_id: str) -> str:
    fig.update_layout(margin=dict(l=60, r=20, t=10, b=100), height=_CHART_HEIGHT_PX,
                       paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)", font=_plotly_font())
    # Matplotlib draws a full rectangular border (all four spines) by default -- Plotly
    # doesn't unless told to, so mirror the axis line to the opposite side to match the
    # look of the charts these replaced.
    fig.update_xaxes(showline=True, linewidth=1, linecolor="black", mirror=True)
    fig.update_yaxes(showline=True, linewidth=1, linecolor="black", mirror=True)
    return pio.to_html(fig, full_html=False, include_plotlyjs=False, div_id=div_id,
                        config={"responsive": True, "displaylogo": False})


def _make_bar_chart(df: pd.DataFrame, col: str, div_id: str):
    # No title set on the figure -- the HTML <h2> card header is the only title, so it
    # doesn't appear twice (once baked into the chart, once in the card).
    if col not in df.columns:
        return None
    d = df[["display_name", col]].dropna().sort_values(col, ascending=False)
    if d.empty:
        return None
    n = len(d)
    x = list(range(n))
    fig = go.Figure(go.Bar(x=x, y=d[col].tolist(), width=_BAR_WIDTH, marker_color="#4C72B0"))
    fig.update_xaxes(tickmode="array", tickvals=x, ticktext=d["display_name"].tolist(), tickangle=-75,
                      tickfont=dict(size=_TICK_FONTSIZE), range=[-0.75, max(n - 0.25, _BAR_MIN_SLOTS - 0.75)])
    fig.update_yaxes(title_text=col, title_font=dict(size=_AXIS_LABEL_FONTSIZE), tickfont=dict(size=_TICK_FONTSIZE))
    return _plotly_to_div(fig, div_id)


def _family_label_map(df: pd.DataFrame) -> dict:
    """Maps the internal per-variant family_id to its human-readable {group}_{partners}
    label (e.g. "H2ANG" -> "5HT2A", "5HT2A" -> "5HT2A_GNAQ+GNB1+GNG2"), derived straight
    from df's own family_group/partner_ids columns -- used for family-level axis/column
    labels (the selectivity heatmap/pivot) that must still group by the real family_id."""
    def label(r):
        return f"{r['family_group']}_{r['partner_ids'].replace(', ', '+')}" if r["partner_ids"] else r["family_group"]
    return df.drop_duplicates("family_id").set_index("family_id").apply(label, axis=1).to_dict()


def _make_selectivity_heatmap(df: pd.DataFrame):
    if "pIC50" not in df.columns or df["family_id"].nunique() < 2:
        return None
    pivot = df.pivot_table(index="ligand_id", columns="family_id", values="pIC50", aggfunc="mean")
    fam_label = _family_label_map(df)
    fig, ax = plt.subplots(figsize=(1.2 * len(pivot.columns) + 2, 0.4 * len(pivot.index) + 2))
    im = ax.imshow(pivot.values, cmap="viridis", aspect="auto")
    ax.set_xticks(range(len(pivot.columns)))
    ax.set_xticklabels([fam_label.get(c, c) for c in pivot.columns], rotation=45, ha="right",
                        fontsize=_TICK_FONTSIZE)
    ax.set_yticks(range(len(pivot.index)))
    ax.set_yticklabels(pivot.index, fontsize=_TICK_FONTSIZE)
    for i in range(len(pivot.index)):
        for j in range(len(pivot.columns)):
            v = pivot.values[i, j]
            if pd.notna(v):
                ax.text(j, i, f"{v:.1f}", ha="center", va="center", color="white", fontsize=_ANNOTATION_FONTSIZE)
    cbar = fig.colorbar(im)
    cbar.set_label("pIC50", fontsize=_AXIS_LABEL_FONTSIZE)
    cbar.ax.tick_params(labelsize=_TICK_FONTSIZE)
    fig.tight_layout()
    return _fig_to_base64(fig)


_ROLE_MARKER_SYMBOL = {"agonist": "circle", "antagonist": "diamond"}
_ROLE_MARKER_DEFAULT = "circle"

# Plotly's own legend defaults to sitting outside the plot on the right, exactly where a
# colorbar also sits by default -- the two overlap on any chart with both (confirmed
# directly: the shape legend and the colorbar text visibly collided). Placing the shape
# legend *inside* the plot area's top-left corner instead keeps it clear of the colorbar
# (which stays in its default outside-right position) without needing to shrink the plot.
_INSET_LEGEND = dict(font=dict(size=_LEGEND_FONTSIZE), x=0.01, y=0.99, xanchor="left", yanchor="top",
                      bgcolor="rgba(255,255,255,0.85)", bordercolor="#dde4ed", borderwidth=1)


def _role_groups(d: pd.DataFrame) -> list:
    """Splits rows into (legend_label, marker_symbol, sub_df) groups by ligand_role, for
    scatter plots that shape-code agonist vs antagonist. Collapses to one unlabeled group
    (no legend entries) when no target in the campaign has a Role: set, so campaigns that
    don't use that optional Ligand: field see no spurious legend split.
    """
    if "ligand_role" not in d.columns or d["ligand_role"].isna().all():
        return [(None, _ROLE_MARKER_DEFAULT, d)]
    groups = []
    for role, symbol in _ROLE_MARKER_SYMBOL.items():
        sub = d[d["ligand_role"] == role]
        if not sub.empty:
            groups.append((role.capitalize(), symbol, sub))
    other = d[~d["ligand_role"].isin(_ROLE_MARKER_SYMBOL)]
    if not other.empty:
        groups.append((None, _ROLE_MARKER_DEFAULT, other))
    return groups


def _tier_marker(values, colorscale: list, colorbar_title: str, is_first_trace: bool, symbol: str) -> dict:
    marker = dict(color=values, colorscale=colorscale, cmin=0, cmax=1, size=9, symbol=symbol,
                  line=dict(width=1, color="#333333"), showscale=is_first_trace)
    if is_first_trace:
        marker["colorbar"] = dict(title=colorbar_title, thickness=14)
    return marker


def _make_scatter(df: pd.DataFrame, div_id: str):
    conf_col = "confidence_score" if "confidence_score" in df.columns else ("ptm" if "ptm" in df.columns else None)
    if not conf_col or "pIC50" not in df.columns:
        return None
    d = df.dropna(subset=[conf_col, "pIC50"])
    if d.empty:
        return None
    fig = go.Figure()
    # Colour by confidence tier (same green/amber/red bands, and boundaries, as the
    # Summary table's shield icon) via a continuous colourscale + colorbar legend (like
    # the Family x ligand selectivity heatmap's own colorbar), shape by agonist/
    # antagonist -- so a single point answers both "how confident is this structure" and
    # "which pharmacology is this" at a glance.
    colorscale = _tier_colorscale(LOW_CONFIDENCE_THRESHOLD, CONFIDENCE_GREEN_THRESHOLD)
    for i, (label, symbol, sub) in enumerate(_role_groups(d)):
        fig.add_trace(go.Scatter(
            x=sub[conf_col], y=sub["pIC50"], mode="markers+text", text=sub["display_name"],
            textposition="top center", textfont=dict(size=_ANNOTATION_FONTSIZE),
            marker=_tier_marker(sub[conf_col], colorscale, "confidence_score", i == 0, symbol),
            name=label or "Target", showlegend=label is not None,
        ))
    fig.update_layout(legend=_INSET_LEGEND)
    fig.update_xaxes(title_text=conf_col, title_font=dict(size=_AXIS_LABEL_FONTSIZE), tickfont=dict(size=_TICK_FONTSIZE))
    fig.update_yaxes(title_text="pIC50", title_font=dict(size=_AXIS_LABEL_FONTSIZE), tickfont=dict(size=_TICK_FONTSIZE))
    return _plotly_to_div(fig, div_id)


def _make_pic50_vs_binder_chart(df: pd.DataFrame, div_id: str):
    if "pIC50" not in df.columns or "affinity_probability_binary" not in df.columns:
        return None
    d = df.dropna(subset=["pIC50", "affinity_probability_binary"])
    if d.empty:
        return None
    fig = go.Figure()
    # Colour by affinity tier (same green/amber/red bands/boundaries as the Summary
    # table's bullseye icon) via a continuous colourscale + colorbar legend, shape by
    # agonist/antagonist. Binder probability on x, pIC50 on y.
    colorscale = _tier_colorscale(AFFINITY_RED_THRESHOLD, AFFINITY_GREEN_THRESHOLD)
    for i, (label, symbol, sub) in enumerate(_role_groups(d)):
        fig.add_trace(go.Scatter(
            x=sub["affinity_probability_binary"], y=sub["pIC50"], mode="markers+text",
            text=sub["display_name"], textposition="top center", textfont=dict(size=_ANNOTATION_FONTSIZE),
            marker=_tier_marker(sub["affinity_probability_binary"], colorscale, "Binder probability", i == 0, symbol),
            name=label or "Target", showlegend=label is not None,
        ))
    fig.update_layout(legend=_INSET_LEGEND)
    fig.update_xaxes(title_text="Binder probability", title_font=dict(size=_AXIS_LABEL_FONTSIZE), tickfont=dict(size=_TICK_FONTSIZE))
    fig.update_yaxes(title_text="pIC50", title_font=dict(size=_AXIS_LABEL_FONTSIZE), tickfont=dict(size=_TICK_FONTSIZE))
    return _plotly_to_div(fig, div_id)


def _make_interaction_count_chart(df: pd.DataFrame, div_id: str):
    count_cols = [c for c in df.columns if c.startswith("plip_") and c.endswith("_count")]
    if not count_cols:
        return None
    d = df[["display_name"] + count_cols].fillna(0)
    if d[count_cols].to_numpy().sum() == 0:
        return None
    n = len(d)
    x = list(range(n))
    fig = go.Figure()
    for col in count_cols:
        label = col[len("plip_"):-len("_count")]
        fig.add_trace(go.Bar(x=x, y=d[col].tolist(), width=_BAR_WIDTH, name=label))
    fig.update_layout(barmode="stack", legend=dict(font=dict(size=_LEGEND_FONTSIZE)))
    fig.update_xaxes(tickmode="array", tickvals=x, ticktext=d["display_name"].tolist(), tickangle=-75,
                      tickfont=dict(size=_TICK_FONTSIZE), range=[-0.75, max(n - 0.25, _BAR_MIN_SLOTS - 0.75)])
    fig.update_yaxes(title_text="interactions", title_font=dict(size=_AXIS_LABEL_FONTSIZE), tickfont=dict(size=_TICK_FONTSIZE))
    return _plotly_to_div(fig, div_id)


def _make_fingerprint_heatmaps(df: pd.DataFrame, interactions_df) -> list:
    # One heatmap per protein family with interaction data, even a single ligand -- a
    # lone ligand's contacted-residue row is still useful (shows what it touches), it
    # just can't be compared/reordered against anything else. Binary ligand x
    # contacted-residue matrix; with >=3 ligands the ligand axis is reordered by
    # Jaccard-distance hierarchical clustering so ligands with a similar interaction
    # pattern group together (SAR ranking) -- guards the all-zero-row case (two ligands
    # sharing zero contacted residues -> Jaccard distance is 0/0 -> NaN).
    if interactions_df is None or interactions_df.empty:
        return []
    target_meta = df.set_index("target_id")[["family_id", "ligand_id"]]
    merged = interactions_df.merge(target_meta, left_on="target_id", right_index=True, how="left")
    fam_label = _family_label_map(df)

    results = []
    for family_id, fam_df in merged.groupby("family_id"):
        fam_df = fam_df.copy()
        fam_df["residue"] = fam_df["prot_restype"].astype(str) + fam_df["prot_resnr"].astype(str)
        pivot = fam_df.pivot_table(index="ligand_id", columns="residue", values="interaction_type",
                                    aggfunc="count", fill_value=0)
        pivot = (pivot > 0).astype(int)
        if pivot.shape[0] < 1 or pivot.shape[1] < 1:
            continue

        if pivot.shape[0] >= 3:
            import numpy as np
            from scipy.cluster.hierarchy import leaves_list, linkage
            from scipy.spatial.distance import pdist
            dist = np.nan_to_num(pdist(pivot.values.astype(float), metric="jaccard"))
            order = leaves_list(linkage(dist, method="average"))
            pivot = pivot.iloc[order]

        div_id = f"chart-fingerprint-{re.sub(r'[^a-zA-Z0-9_-]', '_', str(family_id))}"
        # Hard-step red/green colorscale (not a gradient) since z is strictly binary --
        # same red/green hex pair used for flags elsewhere on the dashboard (_make_scatter).
        fig = go.Figure(go.Heatmap(z=pivot.values, x=list(pivot.columns), y=list(pivot.index),
                                    colorscale=[[0, "#d62728"], [0.5, "#d62728"], [0.5, "#2ca02c"], [1, "#2ca02c"]],
                                    zmin=0, zmax=1, showscale=False, xgap=2, ygap=2))
        # Heatmaps don't get a named-trace legend on their own -- two invisible marker
        # traces populate one, matching the interaction-counts chart's legend styling.
        fig.add_trace(go.Scatter(x=[None], y=[None], mode="markers",
                                  marker=dict(size=10, color="#2ca02c"), name="Interacting"))
        fig.add_trace(go.Scatter(x=[None], y=[None], mode="markers",
                                  marker=dict(size=10, color="#d62728"), name="Non-interacting"))
        fig.update_layout(legend=dict(font=dict(size=_LEGEND_FONTSIZE)))
        fig.update_xaxes(tickangle=-90, tickfont=dict(size=_TICK_FONTSIZE))
        fig.update_yaxes(tickfont=dict(size=_TICK_FONTSIZE), autorange="reversed")
        results.append((fam_label.get(family_id, family_id), _plotly_to_div(fig, div_id)))
    return results


# ==========================================================================
# Ligand grid panel -- paginated 5x5 rendered-structure grid, building on the
# design of github.com/bellcheddar/smiles2grid (same author): per-cell 2D
# depiction + descriptors, with shared-substructure highlighting across the
# set. Adapted for a single campaign's typical scale (a handful to a few dozen
# ligands, often close SAR analogues) rather than smiles2grid's screening-scale
# use case, and wired into this campaign's own ligand-preparation findings.
# ==========================================================================

_LIGAND_GRID_PAGE_SIZE = 25  # 5x5, matching smiles2grid's page convention
_LIGAND_GRID_MIN_SCAFFOLD_ATOMS = 8  # suppress "they all contain a benzene ring"-style trivia
_LIGAND_GRID_MCS_SIMILARITY_THRESHOLD = 0.6
_LIGAND_GRID_IMG_SIZE = (260, 190)
_LIGAND_GRID_STEREO_COLOR = (0.847, 0.106, 0.549)     # magenta -- undefined stereocentre
_LIGAND_GRID_IONIZABLE_COLOR = (0.961, 0.620, 0.043)  # amber -- ionizable group
_LIGAND_GRID_FRAGMENT_COLOR = (0.70, 0.10, 0.10)      # red -- salt/counterion badge only (not atom-highlighted)
_LIGAND_GRID_CLUSTER_PALETTE = [                      # colour-blind-safe qualitative palette (Okabe-Ito derived)
    (0.000, 0.447, 0.698), (0.902, 0.624, 0.000), (0.000, 0.620, 0.451),
    (0.800, 0.475, 0.655), (0.835, 0.369, 0.000), (0.337, 0.706, 0.914),
]
_LIGAND_GRID_BADGE_LABELS = {
    "carboxylic acid": "A", "primary/secondary amine": "N", "phenol": "Ph", "sulfonic acid": "SO3",
}


def _rgb_css(color: tuple) -> str:
    return f"rgb({round(color[0] * 255)},{round(color[1] * 255)},{round(color[2] * 255)})"


def _rgb_hex(color: tuple) -> str:
    return f"#{round(color[0] * 255):02x}{round(color[1] * 255):02x}{round(color[2] * 255):02x}"


def _cluster_ligands_by_scaffold(mols_by_ligand: dict) -> list:
    # Two tiers, both defensible without a single hand-tuned "looks similar enough" call
    # driving what gets highlighted:
    #  1. Exact Bemis-Murcko scaffold match -- threshold-free, the dominant real case for
    #     an SAR series of close analogues sharing one core.
    #  2. For ligands left over, group by Morgan/Tanimoto similarity (a similarity
    #     *decision* is unavoidable here, so it's isolated to this fallback tier only) and
    #     verify the group with a real, whole-group MCS rather than asserting similarity
    #     alone -- the MCS substructure match is what actually gets highlighted, so the
    #     claim is geometrically proven, not just scored.
    #     (A "generic scaffold" tier -- bond/atom-type-abstracted Murcko cores, matched via
    #     RDKit's query-adjustment machinery -- was tried and dropped: it didn't reliably
    #     bridge aromatic vs. Kekulized-single-bond queries across independently-built
    #     molecules in testing, so a match could silently fail. A verified whole-group MCS
    #     has no such failure mode.)
    from rdkit import Chem, DataStructs
    from rdkit.Chem import rdFingerprintGenerator, rdFMCS
    from rdkit.Chem.Scaffolds import MurckoScaffold

    clusters = []
    assigned = set()

    scaffold_by_lig = {}
    for lig_id, mol in mols_by_ligand.items():
        try:
            scaffold = MurckoScaffold.GetScaffoldForMol(mol)
        except Exception:
            continue
        if scaffold is not None and scaffold.GetNumHeavyAtoms() >= _LIGAND_GRID_MIN_SCAFFOLD_ATOMS:
            scaffold_by_lig[lig_id] = scaffold

    groups = {}
    for lig_id, scaffold in scaffold_by_lig.items():
        groups.setdefault(Chem.MolToSmiles(scaffold), []).append(lig_id)
    for members in groups.values():
        if len(members) >= 2:
            ref = scaffold_by_lig[members[0]]
            clusters.append({"level": "exact", "member_ids": members, "match_mol": ref, "template_mol": ref})
            assigned.update(members)

    remaining = [lig_id for lig_id in mols_by_ligand if lig_id not in assigned]
    fp_gen = rdFingerprintGenerator.GetMorganGenerator(radius=2, fpSize=2048)
    fps = {lig_id: fp_gen.GetFingerprint(mols_by_ligand[lig_id]) for lig_id in remaining}

    parent = {lig_id: lig_id for lig_id in remaining}

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    for i, a in enumerate(remaining):
        for b in remaining[i + 1:]:
            if DataStructs.TanimotoSimilarity(fps[a], fps[b]) >= _LIGAND_GRID_MCS_SIMILARITY_THRESHOLD:
                ra, rb = find(a), find(b)
                if ra != rb:
                    parent[ra] = rb

    fallback_groups = {}
    for lig_id in remaining:
        fallback_groups.setdefault(find(lig_id), []).append(lig_id)

    for members in fallback_groups.values():
        if len(members) < 2:
            continue
        mols = [mols_by_ligand[m] for m in members]
        try:
            res = rdFMCS.FindMCS(mols, timeout=5, ringMatchesRingOnly=True, completeRingsOnly=True)
        except Exception:
            continue
        if res.canceled or not res.smartsString:
            continue
        patt = Chem.MolFromSmarts(res.smartsString)
        if patt is None or patt.GetNumAtoms() < _LIGAND_GRID_MIN_SCAFFOLD_ATOMS:
            continue
        verified = [m for m in members if mols_by_ligand[m].HasSubstructMatch(patt)]
        if len(verified) >= 2:
            clusters.append({"level": "mcs", "member_ids": verified, "match_mol": patt, "template_mol": patt})
            assigned.update(verified)

    clusters.sort(key=lambda c: -len(c["member_ids"]))
    for i, cl in enumerate(clusters):
        cl["id"] = i
        cl["color"] = _LIGAND_GRID_CLUSTER_PALETTE[i % len(_LIGAND_GRID_CLUSTER_PALETTE)]
    return clusters


def _ligand_grid_descriptors(mol) -> dict:
    from rdkit.Chem import Crippen, Descriptors, rdMolDescriptors
    return {"mw": Descriptors.MolWt(mol), "clogp": Crippen.MolLogP(mol), "tpsa": rdMolDescriptors.CalcTPSA(mol)}


def _render_ligand_cell_image(mol, stereo_atoms: list, ionizable_atoms: dict, cluster) -> str:
    from rdkit.Chem import AllChem
    from rdkit.Chem.Draw import rdMolDraw2D

    # Scaffold-templated alignment: ligands sharing a cluster draw their common substructure
    # in the same position/orientation, so shared cores visibly line up across cells.
    aligned = False
    if cluster is not None:
        template = cluster["template_mol"]
        if not template.GetNumConformers():
            AllChem.Compute2DCoords(template)
        try:
            match = AllChem.GenerateDepictionMatching2DStructure(mol, template, refPatt=cluster["match_mol"],
                                                                   acceptFailure=True)
            aligned = bool(match)
        except Exception:
            aligned = False
    if not aligned:
        AllChem.Compute2DCoords(mol)

    # Highlight priority: a specific finding (stereocentre/ionizable atom) always wins over
    # the softer cluster-membership highlight, since it's the more actionable signal --
    # cluster colors are written first, then overwritten by finding colors.
    highlight_atoms, highlight_bonds = set(), set()
    atom_colors, bond_colors = {}, {}

    if cluster is not None:
        match = mol.GetSubstructMatch(cluster["match_mol"])
        if match:
            match_set = set(match)
            highlight_atoms |= match_set
            for idx in match_set:
                atom_colors[idx] = cluster["color"]
            for bond in mol.GetBonds():
                if bond.GetBeginAtomIdx() in match_set and bond.GetEndAtomIdx() in match_set:
                    highlight_bonds.add(bond.GetIdx())
                    bond_colors[bond.GetIdx()] = cluster["color"]

    for idx in {i for atoms in ionizable_atoms.values() for i in atoms}:
        highlight_atoms.add(idx)
        atom_colors[idx] = _LIGAND_GRID_IONIZABLE_COLOR
    for idx in stereo_atoms:
        highlight_atoms.add(idx)
        atom_colors[idx] = _LIGAND_GRID_STEREO_COLOR

    drawer = rdMolDraw2D.MolDraw2DCairo(*_LIGAND_GRID_IMG_SIZE)
    opts = drawer.drawOptions()
    opts.addStereoAnnotation = True  # draws RDKit's own (?) marker at undefined centres
    opts.padding = 0.08
    rdMolDraw2D.PrepareAndDrawMolecule(drawer, mol, highlightAtoms=list(highlight_atoms),
                                        highlightAtomColors=atom_colors, highlightBonds=list(highlight_bonds),
                                        highlightBondColors=bond_colors)
    drawer.FinishDrawing()
    return base64.b64encode(drawer.GetDrawingText()).decode("ascii")


def _render_scaffold_thumbnail(mol) -> str:
    from rdkit.Chem import AllChem
    from rdkit.Chem.Draw import rdMolDraw2D
    if not mol.GetNumConformers():
        AllChem.Compute2DCoords(mol)
    drawer = rdMolDraw2D.MolDraw2DCairo(90, 70)
    drawer.drawOptions().padding = 0.1
    rdMolDraw2D.PrepareAndDrawMolecule(drawer, mol)
    drawer.FinishDrawing()
    return base64.b64encode(drawer.GetDrawingText()).decode("ascii")


def _compute_ligand_grid_cells(campaign: Campaign):
    # Shared by the HTML panel and the PDF export -- both renderers get the same
    # per-ligand data (rendered structure, severity, badges, cluster) so they can never
    # show different content, and the RDKit clustering/rendering work only happens once.
    from rdkit import Chem

    ccd_ligands = [lig for lig in campaign.ligands if lig.smiles is None]
    mols_by_ligand = {}
    for lig in campaign.ligands:
        if lig.smiles is None:
            continue
        mol = Chem.MolFromSmiles(lig.smiles)
        if mol is not None:
            mols_by_ligand[lig.id] = mol

    if not mols_by_ligand:
        return [], [], ccd_ligands, 0  # nothing renderable (all-CCD campaign)

    findings = _ligand_chemistry_notes(campaign)
    clusters = _cluster_ligands_by_scaffold(mols_by_ligand) if len(mols_by_ligand) >= 2 else []
    cluster_by_ligand = {m: cl for cl in clusters for m in cl["member_ids"]}

    cells = []
    for lig in campaign.ligands:
        if lig.id in mols_by_ligand:
            mol = mols_by_ligand[lig.id]
            info = findings.get(lig.id, {})
            stereo_atoms = info.get("stereo_atoms", [])
            ionizable_atoms = info.get("ionizable_atoms", {})
            has_fragments = info.get("has_fragments", False)
            cluster = cluster_by_ligand.get(lig.id)

            severity = "error" if (stereo_atoms or has_fragments) else ("review" if ionizable_atoms else "clean")
            badges = []
            if stereo_atoms:
                badges.append(("S", _LIGAND_GRID_STEREO_COLOR))
            if has_fragments:
                badges.append(("salt", _LIGAND_GRID_FRAGMENT_COLOR))
            for name in ionizable_atoms:
                badges.append((_LIGAND_GRID_BADGE_LABELS.get(name, name[:2]), _LIGAND_GRID_IONIZABLE_COLOR))

            img_b64 = _render_ligand_cell_image(mol, stereo_atoms, ionizable_atoms, cluster)
            desc = _ligand_grid_descriptors(mol)
            cells.append({"lig_id": lig.id, "kind": "smiles", "smiles": lig.smiles, "img_b64": img_b64,
                          "desc": desc, "severity": severity, "badges": badges, "cluster": cluster,
                          "stereo_atoms": stereo_atoms, "ionizable_atoms": ionizable_atoms,
                          "has_fragments": has_fragments})
        else:
            cells.append({"lig_id": lig.id, "kind": "ccd", "ccd_code": lig.ccd or "?"})

    return cells, clusters, ccd_ligands, len(mols_by_ligand)


def _write_ligand_grid_pdf(cells: list, path: Path) -> None:
    # Prints the same cells shown in the HTML grid (same 5x5 pagination, same rendered
    # PNGs decoded and re-embedded rather than redrawn -- no duplicate RDKit work) as a
    # PDF, matching the print-oriented output github.com/bellcheddar/smiles2grid (the
    # tool this panel's design builds on) produces natively.
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.utils import ImageReader
    from reportlab.pdfgen import canvas

    smiles_cells = [c for c in cells if c["kind"] == "smiles"]
    if not smiles_cells:
        return

    page_w, page_h = A4
    margin, gap = 24, 8
    cols, rows = 5, 5
    cell_w = (page_w - 2 * margin - (cols - 1) * gap) / cols
    cell_h = (page_h - 2 * margin - (rows - 1) * gap) / rows
    severity_colors = {"error": colors.HexColor(_rgb_hex(_LIGAND_GRID_STEREO_COLOR)),
                        "review": colors.HexColor(_rgb_hex(_LIGAND_GRID_IONIZABLE_COLOR)),
                        "clean": colors.HexColor("#dde4ed")}

    c = canvas.Canvas(str(path), pagesize=A4)
    pages = [cells[i:i + rows * cols] for i in range(0, len(cells), rows * cols)]
    for page_cells in pages:
        for idx, cell in enumerate(page_cells):
            row, col = divmod(idx, cols)
            x = margin + col * (cell_w + gap)
            y = page_h - margin - (row + 1) * cell_h - row * gap

            border = (colors.HexColor(_rgb_hex(cell["cluster"]["color"])) if cell.get("cluster")
                      else severity_colors.get(cell.get("severity"), colors.HexColor("#dde4ed")))
            c.setLineWidth(1.2)
            c.setStrokeColor(border)
            c.rect(x, y, cell_w, cell_h, stroke=1, fill=0)

            c.setFont("Helvetica-Bold", 9)
            c.setFillColor(colors.black)
            c.drawCentredString(x + cell_w / 2, y + cell_h - 12, cell["lig_id"])

            if cell["kind"] == "smiles":
                img = ImageReader(io.BytesIO(base64.b64decode(cell["img_b64"])))
                iw, ih = img.getSize()
                avail_w, avail_h = cell_w - 12, cell_h - 42
                scale = min(avail_w / iw, avail_h / ih)
                draw_w, draw_h = iw * scale, ih * scale
                c.drawImage(img, x + (cell_w - draw_w) / 2, y + 24 + (avail_h - draw_h) / 2,
                            width=draw_w, height=draw_h, mask="auto")

                desc = cell["desc"]
                c.setFont("Helvetica", 6.5)
                c.drawCentredString(x + cell_w / 2, y + 12,
                                    f"MW {desc['mw']:.0f}  cLogP {desc['clogp']:.1f}  TPSA {desc['tpsa']:.0f}")
                if cell["badges"]:
                    c.setFont("Helvetica-Bold", 6.5)
                    c.setFillColor(colors.HexColor(_rgb_hex(cell["badges"][0][1])))
                    c.drawCentredString(x + cell_w / 2, y + 4, " ".join(lbl for lbl, _ in cell["badges"]))
                    c.setFillColor(colors.black)
            else:
                c.setFont("Courier-Bold", 9)
                c.drawCentredString(x + cell_w / 2, y + cell_h / 2 + 4, cell.get("ccd_code", "?"))
                c.setFont("Helvetica", 6.5)
                c.drawCentredString(x + cell_w / 2, y + cell_h / 2 - 8, "No 2D structure (CCD)")
        c.showPage()
    c.save()


def _build_ligand_grid_panel(campaign: Campaign, campaign_dir: Path) -> str:
    if not campaign.ligands:
        return ""

    cells, clusters, ccd_ligands, n_smiles = _compute_ligand_grid_cells(campaign)
    if not cells:
        return ""  # nothing renderable (all-CCD campaign) -- skip the panel entirely

    cells_html = []
    for cell in cells:
        if cell["kind"] == "smiles":
            badge_html = "".join(f"<span class='lig-badge' style='background:{_rgb_css(c)}'>{lbl}</span>"
                                  for lbl, c in cell["badges"])
            cluster, desc = cell["cluster"], cell["desc"]
            border = f"border-color:{_rgb_css(cluster['color'])};" if cluster else ""
            cells_html.append(
                f"<div class='lig-cell lig-severity-{cell['severity']}' style='{border}'>"
                f"<div class='lig-cell-header'><span>{cell['lig_id']}</span><span class='lig-badges'>{badge_html}</span></div>"
                f"<img src='data:image/png;base64,{cell['img_b64']}' alt='{cell['lig_id']} structure'>"
                f"<div class='lig-cell-desc'>MW {desc['mw']:.0f} &middot; cLogP {desc['clogp']:.1f} "
                f"&middot; TPSA {desc['tpsa']:.0f}</div></div>"
            )
        else:
            cells_html.append(
                f"<div class='lig-cell lig-cell-ccd-wrap'>"
                f"<div class='lig-cell-header'><span>{cell['lig_id']}</span></div>"
                f"<div class='lig-cell-ccd'>{cell['ccd_code']}<br><small>No 2D structure (CCD ligand)</small></div></div>"
            )

    pages = [cells_html[i:i + _LIGAND_GRID_PAGE_SIZE] for i in range(0, len(cells_html), _LIGAND_GRID_PAGE_SIZE)]
    pages_html = "".join(
        f"<div class='lig-page' data-page='{i}'{'' if i == 0 else ' hidden'}>{''.join(page)}</div>"
        for i, page in enumerate(pages))
    pager_html = ("<div id='lig-pager'><button id='lig-prev'>&lsaquo; Prev</button>"
                  "<span id='lig-pageinfo'></span><button id='lig-next'>Next &rsaquo;</button>"
                  "<button id='lig-all'>Show all</button></div>") if len(pages) > 1 else ""

    legend_items = []
    for cl in clusters:
        thumb = _render_scaffold_thumbnail(cl["match_mol"])
        label = "shared scaffold" if cl["level"] == "exact" else "shared substructure (fallback match)"
        legend_items.append(
            f"<span class='lig-legend-item'><img class='lig-legend-thumb' style='border-color:{_rgb_css(cl['color'])}' "
            f"src='data:image/png;base64,{thumb}'>{label} -- {len(cl['member_ids'])}/{n_smiles} ligands</span>")
    # Badge key: same chip styling as the actual grid cells (.lig-badge), so each
    # abbreviation shown on a ligand can be looked up directly against its meaning here.
    badge_key = [("S", _LIGAND_GRID_STEREO_COLOR, "undefined stereocentre")]
    badge_key += [(lbl, _LIGAND_GRID_IONIZABLE_COLOR, name) for name, lbl in _LIGAND_GRID_BADGE_LABELS.items()]
    badge_key.append(("salt", _LIGAND_GRID_FRAGMENT_COLOR, "salt/disconnected fragment"))
    for lbl, color, meaning in badge_key:
        legend_items.append(
            f"<span class='lig-legend-item'><span class='lig-badge' style='background:{_rgb_css(color)}'>{lbl}</span>"
            f"{meaning}</span>")
    legend_html = f"<div class='lig-grid-legend'>{''.join(legend_items)}</div>"

    if not clusters and n_smiles >= 2:
        commonality_note = ("<p>No shared scaffold or substructure detected across the set -- "
                             "ligands are structurally distinct.</p>")
    elif clusters and len(clusters[0]["member_ids"]) == n_smiles:
        commonality_note = f"<p>All {n_smiles} SMILES ligands share scaffold {clusters[0]['id'] + 1} below.</p>"
    else:
        commonality_note = ""

    ccd_note = (f"<p>{n_smiles} SMILES ligand(s) analyzed; {len(ccd_ligands)} CCD-code ligand(s) not depicted "
                "(no SMILES to render).</p>") if ccd_ligands else ""

    footnote = ("<p class='lig-footnote'>Scaffolds: Bemis-Murcko, exact match first, then Tanimoto-clustered "
                f"(Morgan r=2, 2048-bit, threshold {_LIGAND_GRID_MCS_SIMILARITY_THRESHOLD:.2f}) whole-group MCS as "
                f"a verified fallback. Minimum highlighted substructure size: {_LIGAND_GRID_MIN_SCAFFOLD_ATOMS} "
                "heavy atoms. Stereocentre/ionizable-group highlighting from this campaign's own ligand-preparation "
                "check (see above).</p>")

    download_links = []
    smiles_cells = [c for c in cells if c["kind"] == "smiles"]
    if smiles_cells:
        pdf_path = campaign_dir / "boltz_ligand_grid.pdf"
        _write_ligand_grid_pdf(cells, pdf_path)
        download_links.append(f"<a href='{pdf_path.name}' download>Download PDF</a>")

        smiles_rows = [{
            "ID": c["lig_id"], "SMILES": c["smiles"],
            "Undefined stereocentres": len(c["stereo_atoms"]),
            "Ionizable groups": "; ".join(c["ionizable_atoms"]),
            "Salts/disconnected fragments": "Yes" if c["has_fragments"] else "No",
            "MW": round(c["desc"]["mw"], 2), "cLogP": round(c["desc"]["clogp"], 2),
            "TPSA": round(c["desc"]["tpsa"], 2),
        } for c in smiles_cells]
        smiles_csv = pd.DataFrame(smiles_rows).to_csv(index=False)
        smiles_csv_b64 = base64.b64encode(smiles_csv.encode("utf-8")).decode("ascii")
        download_links.append(f"<a href='data:text/csv;base64,{smiles_csv_b64}' "
                              "download='boltz_ligands.csv'>Download SMILES</a>")
    # Same "one <p>, links joined by a middle dot" style as the Summary table's CSV links.
    download_links_html = f"<p>{' &middot; '.join(download_links)}</p>" if download_links else ""

    return (f"<div class='md-card table-card'><h2>Ligand structures</h2>{commonality_note}{ccd_note}"
            f"{legend_html}<div id='lig-grid'>{pages_html}</div>{pager_html}{download_links_html}{footnote}</div>")


# marcdeller.com brand theme (see marcs-vibe-coding skill) -- keep in sync with any
# future BoltzMaker HTML output so every generated report looks the same.
_BRAND_CSS = """
:root {
  --md-primary: #1e73be;
  --md-primary-dark: #155a9c;
  --md-primary-light: #4a9fd4;
  --md-bg: #ffffff;
  --md-bg-alt: #f4f7fb;
  --md-surface: #ffffff;
  --md-border: #dde4ed;
  --md-text: #1a1a2e;
  --md-text-muted: #6b7c93;
  --md-text-light: #ffffff;
  --md-accent-green: #00d084;
  --md-accent-orange: #ff6900;
  --md-accent-purple: #9b51e0;
  --md-accent-amber: #fcb900;
  --md-shadow-sm: 6px 6px 9px rgba(0,0,0,0.12);
  --md-shadow-md: 12px 12px 50px rgba(0,0,0,0.18);
  --md-radius: 8px;
  --md-radius-lg: 16px;
}
*, *::before, *::after { box-sizing: border-box; }
body {
  margin: 0;
  overflow-x: hidden;
  font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
  font-size: 15px;
  line-height: 1.6;
  color: var(--md-text);
  background: var(--md-bg-alt);
}
img, canvas { max-width: 100%; height: auto; }
.md-header {
  background: linear-gradient(135deg, var(--md-primary-dark) 0%, var(--md-primary) 100%);
  color: var(--md-text-light);
  box-shadow: var(--md-shadow-sm);
  position: sticky;
  top: 0;
  z-index: 100;
}
.md-header-inner {
  max-width: 1400px;
  margin: 0 auto;
  padding: 12px 24px;
  display: flex;
  align-items: center;
  gap: 16px;
  flex-wrap: wrap;
}
.md-header-brand a {
  display: flex;
  align-items: center;
  gap: 8px;
  color: rgba(255,255,255,0.9);
  text-decoration: none;
  font-weight: 600;
  font-size: 13px;
  letter-spacing: 0.02em;
  white-space: nowrap;
}
.md-logo-dot { width: 10px; height: 10px; background: var(--md-accent-amber); border-radius: 50%; flex-shrink: 0; }
.md-header-title { flex: 1; min-width: 0; }
.md-header-title h1 { font-size: 16px; font-weight: 700; color: #fff; margin: 0; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }
.md-header-links { display: flex; gap: 10px; flex-shrink: 0; flex-wrap: wrap; }
.md-header-links a {
  color: rgba(255,255,255,0.85);
  text-decoration: none;
  font-size: 12px;
  padding: 4px 12px;
  border: 1px solid rgba(255,255,255,0.3);
  border-radius: 20px;
  white-space: nowrap;
}
.md-main { max-width: 1400px; margin: 0 auto; padding: 24px; }
.md-card { background: var(--md-surface); border: 1px solid var(--md-border); border-radius: var(--md-radius); padding: 20px; box-shadow: var(--md-shadow-sm); margin-bottom: 24px; }
.md-card h2 { margin-top: 0; font-size: 16px; }
.md-card.table-card { overflow-x: auto; max-width: 100%; }
.md-chart-grid, .md-side-by-side { display: grid; gap: 16px; grid-template-columns: repeat(2, 1fr); margin-bottom: 24px; }
.md-side-by-side.md-side-3col { grid-template-columns: repeat(3, 1fr); }
.md-chart-grid .md-card, .md-side-by-side { margin-bottom: 0; }
.md-chart-grid .md-card img, .md-side-image img { width: 100%; height: 260px; object-fit: contain; display: block; }
.md-chart-grid .md-card-span2 { grid-column: 1 / -1; }
.md-side-table { overflow: visible; }
.md-side-table table { font-size: 10px; }
.md-side-table th, .md-side-table td { padding: 3px 6px; }
/* Each binding-site column stretches to the row's full height by default (grid items
   stretch to match the tallest cell in the row); turning each into a column flexbox and
   pushing its trailing download link down with margin-top:auto keeps every link aligned
   to the bottom of the row, even when the contacts table is much taller than the fixed
   260px image/3D-viewer columns next to it. */
.md-side-viewer, .md-side-image, .md-side-table-col { display: flex; flex-direction: column; }
.md-side-viewer .md-3dmol-viewer, .md-side-image img { flex-shrink: 0; }
.md-side-viewer p, .md-side-image p, .md-side-table-col p { margin-top: auto; }
.md-3dmol-viewer { width: 100%; height: 260px; position: relative; background: #fff; border-radius: var(--md-radius); }
table { border-collapse: collapse; font-family: 'Roboto Mono', monospace; font-size: 12px; width: 100%; max-width: 100%; }
th, td { border: 1px solid var(--md-border); padding: 5px 9px; text-align: left; white-space: nowrap; }
th { background: var(--md-bg-alt); font-weight: 600; position: sticky; top: 0; }
tr:nth-child(even) { background: var(--md-bg-alt); }
/* Full table: many columns of mostly-numeric data -- fixed layout + wrapping keeps
   the whole table within the viewport instead of forcing one-line-per-cell widths
   that overflow far past the browser width. */
.md-card.table-card table { table-layout: fixed; }
.md-card.table-card th, .md-card.table-card td { white-space: normal; word-break: break-word; font-size: 11px; }
/* Full table: grouped header (colspan) + narrow right-aligned numeric columns, so a
   campaign with 20+ raw fields still fits without the whole table needing to scroll --
   only the identity columns (family/ligand/flags) stay left-aligned and get to wrap. */
.full-table { table-layout: auto; }
.full-table th, .full-table td { text-align: right; white-space: nowrap; font-variant-numeric: tabular-nums; padding: 4px 8px; }
.full-table th.ft-group { text-align: center; background: var(--md-bg-alt); }
.full-table td:not(.ft-num), .full-table th:not(.ft-group):not(.ft-num) { text-align: left; white-space: normal; }
.full-table th.ft-group-start, .full-table td.ft-group-start { border-left: 2px solid var(--md-primary); }
.full-table tr.row-group-start td { border-top: 2px solid var(--md-primary); }
.full-table th.ft-flags, .full-table td.ft-flags { text-align: center; }
.flag-ok, .flag-warn, .flag-bad { cursor: default; }
.cell-na { color: var(--md-text-muted); font-style: italic; }
.summary-table-footer { display: flex; flex-wrap: wrap; justify-content: space-between;
  align-items: center; gap: 12px; }
.summary-table-footer p { margin: 0; }
.summary-legend { display: flex; flex-wrap: wrap; gap: 4px 14px; align-items: center;
  font-size: 12px; color: var(--md-text-muted); }
.legend-title { font-weight: 600; color: var(--md-text); }
.legend-item { display: inline-flex; align-items: center; gap: 4px; white-space: nowrap; }
.md-footer { text-align: center; padding: 24px; color: var(--md-text-muted); font-size: 13px; }
.md-footer a { color: var(--md-primary); text-decoration: none; }
.lig-grid-legend { display: flex; flex-wrap: wrap; gap: 6px 18px; margin-bottom: 14px; font-size: 12px; color: var(--md-text-muted); }
.lig-legend-item { display: inline-flex; align-items: center; gap: 6px; }
.lig-swatch { width: 13px; height: 13px; border-radius: 3px; display: inline-block; flex-shrink: 0; }
.lig-legend-thumb { width: 40px; height: 32px; object-fit: contain; border: 2px solid; border-radius: 4px; background: #fff; flex-shrink: 0; }
.lig-footnote { font-size: 11px; color: var(--md-text-muted); margin-top: 12px; }
#lig-grid { margin-bottom: 12px; }
.lig-page { display: grid; grid-template-columns: repeat(5, 1fr); gap: 10px; }
.lig-page[hidden] { display: none; }
.lig-cell { border: 2px solid var(--md-border); border-radius: var(--md-radius); padding: 8px; background: var(--md-bg-alt); display: flex; flex-direction: column; }
.lig-cell.lig-severity-error { border-color: #d81b8c; }
.lig-cell.lig-severity-review { border-color: #f59e0b; }
.lig-cell-header { display: flex; justify-content: space-between; align-items: center; font-size: 11px; font-weight: 600; margin-bottom: 4px; gap: 4px; }
.lig-badges { display: flex; gap: 3px; flex-wrap: wrap; justify-content: flex-end; }
.lig-badge { font-size: 9px; font-weight: 700; color: #fff; padding: 1px 5px; border-radius: 3px; white-space: nowrap; }
.lig-cell img { width: 100%; height: 140px; object-fit: contain; display: block; background: #fff; border-radius: 4px; }
.lig-cell-desc { font-size: 10px; color: var(--md-text-muted); text-align: center; margin-top: 4px; }
.lig-cell-ccd-wrap { align-items: center; }
.lig-cell-ccd { display: flex; flex-direction: column; align-items: center; justify-content: center; height: 140px; width: 100%; color: var(--md-text-muted); font-family: 'Roboto Mono', monospace; text-align: center; }
#lig-pager { display: flex; align-items: center; gap: 12px; margin-top: 12px; font-size: 13px; }
#lig-pager button { padding: 4px 14px; border-radius: 20px; border: 1px solid var(--md-border); background: var(--md-surface); color: var(--md-text); cursor: pointer; }
#lig-pager button:disabled { opacity: 0.4; cursor: default; }
@media (max-width: 768px) {
  .md-header-inner { padding: 10px 16px; gap: 8px; }
  .md-header-title h1 { font-size: 14px; }
  .md-header-links a { font-size: 11px; padding: 3px 8px; }
  .md-main { padding: 14px; }
  .md-card { padding: 14px; }
  .md-chart-grid, .md-side-by-side, .md-side-by-side.md-side-3col { grid-template-columns: 1fr; }
  .lig-page { grid-template-columns: repeat(2, 1fr); }
}
"""

_BRAND_HEADER = """
<header class="md-header">
  <div class="md-header-inner">
    <div class="md-header-brand">
      <a href="https://marcdeller.com" target="_blank" rel="noopener noreferrer">
        <span class="md-logo-dot"></span>
        <span class="md-logo-text">Marc C. Deller, D.Phil.</span>
      </a>
    </div>
    <div class="md-header-title"><h1>BoltzMaker Report</h1></div>
    <div class="md-header-links">
      <a href="https://marcdeller.com" target="_blank" rel="noopener noreferrer">marcdeller.com</a>
      <a href="mailto:marc@marcdeller.com">marc@marcdeller.com</a>
    </div>
  </div>
</header>
"""

_BRAND_FOOTER = """
<footer class="md-footer">
  Built with BoltzMaker by
  <a href="https://marcdeller.com" target="_blank" rel="noopener noreferrer">Marc C. Deller, D.Phil.</a>
  &middot; <a href="mailto:marc@marcdeller.com">marc@marcdeller.com</a>
</footer>
"""

# All pages are already in the DOM (base64 PNGs pre-embedded by Python); this only
# toggles `hidden`. No lazy-loading/observers needed at this dashboard's scale.
_LIGAND_GRID_PAGER_JS = """
(function () {
  var pages = Array.prototype.slice.call(document.querySelectorAll('#lig-grid .lig-page'));
  var pager = document.getElementById('lig-pager');
  if (!pages.length || !pager) return;
  var prev = document.getElementById('lig-prev');
  var next = document.getElementById('lig-next');
  var showAllBtn = document.getElementById('lig-all');
  var info = document.getElementById('lig-pageinfo');
  var cur = 0;
  var allShown = false;
  function show(i) {
    allShown = false;
    cur = Math.max(0, Math.min(i, pages.length - 1));
    pages.forEach(function (p, idx) { p.hidden = (idx !== cur); });
    info.textContent = 'Page ' + (cur + 1) + ' / ' + pages.length;
    prev.disabled = (cur === 0);
    next.disabled = (cur === pages.length - 1);
  }
  prev.addEventListener('click', function () { show(cur - 1); });
  next.addEventListener('click', function () { show(cur + 1); });
  showAllBtn.addEventListener('click', function () {
    allShown = !allShown;
    if (allShown) {
      pages.forEach(function (p) { p.hidden = false; });
      info.textContent = 'Showing all ' + pages.length + ' pages';
      prev.disabled = true; next.disabled = true;
    } else {
      show(cur);
    }
  });
  show(0);
})();
"""


def _format_duration(seconds: float) -> str:
    seconds = int(seconds)
    h, rem = divmod(seconds, 3600)
    m, s = divmod(rem, 60)
    if h:
        return f"{h}h {m}m {s}s"
    if m:
        return f"{m}m {s}s"
    return f"{s}s"


def _partner_display_id(partner_id) -> str:
    return partner_id if isinstance(partner_id, str) else "/".join(partner_id)


def _family_display_name(fam: object) -> str:
    """Human-readable family-level label: {group}_{partners} (partners omitted when
    there are none), e.g. "5HT2A_GNAQ+GNB1+GNG2" or "5HT2A". Used everywhere a whole
    family (not a specific target) needs one label -- e.g. the residue-interaction
    fingerprint card title and the selectivity pivot's column headers -- instead of the
    internal per-variant family id (e.g. "H2ANG") BoltzMaker uses to keep with/without-
    partner and apo variants of the same receptor as distinct families under the hood.
    """
    if fam is None:
        return "?"
    group = fam.group if fam.group else fam.id
    if fam.partners:
        return f"{group}_{'+'.join(_partner_display_id(p) for p in fam.partners)}"
    return group


def _target_display_name(fam: object, ligand_id: object) -> str:
    """Human-readable, still-unique-per-target label: {group}_{partners}_{ligand-or-apo}
    -- e.g. "5HT2A_GNAQ+GNB1+GNG2_RISP", "5HT2A_RISP" (no partner), "5HT2A_apo" (no
    ligand). Used everywhere a single target needs one display string (chart tick
    labels/legends, card titles, the campaign-summary target list) instead of the
    internal disambiguation stem (e.g. "H2ANG_RISP") BoltzMaker uses internally.
    """
    return f"{_family_display_name(fam)}_{ligand_id if ligand_id else 'apo'}"


def _build_campaign_summary(campaign: Campaign, campaign_dir: Path) -> list:
    # Three columns: Field/Value stay short and scannable, Details carries everything
    # that would otherwise clutter Value -- ids, lengths, pointers to other cards, and a
    # plain-English gloss for the more cryptic run-parameter names.
    targets = _expand_targets(campaign)
    rows = []

    if campaign.source_path:
        fname = campaign.source_path.name
        rows.append(("Input file", fname, f"<a href='{fname}'>{fname}</a>"))
    else:
        rows.append(("Input file", "n/a", ""))

    fam_details = "; ".join(f"{f.id} ({len(f.sequence)} aa)" for f in campaign.families)
    rows.append(("Proteins", str(len(campaign.families)), fam_details))

    partner_list = list(campaign.partners.values())
    partner_details = "; ".join(f"{_partner_display_id(p.id)} ({p.type}, {len(p.sequence)} aa)"
                                 for p in partner_list) if partner_list else "none"
    rows.append(("Partners", str(len(partner_list)), partner_details))

    lig_details = "; ".join(f"{l.id} ({'SMILES' if l.smiles else f'CCD {l.ccd}'})" for l in campaign.ligands)
    rows.append(("Ligands", str(len(campaign.ligands)), lig_details))

    target_stems = ", ".join(_target_stem(fam, lig, code) for fam, lig, code in targets)
    rows.append(("Targets (protein x ligand)", str(len(targets)), target_stems))

    aff_detail = ("pIC50 predicted for every target" if campaign.settings.predict_affinity
                  else "structure only -- no affinity model run")
    rows.append(("Predict affinity", "yes" if campaign.settings.predict_affinity else "no", aff_detail))

    lig_notes = _ligand_chemistry_notes(campaign)
    if lig_notes:
        rows.append(("Ligand chemistry", f"{len(lig_notes)} of {len(campaign.ligands)} flagged",
                     f'{", ".join(lig_notes)} -- see "Ligand preparation" below'))
    else:
        rows.append(("Ligand chemistry", "clean", "no stereo/protonation/fragment concerns detected"))

    history_path = campaign_dir / RUN_HISTORY_FILENAME
    if history_path.exists():
        records = [json.loads(line) for line in history_path.read_text().splitlines() if line.strip()]
        if records:
            total_duration = sum(r.get("duration_seconds", 0) for r in records)
            last = records[-1]
            invocation_detail = (f"across {len(records)} run invocations" if len(records) > 1
                                  else "single run invocation")
            rows.append(("Boltz predict runtime", _format_duration(total_duration), invocation_detail))
            rows.append(("Accelerator", str(last.get("accelerator", "n/a")),
                         "gpu = Metal/CUDA backend used; cpu = no GPU available"))
            rows.append(("Workers", str(last.get("workers", "n/a")),
                         "parallel data-loading workers (Boltz's own default is 2)"))
            rows.append(("MPS watermark", str(last.get("mps_watermark", "n/a")),
                         "PYTORCH_MPS_HIGH_WATERMARK_RATIO cap -- lower avoids swap on Apple unified memory"))
            rows.append(("Max parallel samples", str(last.get("max_parallel_samples", "n/a")),
                         "Boltz's own --max_parallel_samples"))
            param_details = {
                "recycling_steps": "structure-refinement recycling iterations",
                "sampling_steps": "diffusion sampling steps for structure prediction",
                "diffusion_samples_affinity": "independent affinity-model ensemble members",
                "sampling_steps_affinity": "diffusion sampling steps for the affinity model",
                "max_msa_seqs": "cap on MSA sequences used for co-evolution features",
            }
            for key, label in (("recycling_steps", "Recycling steps"), ("sampling_steps", "Sampling steps"),
                               ("diffusion_samples_affinity", "Diffusion samples (affinity)"),
                               ("sampling_steps_affinity", "Sampling steps (affinity)"),
                               ("max_msa_seqs", "Max MSA sequences")):
                if last.get(key) is not None:
                    rows.append((label, str(last[key]), param_details.get(key, "")))
    return rows


_FULL_TABLE_HIDE_PATTERNS = [
    # Regex, not a fixed list -- so per-chain/per-pair columns are hidden regardless of
    # how many chains a campaign has (a hardcoded 2-chain list previously leaked
    # chains_ptm_2 and all six pair_chains_iptm_*_2/2_* columns for a 3-chain campaign).
    # Everything hidden here is either a raw concatenation of columns already shown, an
    # ensemble sub-model's individual number (the ensemble/primary value is shown
    # instead), a granular per-chain(-pair) breakdown of an interface summary number
    # already shown, or an internal file path already surfaced elsewhere in the
    # dashboard. All of it remains in the full "Download CSV" export.
    r"^target_id$", r"^ligand_smiles$", r"^notes$",
    r"^complex_iplddt$", r"^complex_pde$", r"^complex_ipde$",
    r"^chains_ptm_\d+$", r"^pair_chains_iptm_\d+_\d+$",
    r".*_path$", r"^plip_status$",
    r"^affinity_pred_value\d*$", r"^affinity_probability_binary[12]$",
    r"^pIC50_[12]$", r"^pIC50_ensemble_mean$",
    r"^family_group$",  # substituted in for family_id's column slot below, not its own entry
    r"^display_name$",  # composed {group}_{partners}_{ligand} label used in charts/titles,
                          # redundant here since Target/Partner/Ligand are already separate columns
    r"^ligand_role$",  # optional agonist/antagonist label, chart-only (marker shape)
]

_FULL_TABLE_RENAME = {
    "family_id": "Target", "family_group": "Target", "partner_ids": "Partner", "ligand_id": "Ligand",
    "flags": "Summary", "confidence_score": "Score", "ptm": "pTM", "iptm": "ipTM",
    "ligand_iptm": "Lig ipTM", "protein_iptm": "PPI ipTM",
    "complex_plddt": "pLDDT", "pIC50": "pIC50", "pIC50_ensemble_std": "pIC50 SD",
    "affinity_probability_binary": "Binder p", "cif_file": "CIF",
}
_FULL_TABLE_GROUPS = {
    "family_id": "Identity", "family_group": "Identity", "partner_ids": "Identity",
    "ligand_id": "Identity", "flags": "Identity",
    "confidence_score": "Confidence", "ptm": "Confidence", "iptm": "Confidence",
    "ligand_iptm": "Confidence", "protein_iptm": "Confidence", "complex_plddt": "Confidence",
    "pIC50": "Affinity", "affinity_probability_binary": "Affinity",
    "cif_file": "Structure",
}
_FULL_TABLE_GROUP_ORDER = ["Identity", "Confidence", "Affinity", "Interactions", "Structure", "Other"]
_FULL_TABLE_TEXT_COLS = {"family_id", "family_group", "partner_ids", "ligand_id", "flags"}
_PLIP_COUNT_LABELS = {
    "hydrogen_bonds": "H-bond", "hydrophobic": "Phobic", "pi_stacks": "π-stack",
    "salt_bridges": "Salt", "pi_cation": "π-cation", "halogen_bonds": "Halogen",
    "water_bridges": "Water",
}

# Columns that mean "no interface exists" / "no ligand to score" rather than a genuine
# poor result whenever the row is a ligand-free (apo) target -- shown as an explicit
# N/A (see cell_html) instead of a blank cell or a misleading 0.00.
_APO_NA_COLS = {"ligand_id", "iptm", "ligand_iptm", "protein_iptm"}

_FLAG_ICON_WARN = "&#9888;&#65039;"  # amber warning triangle -- used for advisories the two icons below don't otherwise show
_FLAG_ICON_BAD = "&#10060;"  # red cross -- prediction never completed, nothing to score

# Emoji glyphs render pre-coloured by the OS/font and can't be recoloured via CSS, so the
# bullseye (affinity) and shield (confidence) icons are small inline SVGs with
# fill/stroke="currentColor", tinted per-tier via a wrapping <span style="color:...">.
_BULLSEYE_SVG = ("<svg width='14' height='14' viewBox='0 0 24 24' fill='none' "
                 "xmlns='http://www.w3.org/2000/svg' aria-hidden='true'>"
                 "<circle cx='12' cy='12' r='9' stroke='currentColor' stroke-width='2'/>"
                 "<circle cx='12' cy='12' r='5' stroke='currentColor' stroke-width='2'/>"
                 "<circle cx='12' cy='12' r='1.5' fill='currentColor'/></svg>")
_SHIELD_SVG = ("<svg width='14' height='14' viewBox='0 0 24 24' fill='currentColor' "
               "xmlns='http://www.w3.org/2000/svg' aria-hidden='true'>"
               "<path d='M12 2 L20 5 V11 C20 16.5 16.5 20.8 12 22 C7.5 20.8 4 16.5 4 11 V5 Z'/></svg>")

_TIER_GREEN = "#00d084"
_TIER_AMBER = "#fcb900"
_TIER_RED = "#d62728"

# Boltz's own docs (docs/prediction.md) define confidence_score/ptm/plddt as [0, 1],
# higher = better, and affinity_probability_binary as a [0, 1] binder-probability with
# 0.5 as the natural binder/non-binder decision boundary -- but publish no tri-colour
# bands. These reuse BoltzMaker's own existing LOW_CONFIDENCE_THRESHOLD (0.5) as the
# confidence red/amber boundary (so the shield agrees with the LOW_CONFIDENCE flag it's
# partly replacing) and a symmetric +/-0.2 buffer around Boltz's 0.5 binder boundary for
# affinity, rather than inventing unrelated cutoffs.
CONFIDENCE_GREEN_THRESHOLD = 0.7
AFFINITY_GREEN_THRESHOLD = 0.7
AFFINITY_RED_THRESHOLD = 0.3


def _confidence_tier(score) -> tuple:
    if pd.isna(score):
        return None, "confidence score unavailable"
    if score >= CONFIDENCE_GREEN_THRESHOLD:
        return _TIER_GREEN, f"confidence_score {score:.2f} -- high"
    if score >= LOW_CONFIDENCE_THRESHOLD:
        return _TIER_AMBER, f"confidence_score {score:.2f} -- moderate"
    return _TIER_RED, f"confidence_score {score:.2f} -- low"


def _affinity_tier(prob) -> tuple:
    if pd.isna(prob):
        return None, "not applicable -- ligand-free (apo) target"
    if prob >= AFFINITY_GREEN_THRESHOLD:
        return _TIER_GREEN, f"binder probability {prob:.2f} -- likely binder"
    if prob >= AFFINITY_RED_THRESHOLD:
        return _TIER_AMBER, f"binder probability {prob:.2f} -- uncertain"
    return _TIER_RED, f"binder probability {prob:.2f} -- likely non-binder/decoy"


def _tier_colorscale(red_max: float, green_min: float) -> list:
    """A hard-step (not gradient) Plotly colorscale over [0, 1] with the same tier
    boundaries and colours as the Summary table's bullseye/shield icons -- used as a
    continuous marker colour + colorbar (a real legend for what the colour means),
    rather than the 3 flat per-point colours used elsewhere, matching the same
    duplicate-stop-for-a-hard-edge trick already used for the fingerprint heatmap.
    """
    return [
        [0.0, _TIER_RED], [red_max, _TIER_RED],
        [red_max, _TIER_AMBER], [green_min, _TIER_AMBER],
        [green_min, _TIER_GREEN], [1.0, _TIER_GREEN],
    ]


def _summary_cell_html(row) -> str:
    flags_val = row.get("flags")
    flags_str = "" if pd.isna(flags_val) else str(flags_val)
    flag_list = [f for f in flags_str.split(";") if f]

    if "MISSING_OUTPUTS" in flag_list:
        title = html.escape(_flags_to_note(flags_str), quote=True)
        return f"<span class='flag-bad' title='{title}'>{_FLAG_ICON_BAD}</span>"

    aff_color, aff_title = _affinity_tier(row.get("affinity_probability_binary"))
    conf_color, conf_title = _confidence_tier(row.get("confidence_score"))

    # Advisories not already obvious from the two tier colours alone (a confidence/
    # affinity mismatch *is* visually apparent as contrasting colours; low-pocket-pLDDT
    # is a different, pocket-local metric neither icon otherwise represents).
    other_flags = [f for f in flag_list if f != "LOW_CONFIDENCE"]
    if other_flags:
        note = " -- " + _flags_to_note(";".join(other_flags))
        aff_title += note
        conf_title += note

    def icon(svg: str, color: object, title: str) -> str:
        if color is None:
            return f"<span class='cell-na' title='{html.escape(title, quote=True)}'>N/A</span>"
        return f"<span style='color:{color}' title='{html.escape(title, quote=True)}'>{svg}</span>"

    cell = f"{icon(_BULLSEYE_SVG, aff_color, aff_title)} {icon(_SHIELD_SVG, conf_color, conf_title)}"
    if "LOW_POCKET_PLDDT" in flag_list:
        cell += (f" <span class='flag-warn' title='{html.escape(_FLAG_TEMPLATES['LOW_POCKET_PLDDT'], quote=True)}'>"
                 f"{_FLAG_ICON_WARN}</span>")
    return cell


def _build_summary_legend_html() -> str:
    def swatch(svg: str, color: str, label: str) -> str:
        return f"<span class='legend-item'><span style='color:{color}'>{svg}</span> {label}</span>"

    items = [
        swatch(_BULLSEYE_SVG, _TIER_GREEN, f"Likely binder (p &ge; {AFFINITY_GREEN_THRESHOLD:.1f})"),
        swatch(_BULLSEYE_SVG, _TIER_AMBER, f"Uncertain ({AFFINITY_RED_THRESHOLD:.1f}-{AFFINITY_GREEN_THRESHOLD:.1f})"),
        swatch(_BULLSEYE_SVG, _TIER_RED, f"Likely non-binder (p &lt; {AFFINITY_RED_THRESHOLD:.1f})"),
        swatch(_SHIELD_SVG, _TIER_GREEN, f"High confidence (&ge; {CONFIDENCE_GREEN_THRESHOLD:.1f})"),
        swatch(_SHIELD_SVG, _TIER_AMBER, f"Moderate confidence ({LOW_CONFIDENCE_THRESHOLD:.1f}-{CONFIDENCE_GREEN_THRESHOLD:.1f})"),
        swatch(_SHIELD_SVG, _TIER_RED, f"Low confidence (&lt; {LOW_CONFIDENCE_THRESHOLD:.1f})"),
    ]
    return (f"<div class='summary-legend'><span class='legend-title'>&#127919; affinity "
            f"&middot; &#128737;&#65039; confidence:</span>{''.join(items)}</div>")


def _full_table_label(col: str) -> str:
    if col in _FULL_TABLE_RENAME:
        return _FULL_TABLE_RENAME[col]
    if col.startswith("plip_") and col.endswith("_count"):
        itype = col[len("plip_"):-len("_count")]
        return _PLIP_COUNT_LABELS.get(itype, itype.replace("_", " ").title())
    return col


def _full_table_group(col: str) -> str:
    if col in _FULL_TABLE_GROUPS:
        return _FULL_TABLE_GROUPS[col]
    if col.startswith("plip_") and col.endswith("_count"):
        return "Interactions"
    return "Other"


def _resolve_summary_table_columns(df: pd.DataFrame) -> list:
    # Shared by the HTML summary table and the summary CSV export, so the two can never
    # show different columns. Regex-based hiding (not a fixed list) so per-chain/
    # per-chain-pair columns are hidden regardless of how many chains a campaign has --
    # a hardcoded 2-chain list previously leaked chains_ptm_2 and all six 3-chain
    # pair_chains_iptm_*_2/2_* columns straight into the table.
    hidden = re.compile("|".join(_FULL_TABLE_HIDE_PATTERNS))
    cols = [c for c in df.columns if not hidden.match(c)]

    # Conditional-by-content, not by config: a column that's empty/zero for every row in
    # *this* campaign (no partner chain -> protein_iptm always 0) is dropped, rather than
    # hardcoding "only show for multi-chain campaigns". Flags is the one exception: it's
    # always kept even when nothing was ever flagged, since the column now renders a
    # per-row confidence/affinity icon pair unconditionally -- see _summary_cell_html.
    kept = []
    for c in cols:
        if c == "pIC50_ensemble_std":
            continue  # merged into the pIC50 cell/column as "± SD", never its own column
        s = df[c]
        if c == "flags":
            pass
        elif c == "protein_iptm":
            if s.fillna(0).eq(0).all():
                continue
        elif c == "partner_ids":
            if s.fillna("").astype(str).str.strip().eq("").all():
                continue
        elif pd.api.types.is_numeric_dtype(s) and s.isna().all():
            continue
        kept.append(c)
    cols = kept
    # family_group substitutes for family_id's column slot (same "Target" header/position)
    # so the dashboard shows the shared receptor name (e.g. "5HT2A") rather than the
    # internal per-variant family id (e.g. "H2ANG", "H2AAP") BoltzMaker uses to keep
    # with/without-partner and apo variants as distinct targets under the hood.
    cols = ["family_group" if c == "family_id" else c for c in cols]
    cols.sort(key=lambda c: _FULL_TABLE_GROUP_ORDER.index(_full_table_group(c)))
    return cols


def write_summary_csv(df: pd.DataFrame, path: Path) -> None:
    # Mirrors the HTML "Summary table" exactly (same columns, same renamed headers) --
    # unlike the HTML cell, pIC50's ensemble stdev stays its own numeric column here
    # rather than a merged "value ± SD" string, since a CSV is for further analysis.
    cols = _resolve_summary_table_columns(df)
    out_cols = list(cols)
    if "pIC50" in out_cols and "pIC50_ensemble_std" in df.columns:
        out_cols.insert(out_cols.index("pIC50") + 1, "pIC50_ensemble_std")
    export_df = df[out_cols].copy() if out_cols else df.iloc[:, :0].copy()
    export_df.columns = [_full_table_label(c) for c in out_cols]
    export_df.to_csv(path, index=False)


def _build_full_table_html(df: pd.DataFrame) -> str:
    cols = _resolve_summary_table_columns(df)
    if not cols:
        return "<p>No columns to display.</p>"

    groups = [_full_table_group(c) for c in cols]
    has_sd = "pIC50" in cols and "pIC50_ensemble_std" in df.columns

    def group_header_row() -> str:
        cells, i = [], 0
        while i < len(groups):
            j = i
            while j < len(groups) and groups[j] == groups[i]:
                j += 1
            cells.append(f"<th colspan='{j - i}' class='ft-group'>{groups[i]}</th>")
            i = j
        return f"<tr>{''.join(cells)}</tr>"

    def column_header_row() -> str:
        cells, prev = [], None
        for c, g in zip(cols, groups):
            classes = ([] if g == prev else ["ft-group-start"])
            classes += ["ft-flags"] if c == "flags" else ([] if c in _FULL_TABLE_TEXT_COLS else ["ft-num"])
            cells.append(f"<th class='{' '.join(classes)}'>{_full_table_label(c)}</th>")
            prev = g
        return f"<tr>{''.join(cells)}</tr>"

    def cell_html(row, c: str) -> str:
        v = row[c]
        if c == "cif_file":
            return "" if pd.isna(v) else f"<a href='boltz_cif/{v}'>CIF</a>"
        if c == "flags":
            return _summary_cell_html(row)
        is_apo = pd.isna(row.get("ligand_id"))
        if is_apo and (c in _APO_NA_COLS or _full_table_group(c) == "Affinity"
                       or (c.startswith("plip_") and c.endswith("_count"))):
            return "<span class='cell-na' title='not applicable -- ligand-free (apo) target'>N/A</span>"
        if c in _FULL_TABLE_TEXT_COLS:
            return "" if pd.isna(v) else str(v)
        if pd.isna(v):
            return ""
        text = f"{v:.2f}" if isinstance(v, float) else str(v)
        if c == "pIC50" and has_sd and not pd.isna(row.get("pIC50_ensemble_std")):
            text += f" ± {row['pIC50_ensemble_std']:.2f}"
        return text

    def body_row(row, is_new_family_group: bool) -> str:
        cells, prev = [], None
        for c, g in zip(cols, groups):
            classes = ([] if g == prev else ["ft-group-start"])
            classes += ["ft-flags"] if c == "flags" else ([] if c in _FULL_TABLE_TEXT_COLS else ["ft-num"])
            cells.append(f"<td class='{' '.join(classes)}'>{cell_html(row, c)}</td>")
            prev = g
        tr_class = " class='row-group-start'" if is_new_family_group else ""
        return f"<tr{tr_class}>{''.join(cells)}</tr>"

    # Rows already come out of the manifest grouped by family (each family's targets are
    # contiguous), so this only needs to notice when family_group changes between
    # consecutive rows -- no re-sorting -- and mark that row with a top border, the same
    # blue divider already used between column groups (border-left, see .ft-group-start).
    body_parts, prev_group = [], None
    for _, row in df.iterrows():
        group = row.get("family_group")
        body_parts.append(body_row(row, prev_group is not None and group != prev_group))
        prev_group = group
    body = "".join(body_parts)
    return (f"<table class='full-table'><thead>{group_header_row()}{column_header_row()}</thead>"
            f"<tbody>{body}</tbody></table>")


def write_html(df: pd.DataFrame, path: Path, campaign_dir: Path, campaign: Campaign) -> None:
    summary_rows = _build_campaign_summary(campaign, campaign_dir)
    summary_html = pd.DataFrame(summary_rows, columns=["Field", "Value", "Details"]).to_html(
        index=False, na_rep="", escape=False)
    parts = [f"<div class='md-card table-card'><h2>Campaign summary</h2>{summary_html}</div>"]

    summary_view_path = campaign_dir / "boltz_summary_view.csv"
    write_summary_csv(df, summary_view_path)
    csv_links = ("<p><a href='boltz_summary.csv'>Download full CSV</a> &middot; "
                 f"<a href='{summary_view_path.name}'>Download summary CSV</a></p>")
    footer = f"<div class='summary-table-footer'>{csv_links}{_build_summary_legend_html()}</div>"
    parts.append(f"<div class='md-card table-card'><h2>Summary table</h2>"
                 f"{_build_full_table_html(df)}{footer}</div>")

    lig_notes = _ligand_chemistry_notes(campaign)
    if lig_notes:
        lig_rows = [{"Ligand": lig_id, "Chemistry notes": "; ".join(info["notes"])} for lig_id, info in lig_notes.items()]
        lig_html = pd.DataFrame(lig_rows).to_html(index=False, na_rep="")
        lig_note_text = (f"{len(lig_notes)} of {len(campaign.ligands)} ligand(s) flagged for chemistry review "
                          "-- these are advisory, not errors; verify the input SMILES reflects what you intended "
                          "before trusting the results below.")
        parts.append(f"<div class='md-card table-card'><h2>Ligand preparation</h2>"
                      f"<p>{lig_note_text}</p>{lig_html}</div>")
    else:
        parts.append("<div class='md-card table-card'><h2>Ligand preparation</h2>"
                      "<p>No stereocentre, protonation-state, or disconnected-fragment concerns detected.</p></div>")

    ligand_grid_html = _build_ligand_grid_panel(campaign, campaign_dir)
    if ligand_grid_html:
        parts.append(ligand_grid_html)

    chart_cards = []
    for col, title, div_id in (("pIC50", "Ranked predicted pIC50", "chart-pic50"),
                               ("confidence_score", "Ranked confidence", "chart-confidence")):
        chart_html = _make_bar_chart(df, col, div_id)
        if chart_html:
            chart_cards.append(f"<div class='md-card'><h2>{title}</h2>{chart_html}</div>")

    heat = _make_selectivity_heatmap(df)
    if heat:
        chart_cards.append(f"<div class='md-card'><h2>Family x ligand selectivity</h2><img src='data:image/png;base64,{heat}'></div>")

    scatter = _make_scatter(df, "chart-scatter")
    if scatter:
        chart_cards.append(f"<div class='md-card'><h2>pIC50 vs confidence score</h2>{scatter}</div>")

    interactions_csv = campaign_dir / "boltz_interactions.csv"
    interactions_df = pd.read_csv(interactions_csv) if interactions_csv.exists() else None

    ichart = _make_interaction_count_chart(df, "chart-interactions")
    if ichart:
        chart_cards.append(f"<div class='md-card'><h2>Interaction counts by type</h2>{ichart}</div>")

    pic50_binder = _make_pic50_vs_binder_chart(df, "chart-pic50-binder")
    if pic50_binder:
        chart_cards.append(f"<div class='md-card'><h2>pIC50 vs binder probability</h2>{pic50_binder}</div>")

    for family_id, chart_html in _make_fingerprint_heatmaps(df, interactions_df):
        chart_cards.append(f"<div class='md-card md-card-span2'><h2>{family_id}: residue interaction fingerprint</h2>"
                            f"{chart_html}</div>")

    if chart_cards:
        parts.append(f"<div class='md-chart-grid'>{''.join(chart_cards)}</div>")

    need_3dmol = False
    if "plip_png_path" in df.columns:
        sessions_dir = campaign_dir / "boltz_dashboard_sessions"
        session_cards, total_bytes = [], 0
        viewer_scripts = []
        for _, row in df.iterrows():
            png_rel = row.get("plip_png_path")
            if not isinstance(png_rel, str):
                continue
            png_path = campaign_dir / png_rel
            if not png_path.exists():
                continue
            target_id = row["target_id"]
            display_name = row.get("display_name", target_id)
            b64 = base64.b64encode(png_path.read_bytes()).decode("ascii")
            img_download = f"boltz_binding_site_{target_id}.png"
            image_links = [f"<a href='data:image/png;base64,{b64}' download='{img_download}'>Download image</a>"]

            pse_rel = row.get("plip_pse_path")
            pse_link_html = None
            if isinstance(pse_rel, str) and (campaign_dir / pse_rel).exists():
                sessions_dir.mkdir(exist_ok=True)
                dest = sessions_dir / f"{target_id}.pse"
                shutil.copy2(campaign_dir / pse_rel, dest)
                total_bytes += dest.stat().st_size
                pse_link_html = f"<a href='boltz_dashboard_sessions/{dest.name}' download>Download PyMOL session</a>"

            # Interactive rotating structure view next to the static PyMOL image, built
            # from the same predicted CIF -- 3Dmol.js parses mmCIF natively (format "cif"),
            # so the raw file goes in directly; no PDB conversion needed (Boltz's own chain
            # names, e.g. a 3-letter family id, are longer than PDB's 1-character chain
            # field allows and would need remapping otherwise).
            # Everything for a grid column must be ONE element -- CSS Grid auto-places every
            # direct child of the grid container into its own cell, so a viewer div followed
            # by a sibling <p> (rather than both nested inside one wrapper) silently shifts
            # every column after it over by one, which is exactly what happened here.
            viewer_col = ""
            cif_rel = row.get("cif_file")
            if isinstance(cif_rel, str):
                cif_path = campaign_dir / "boltz_cif" / cif_rel
                if cif_path.exists():
                    div_id = f"viewer-{re.sub(r'[^a-zA-Z0-9_-]', '_', str(target_id))}"
                    cif_json = json.dumps(cif_path.read_text())
                    need_3dmol = True
                    viewer_scripts.append(f"""
(function() {{
  var el = document.getElementById({json.dumps(div_id)});
  if (!el || typeof $3Dmol === 'undefined') return;
  var viewer = $3Dmol.createViewer(el, {{backgroundColor: 'white'}});
  var model = viewer.addModel({cif_json}, 'cif');
  viewer.setStyle({{}}, {{cartoon: {{color: 'lightgrey'}}}});
  viewer.setStyle({{hetflag: true}}, {{stick: {{colorscheme: 'greenCarbon'}}}});
  var lig = model.selectedAtoms({{hetflag: true}});
  if (lig.length) {{ viewer.zoomTo({{hetflag: true}}); }} else {{ viewer.zoomTo(); }}
  viewer.render();
  viewer.spin('y', 0.5);
}})();""")
                    pse_p = f"<p>{pse_link_html}</p>" if pse_link_html else ""
                    viewer_col = f"<div class='md-side-viewer'><div class='md-3dmol-viewer' id='{div_id}'></div>{pse_p}</div>"

            if not viewer_col and pse_link_html:
                image_links.append(pse_link_html)
            image_col = (f"<div class='md-side-image'><img src='data:image/png;base64,{b64}'>"
                         f"<p>{' &middot; '.join(image_links)}</p></div>")

            contacts_table = "<p><em>No interaction data.</em></p>"
            contacts_csv_link = ""
            if interactions_df is not None:
                tdf = interactions_df[interactions_df["target_id"] == target_id]
                rename_map = {"interaction_type": "Interaction", "prot_restype": "Residue",
                              "prot_resnr": "Number", "prot_chain": "Chain", "distance_A": "Distance"}
                show_cols = [c for c in rename_map if c in tdf.columns]
                if not tdf.empty and show_cols:
                    contacts_df = tdf[show_cols].sort_values("interaction_type").rename(columns=rename_map)
                    contacts_table = contacts_df.to_html(index=False, na_rep="")
                    csv_b64 = base64.b64encode(contacts_df.to_csv(index=False).encode("utf-8")).decode("ascii")
                    contacts_csv_link = (f"<p><a href='data:text/csv;base64,{csv_b64}' "
                                         f"download='boltz_contacts_{target_id}.csv'>Download CSV</a></p>")
            layout_cls = "md-side-by-side md-side-3col" if viewer_col else "md-side-by-side"
            session_cards.append(
                f"<div class='md-card'><h2>{display_name}: binding site</h2>"
                f"<div class='{layout_cls}'>"
                f"{viewer_col}"
                f"{image_col}"
                f"<div class='md-side-table-col'><div class='md-side-table'>{contacts_table}</div>{contacts_csv_link}</div>"
                f"</div></div>"
            )
        if session_cards:
            if total_bytes:
                _info(f"bundled {total_bytes / 1e6:.1f}MB of PyMOL session(s) into "
                      f"{sessions_dir} (this is why the dashboard is no longer a single file)")
            parts.extend(session_cards)
            if viewer_scripts:
                parts.append(f"<script>{''.join(viewer_scripts)}</script>")

    sse_csv = campaign_dir / "boltz_sse_comparison.csv"
    sse_status_path = campaign_dir / "boltz_sse_family_status.json"
    if sse_csv.exists() or sse_status_path.exists():
        # Core, not optional: `analyze`/`all` run compare-sse automatically (see main()),
        # so this card reflects whatever it found -- including families with no 'Apo
        # structure:' configured at all, reported rather than silently omitted. Reads
        # from disk (rather than taking a df/status param) so a standalone `compare-sse`
        # re-run followed by a plain `analyze` still picks up the latest result, matching
        # how the PLIP interaction cards above already work.
        from sse_comparison.report import (_make_sse_heatmap, _make_sse_shift_chart,
                                            build_family_status_html, build_sse_table_html,
                                            build_summary_stats_html, compute_summary_stats)
        sse_df = pd.read_csv(sse_csv) if sse_csv.exists() else pd.DataFrame()
        family_status = json.loads(sse_status_path.read_text()) if sse_status_path.exists() else {}

        sse_intro = build_family_status_html(family_status) + build_summary_stats_html(compute_summary_stats(sse_df))
        card_html = f"<div class='md-card table-card'><h2>Secondary structure shifts (apo vs holo)</h2>{sse_intro}"
        if not sse_df.empty:
            card_html += (f"{build_sse_table_html(sse_df)}"
                          f"<p><a href='boltz_sse_comparison.csv'>Download CSV</a></p>")
        card_html += "</div>"
        parts.append(card_html)

        if not sse_df.empty:
            sse_charts = []
            sse_bar = _make_sse_shift_chart(sse_df, "chart-sse-shift")
            if sse_bar:
                sse_charts.append(f"<div class='md-card'><h2>Per-motif Ca RMSD</h2>{sse_bar}</div>")
            sse_heat = _make_sse_heatmap(sse_df, "chart-sse-heatmap")
            if sse_heat:
                sse_charts.append(f"<div class='md-card'><h2>Motif x target RMSD</h2>{sse_heat}</div>")
            if sse_charts:
                parts.append(f"<div class='md-chart-grid'>{''.join(sse_charts)}</div>")

    if PLOTLY_JS_PATH.exists():
        plotly_script = f"<script>{PLOTLY_JS_PATH.read_text()}</script>"
    else:
        # Fresh checkouts always have vendor/plotly-2.35.2.min.js committed; this is only
        # a safety net (e.g. a shallow/sparse clone), and it reintroduces the exact
        # htmlpreview-breaking failure mode the vendored copy exists to avoid.
        _warn("vendor/plotly-2.35.2.min.js not found -- falling back to "
              "the plotly.js CDN, which is known not to render in some HTML-preview contexts")
        plotly_script = "<script src='https://cdn.plot.ly/plotly-2.35.2.min.js'></script>"

    threedmol_script = ""
    if need_3dmol:
        if THREEDMOL_JS_PATH.exists():
            threedmol_script = f"<script>{THREEDMOL_JS_PATH.read_text()}</script>"
        else:
            _warn("vendor/3Dmol-2.5.5-min.js not found -- falling back to the 3Dmol.js CDN")
            threedmol_script = "<script src='https://cdn.jsdelivr.net/npm/3dmol@2.5.5/build/3Dmol-min.js'></script>"

    # Posts this page's real content height to any parent window embedding it in an
    # iframe (e.g. findings.md's "Interactive dashboard" section) -- postMessage works
    # across origins even though direct DOM/height reads on a cross-origin iframe don't,
    # so the embedding page can size the iframe to the actual content instead of guessing
    # a fixed height and either clipping content or leaving a scrollbar. A ResizeObserver
    # on <body> (not just a couple of fixed-delay retries after load) is what actually
    # keeps this correct: this page's own embedded PLIP images, web fonts, and the
    # ligand-grid pager can all reflow content well after the load event fires, and a
    # fixed timeout window doesn't reliably outlast every one of those -- confirmed
    # directly (a fixed-timeout version still left a residual scrollbar). Falls back to
    # a bounded polling interval on browsers without ResizeObserver.
    _IFRAME_RESIZE_JS = """
(function () {
  function postHeight() {
    if (window.parent && window.parent !== window) {
      window.parent.postMessage({source: 'boltzmaker-dashboard', height: document.documentElement.scrollHeight}, '*');
    }
  }
  window.addEventListener('load', postHeight);
  window.addEventListener('resize', postHeight);
  if (window.ResizeObserver) {
    new ResizeObserver(postHeight).observe(document.body);
  } else {
    var ticks = 0;
    var poll = setInterval(function () {
      postHeight();
      if (++ticks > 30) clearInterval(poll);
    }, 300);
  }
})();
"""
    doc = (
        "<!DOCTYPE html><html lang='en'><head><meta charset='utf-8'>"
        "<meta name='viewport' content='width=device-width, initial-scale=1.0'>"
        "<title>BoltzMaker Report | Marc C. Deller</title>"
        "<link href='https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700"
        "&family=Roboto+Mono:wght@400;500&display=swap' rel='stylesheet'>"
        + plotly_script + threedmol_script
        + f"<style>{_BRAND_CSS}</style></head><body>"
        + _BRAND_HEADER + "<main class='md-main'>" + "".join(parts) + "</main>" + _BRAND_FOOTER
        + f"<script>{_LIGAND_GRID_PAGER_JS}</script>"
        + f"<script>{_IFRAME_RESIZE_JS}</script>"
        + "</body></html>"
    )
    path.write_text(doc)


# ==========================================================================
# CLI
# ==========================================================================

def _build_argparser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(prog="BoltzMaker.py", description="Manage Boltz-2 batch campaigns end-to-end.")
    sub = p.add_subparsers(dest="command")

    fmt = sub.add_parser("format", help="auto-align comments/blank-lines in a boltz_input.md (cosmetic only)")
    fmt.add_argument("md_path", type=Path, help="path to boltz_input.md")
    fmt.add_argument("--check", action="store_true", help="report whether reformatting is needed; exit 1 if so, without writing")

    new = sub.add_parser("new", help="interactively write a new boltz_input.md by answering plain questions")
    new.add_argument("md_path", type=Path, nargs="?", default=Path("boltz_input.md"), help="output path (default boltz_input.md)")

    cs = sub.add_parser("compare-sse", help="compare secondary-structure motif shifts between a "
                         "family's apo reference structure and its predicted holo target(s)")
    cs.add_argument("md_path", type=Path, help="path to boltz_input.md")
    cs.add_argument("--family", type=str, default=None, help="restrict to one Protein family id "
                     "(default: every family with an 'Apo structure:' set)")
    cs.add_argument("--target", type=str, default=None, help="restrict to one target stem "
                     "(default: every target for the selected family)")
    cs.add_argument("--out-dir", type=Path, default=None, help="default: alongside boltz_input.md")
    cs.add_argument("--phi-psi-threshold", type=float, default=30.0, help="degrees; per-residue "
                     "phi/psi delta above this is flagged (default 30)")
    cs.add_argument("--dfg-distance-threshold", type=float, default=8.0, help="angstroms; DFG-Asp to "
                     "catalytic-Lys Ca-Ca distance below this is classified DFG-in (default 8.0)")
    cs.add_argument("--alphac-distance-threshold", type=float, default=10.0, help="angstroms; "
                     "alphaC-Glu to catalytic-Lys Ca-Ca distance below this is classified alphaC-in (default 10.0)")
    cs.add_argument("--no-pymol", action="store_true", help="skip writing .pml session scripts")
    cs.add_argument("--refresh-cache", action="store_true", help="bypass the GPCRdb/KLIFS/PDBe "
                     "disk cache for this run")

    for name in ("generate", "preflight", "run", "analyze", "all"):
        sp = sub.add_parser(name)
        sp.add_argument("md_path", type=Path, help="path to boltz_input.md")
        sp.add_argument("--output-dir", type=Path, default=None, help="override settings.output_dir")
        sp.add_argument("--out-dir", type=Path, default=None, help="boltz predict --out_dir (default ./boltz_output beside the md)")
        sp.add_argument("--workers", type=int, default=0, help="dataloader workers. Boltz's own default "
                        "is 2, but each worker duplicates large in-memory structures, and on unified-memory "
                        "hardware that is paid for out of the same pool the model is using -- 0 is what a "
                        "26-target GPCR campaign (~1300 tokens/target) actually needed on a 64GB M1 Max")
        sp.add_argument("--accelerator", choices=["auto", "gpu", "cpu"], default="auto")
        sp.add_argument("--limit", type=int, default=None, help="cap how many pending targets `run` submits")
        sp.add_argument("--max-retries", type=int, default=2, help="if a target doesn't complete (e.g. an "
                        "OOM crash), automatically retry it up to this many times, in isolation one target "
                        "at a time -- lets a multi-hour campaign recover unattended instead of stopping on "
                        "a transient memory failure (0 to disable and match the old fail-once behavior)")
        sp.add_argument("--strict", action="store_true", help="promote preflight WARN to FAIL")
        sp.add_argument("-y", "--yes", action="store_true")
        sp.add_argument("--mps-watermark", type=float, default=1.0, help="PYTORCH_MPS_HIGH_WATERMARK_RATIO -- "
                        "caps MPS memory at this x the device's recommended max, so an oversized complex fails "
                        "fast with a clear OOM instead of swap-thrashing (default 1.0; set higher to allow more "
                        "overcommit, 0 to disable the cap entirely). This is a HARD allocation ceiling, not a "
                        "swap-avoidance dial: 0.7 on a 64GB M1 Max caps allocation at 36GB against a ~34GB "
                        "requirement and every batch OOMs immediately. Lower it only above a measured peak")
        sp.add_argument("--max-parallel-samples", type=int, default=1, help="boltz --max_parallel_samples "
                        "(default 1 here for Mac memory safety; Boltz's own default is unbounded)")
        sp.add_argument("--recycling-steps", type=int, default=None, help="boltz --recycling_steps passthrough "
                        "(default: Boltz's own default of 3)")
        sp.add_argument("--sampling-steps", type=int, default=None, help="boltz --sampling_steps passthrough "
                        "(default: Boltz's own default of 200)")
        sp.add_argument("--diffusion-samples", type=int, default=1, help="boltz --diffusion_samples "
                        "passthrough -- how many independent structure samples to predict per target "
                        "(default 1; each additional sample costs roughly proportional diffusion time). "
                        "Boltz writes {stem}_model_0.cif .. {stem}_model_{N-1}.cif; BoltzMaker's own "
                        "analyze/compare-sse only ever look at model_0)")
        sp.add_argument("--diffusion-samples-affinity", type=int, default=None, help="boltz "
                        "--diffusion_samples_affinity passthrough (default: Boltz's own default of 5)")
        sp.add_argument("--sampling-steps-affinity", type=int, default=None, help="boltz "
                        "--sampling_steps_affinity passthrough (default: Boltz's own default of 200)")
        sp.add_argument("--max-msa-seqs", type=int, default=4096, help="boltz --max_msa_seqs passthrough. "
                        "Boltz's own default is 8192; 4096 halves the co-evolution feature block, which is one "
                        "of the few levers that measurably cuts peak memory on large complexes. Pass 8192 to "
                        "restore Boltz's default if you have the headroom")
        sp.add_argument("--no-potentials", dest="use_potentials", action="store_false", default=True,
                        help="disable boltz --use_potentials (FK steering and the physical-guidance "
                             "coordinate update). On by default, matching Boltz's recommended setting; "
                             "turning it off is worth trying when a target's diffusion diverges to NaN.")
        sp.add_argument("--memory-warn-tokens", type=int, default=1500, help="preflight WARNs if a target's "
                        "total residue/atom count exceeds this (empirical heuristic, see preflight check)")
        sp.add_argument("--skip-interactions", action="store_true", help="skip cif2plip protein-ligand "
                        "interaction analysis during `analyze`, even if `setup-plip` has been run")
        sp.add_argument("--skip-sse", action="store_true", help="skip compare-sse apo-vs-holo secondary-"
                        "structure analysis during `analyze`, even if a family has 'Apo structure:' set")
        if name == "preflight":
            sp.add_argument("--json", action="store_true",
                             help="emit check results as JSON instead of a rich table (for scripting/tooling)")
    return p


def main() -> None:
    argv = sys.argv[1:]
    if not argv:
        _info(f"{_BOLD}usage:{_RESET}{_BLUE} BoltzMaker.py {_RESET}"
              f"{_DIM}[setup|setup-plip|new|format|compare-sse|generate|preflight|run|analyze|all]{_RESET} "
              f"<boltz_input.md> [options]")
        sys.exit(1)

    known = {"format", "new", "compare-sse", "generate", "preflight", "run", "analyze", "all"}
    if argv[0] not in known:
        argv = ["all"] + argv
    args = _build_argparser().parse_args(argv)

    if args.command == "format":
        cmd_format(args.md_path.resolve(), check=args.check)
        return

    if args.command == "new":
        cmd_new(args.md_path.resolve())
        return

    md_path = args.md_path.resolve()
    campaign_dir = md_path.parent
    campaign = parse_md(md_path)

    if args.command == "compare-sse":
        from sse_comparison.cli import run_compare_sse
        run_compare_sse(campaign, campaign_dir, family_id=args.family, target_stem=args.target,
                         out_dir=args.out_dir or campaign_dir, phi_psi_threshold=args.phi_psi_threshold,
                         dfg_distance_threshold=args.dfg_distance_threshold,
                         alphac_distance_threshold=args.alphac_distance_threshold,
                         render_pymol=not args.no_pymol, refresh_cache=args.refresh_cache)
        return

    output_dir = args.output_dir if args.output_dir else Path(campaign.settings.output_dir)
    if not output_dir.is_absolute():
        output_dir = (campaign_dir / output_dir).resolve()
    out_dir = args.out_dir if args.out_dir else (campaign_dir / "boltz_output")

    if args.command in ("generate", "all"):
        manifest = generate_yamls(campaign, output_dir)
        _ok(f"generated {len(manifest)} target YAML(s) in {output_dir}")
    else:
        manifest = load_manifest(output_dir)

    if args.command in ("preflight", "all"):
        ok = run_preflight(manifest, output_dir, campaign, md_path, strict=args.strict,
                            memory_warn_tokens=args.memory_warn_tokens,
                            json_output=getattr(args, "json", False))
        if not ok and args.command == "all":
            _err("preflight failed, aborting before run.")
            sys.exit(1)

    if args.command in ("run", "all"):
        accelerator = resolve_accelerator(args.accelerator)
        run_boltz(output_dir, out_dir, manifest, args.workers, accelerator, campaign_dir,
                  limit=args.limit, mps_watermark=args.mps_watermark, max_parallel_samples=args.max_parallel_samples,
                  recycling_steps=args.recycling_steps, sampling_steps=args.sampling_steps,
                  diffusion_samples=args.diffusion_samples,
                  diffusion_samples_affinity=args.diffusion_samples_affinity,
                  sampling_steps_affinity=args.sampling_steps_affinity, max_msa_seqs=args.max_msa_seqs,
                  max_retries=args.max_retries, use_potentials=args.use_potentials,
                  targets_per_invocation=campaign.settings.targets_per_invocation)

    if args.command in ("analyze", "all"):
        df = analyze(output_dir, out_dir, campaign_dir, campaign,
                     skip_interactions=args.skip_interactions)
        write_csv(df, campaign_dir / "boltz_summary.csv")
        write_xlsx(df, campaign_dir / "boltz_summary.xlsx")
        if not args.skip_sse:
            # Core, not optional: every family always gets a status (even "no Apo
            # structure: configured"), written to boltz_sse_family_status.json and
            # picked up by write_html below -- never aborts analyze over this.
            from sse_comparison.cli import run_compare_sse
            run_compare_sse(campaign, campaign_dir, strict=False)
        write_html(df, campaign_dir / "boltz_dashboard.html", campaign_dir, campaign)
        _ok(f"analysis written to {campaign_dir} "
            "(boltz_summary.csv / .xlsx / boltz_summary_view.csv / boltz_dashboard.html)")


if __name__ == "__main__":
    main()
